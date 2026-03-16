import os
import sys
import json
import uuid
# Debug timestamp: 2026-01-27 14:33 - Force reload
import smtplib
import secrets
import base64
import time
from datetime import datetime, timedelta, timezone
from flask import Flask, render_template, request, redirect, url_for, jsonify, send_from_directory, send_file, flash
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from flask_wtf.csrf import CSRFProtect
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

# הגדרת encoding ל-UTF-8 ל-Windows console (למניעת שגיאות emoji)
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except:
        pass  # אם זה לא עובד, נתעלם מזה

# טעינת משתני סביבה מקובץ .env (אם קיים)
load_dotenv()

# Use database helpers if USE_DATABASE is enabled (for Railway deployment)
USE_DATABASE = os.environ.get('USE_DATABASE', 'false').lower() == 'true'
if USE_DATABASE:
    # Import database helpers to override JSON functions
    from database_helpers import (
        load_users, save_users, load_data, save_data,
        load_suppliers, save_suppliers, load_quotes, save_quotes,
        load_messages, save_messages, load_events, save_events,
        load_equipment_bank, save_equipment_bank,
        load_checklist_templates, save_checklist_templates,
        load_forms, save_forms, delete_user_record,
        load_time_tracking, save_time_tracking
    )

# Import notifications module
from backend.utils.notifications import create_notification
from backend.utils.email import send_charge_notification_email

app = Flask(__name__)
# SECRET_KEY מ-environment variable (חובה בפרודקשן!)
app.secret_key = os.environ.get('SECRET_KEY') or 'vatkin_master_final_v100_CHANGE_IN_PRODUCTION'
# הגדרת timeout ל-session - 24 שעות
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)
# הגדרות אבטחה ל-session cookies
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
# Railway uses HTTPS, so we need secure cookies
is_production = os.environ.get('RAILWAY_ENVIRONMENT') or os.environ.get('PORT')
if is_production:
    app.config['SESSION_COOKIE_SECURE'] = True

# הגדרת נתיבים ותיקיות
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATIC_FOLDER = os.path.join(BASE_DIR, 'static')
LOGOS_FOLDER = os.path.join(BASE_DIR, 'static', 'logos')
DOCUMENTS_FOLDER = os.path.join(BASE_DIR, 'static', 'documents')
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'static', 'client_docs')
DATA_FILE = os.path.join(BASE_DIR, 'agency_db.json')
USERS_FILE = os.path.join(BASE_DIR, 'users_db.json')
SUPPLIERS_FILE = os.path.join(BASE_DIR, 'suppliers_db.json')
QUOTES_FILE = os.path.join(BASE_DIR, 'quotes_db.json')
MESSAGES_FILE = os.path.join(BASE_DIR, 'messages_db.json')
EVENTS_FILE = os.path.join(BASE_DIR, 'events_db.json')
EQUIPMENT_BANK_FILE = os.path.join(BASE_DIR, 'equipment_bank.json')
CHECKLIST_TEMPLATES_FILE = os.path.join(BASE_DIR, 'checklist_templates.json')
FORMS_FILE = os.path.join(BASE_DIR, 'forms_db.json')
FORMS_UPLOAD_FOLDER = os.path.join(BASE_DIR, 'static', 'forms_uploads')
CHAT_FILES_FOLDER = os.path.join(BASE_DIR, 'static', 'chat_files')
SUPPLIER_FILES_FOLDER = os.path.join(BASE_DIR, 'static', 'supplier_files')
PERMISSIONS_FILE = os.path.join(BASE_DIR, 'permissions_db.json')
USER_ACTIVITY_FILE = os.path.join(BASE_DIR, 'user_activity.json')
ACTIVITY_LOGS_FILE = os.path.join(BASE_DIR, 'activity_logs.json')
TIME_TRACKING_FILE = os.path.join(BASE_DIR, 'time_tracking.json')
# הגדרת תיקיית העלאות
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# יצירת תיקיות אם לא קיימות
if not os.path.exists(STATIC_FOLDER):
    os.makedirs(STATIC_FOLDER, exist_ok=True)
if not os.path.exists(LOGOS_FOLDER): 
    os.makedirs(LOGOS_FOLDER, exist_ok=True)
if not os.path.exists(DOCUMENTS_FOLDER):
    os.makedirs(DOCUMENTS_FOLDER, exist_ok=True)
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
if not os.path.exists(CHAT_FILES_FOLDER):
    os.makedirs(CHAT_FILES_FOLDER, exist_ok=True)
if not os.path.exists(FORMS_UPLOAD_FOLDER):
    os.makedirs(FORMS_UPLOAD_FOLDER, exist_ok=True)
if not os.path.exists(SUPPLIER_FILES_FOLDER):
    os.makedirs(SUPPLIER_FILES_FOLDER, exist_ok=True)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# הוספת CSRF Protection
csrf = CSRFProtect(app)

# הוספת Rate Limiting
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://"  # ניתן לשנות ל-Redis בפרודקשן
)

# Only define JSON-based functions if NOT using database
if not USE_DATABASE:
    def load_users():
        if not os.path.exists(USERS_FILE):
            # יצירת סיסמה מוצפנת למנהל ברירת מחדל
            u = {'admin': {'password': generate_password_hash('1234'), 'name': 'מנהל המשרד', 'role': 'אדמין'}}
            with open(USERS_FILE, 'w', encoding='utf-8') as f: json.dump(u, f, ensure_ascii=False, indent=4)
            return u
        users = {}
        with open(USERS_FILE, 'r', encoding='utf-8') as f: 
            users = json.load(f)
        # וידוא שלכל משתמש יש role (ברירת מחדל: עובד)
        needs_update = False
        for uid, user_info in users.items():
            if 'role' not in user_info:
                user_info['role'] = 'עובד' if uid != 'admin' else 'אדמין'
                needs_update = True
        if needs_update:
            save_users(users)
        return users

    def save_users(users):
        with open(USERS_FILE, 'w', encoding='utf-8') as f: 
            json.dump(users, f, ensure_ascii=False, indent=4)

if not USE_DATABASE:
    def load_data():
        if not os.path.exists(DATA_FILE) or os.stat(DATA_FILE).st_size == 0: return []
        with open(DATA_FILE, 'r', encoding='utf-8') as f: 
            data = json.load(f)
            # אם זה רשימה ריקה, החזר
            if not data:
                return data
            # וידוא שלכל לקוח יש client_number
            needs_update = False
            for client in data:
                if 'client_number' not in client:
                    needs_update = True
                    break
            # אם צריך עדכון, עדכן את כל הלקוחות
            if needs_update:
                assign_client_numbers(data)
            return data

def assign_client_numbers(clients):
    """מקצה מספרים ייחודיים ללקוחות שאין להם"""
    # מצא את המספר הגבוה ביותר הקיים
    max_number = 0
    for client in clients:
        if 'client_number' in client:
            try:
                num = int(client['client_number'])
                if num > max_number:
                    max_number = num
            except:
                pass
    
    # הקצה מספרים ללקוחות שאין להם
    current_number = max_number + 1
    for client in clients:
        if 'client_number' not in client or not client.get('client_number'):
            client['client_number'] = current_number
            current_number += 1
    
    # שמור את העדכון
    save_data(clients)

def get_next_client_number():
    """מחזיר את המספר הבא ללקוח חדש"""
    data = load_data()
    if not data:
        return 1
    
    max_number = 0
    for client in data:
        if 'client_number' in client:
            try:
                num = int(client['client_number'])
                if num > max_number:
                    max_number = num
            except:
                pass
    
    return max_number + 1

def get_next_project_number(client):
    """מחזיר את מספר הפרויקט הבא ללקוח מסוים
    פורמט: 7 ספרות - 3 ספרות לקוח + 4 ספרות פרויקט
    דוגמה: 0010001 (לקוח 1, פרויקט 1)"""
    client_number = client.get('client_number', 1)
    try:
        client_num = int(client_number)
    except (ValueError, TypeError):
        client_num = 1
    
    # מצא את המספר הגבוה ביותר של פרויקטים ללקוח הזה
    max_project_seq = 0
    for project in client.get('projects', []):
        if 'project_number' in project:
            try:
                # קח את 4 הספרות האחרונות (מספר הפרויקט)
                proj_num_str = str(project['project_number'])
                if len(proj_num_str) >= 4:
                    project_seq = int(proj_num_str[-4:])
                    if project_seq > max_project_seq:
                        max_project_seq = project_seq
            except (ValueError, TypeError):
                pass
    
    next_seq = max_project_seq + 1
    # הרכב: 3 ספרות לקוח + 4 ספרות פרויקט
    project_number = f"{client_num:03d}{next_seq:04d}"
    return project_number

def get_next_workday(from_date=None):
    """מחזיר את יום העבודה הקרוב (א'-ה')
    בישראל: ראשון=0, שני=1, שלישי=2, רביעי=3, חמישי=4, שישי=5, שבת=6
    ימי עבודה: ראשון עד חמישי (0-4)"""
    if from_date is None:
        from_date = datetime.now()
    
    # אם היום יום עבודה (א'-ה'), החזר היום
    # weekday() מחזיר: Monday=0, Sunday=6
    # אנחנו צריכים להמיר לשבוע ישראלי
    day_of_week = from_date.weekday()  # 0=Monday, 6=Sunday
    
    # המרה לשבוע ישראלי: Sunday=0, Monday=1, ..., Saturday=6
    israeli_day = (day_of_week + 1) % 7  # Sunday=0, Monday=1, ..., Saturday=6
    
    # ימי עבודה בישראל: ראשון(0) עד חמישי(4)
    if israeli_day <= 4:  # ראשון עד חמישי
        return from_date.replace(hour=0, minute=0, second=0, microsecond=0)
    elif israeli_day == 5:  # שישי - קפוץ לראשון
        return (from_date + timedelta(days=2)).replace(hour=0, minute=0, second=0, microsecond=0)
    else:  # שבת - קפוץ לראשון
        return (from_date + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)

def get_next_task_number(client, project):
    """מחזיר את מספר המשימה הבאה בפרויקט מסוים
    פורמט: 10 ספרות - 7 ספרות פרויקט + 3 ספרות משימה
    דוגמה: 0010001001 (לקוח 1, פרויקט 1, משימה 1)"""
    project_number = project.get('project_number', '')
    if not project_number:
        # אם אין מספר פרויקט, צור אחד חדש
        project_number = get_next_project_number(client)
        project['project_number'] = project_number
    
    # מצא את המספר הגבוה ביותר של משימות בפרויקט הזה
    max_task_seq = 0
    for task in project.get('tasks', []):
        if 'task_number' in task:
            try:
                # קח את 3 הספרות האחרונות (מספר המשימה)
                task_num_str = str(task['task_number'])
                if len(task_num_str) >= 3:
                    task_seq = int(task_num_str[-3:])
                    if task_seq > max_task_seq:
                        max_task_seq = task_seq
            except (ValueError, TypeError):
                pass
    
    next_seq = max_task_seq + 1
    # הרכב: 7 ספרות פרויקט + 3 ספרות משימה
    task_number = f"{project_number}{next_seq:03d}"
    return task_number

def get_next_charge_number(client):
    """מחזיר את מספר החיוב הבא ללקוח מסוים
    פורמט: 7 ספרות - 3 ספרות לקוח + 4 ספרות חיוב
    דוגמה: 0010001 (לקוח 1, חיוב 1)"""
    client_number = client.get('client_number', 1)
    try:
        client_num = int(client_number)
    except (ValueError, TypeError):
        client_num = 1
    
    # מצא את המספר הגבוה ביותר של חיובים ללקוח הזה
    max_charge_seq = 0
    for charge in client.get('extra_charges', []):
        if 'charge_number' in charge:
            try:
                # קח את 4 הספרות האחרונות (מספר החיוב)
                charge_num_str = str(charge['charge_number'])
                if len(charge_num_str) >= 4:
                    charge_seq = int(charge_num_str[-4:])
                    if charge_seq > max_charge_seq:
                        max_charge_seq = charge_seq
            except (ValueError, TypeError):
                pass
    
    next_seq = max_charge_seq + 1
    # הרכב: 3 ספרות לקוח + 4 ספרות חיוב
    charge_number = f"{client_num:03d}{next_seq:04d}"
    return charge_number

if not USE_DATABASE:
    def save_data(data):
        with open(DATA_FILE, 'w', encoding='utf-8') as f: json.dump(data, f, ensure_ascii=False, indent=4)

    def load_suppliers():
        if not os.path.exists(SUPPLIERS_FILE) or os.stat(SUPPLIERS_FILE).st_size == 0: return []
        with open(SUPPLIERS_FILE, 'r', encoding='utf-8') as f: return json.load(f)

    def save_suppliers(suppliers):
        with open(SUPPLIERS_FILE, 'w', encoding='utf-8') as f: json.dump(suppliers, f, ensure_ascii=False, indent=4)

    def load_quotes():
        if not os.path.exists(QUOTES_FILE) or os.stat(QUOTES_FILE).st_size == 0: return []
        with open(QUOTES_FILE, 'r', encoding='utf-8') as f: return json.load(f)

    def save_quotes(quotes):
        with open(QUOTES_FILE, 'w', encoding='utf-8') as f: json.dump(quotes, f, ensure_ascii=False, indent=4)

    def load_messages():    
        if not os.path.exists(MESSAGES_FILE) or os.stat(MESSAGES_FILE).st_size == 0: return []
        with open(MESSAGES_FILE, 'r', encoding='utf-8') as f: return json.load(f)

    def save_messages(messages):
        with open(MESSAGES_FILE, 'w', encoding='utf-8') as f: json.dump(messages, f, ensure_ascii=False, indent=4)

    def load_events():
        if not os.path.exists(EVENTS_FILE) or os.stat(EVENTS_FILE).st_size == 0: return []
        with open(EVENTS_FILE, 'r', encoding='utf-8') as f: return json.load(f)

    def save_events(events):
        with open(EVENTS_FILE, 'w', encoding='utf-8') as f: json.dump(events, f, ensure_ascii=False, indent=4)

if not USE_DATABASE:
    def load_time_tracking():
        """טעינת מדידות זמן מקובץ JSON"""
        if not os.path.exists(TIME_TRACKING_FILE) or os.stat(TIME_TRACKING_FILE).st_size == 0:
            return {'entries': [], 'active_sessions': {}}
        with open(TIME_TRACKING_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)

    def save_time_tracking(data):
        """שמירת מדידות זמן לקובץ JSON"""
        with open(TIME_TRACKING_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)

    def load_equipment_bank():
        if not os.path.exists(EQUIPMENT_BANK_FILE) or os.stat(EQUIPMENT_BANK_FILE).st_size == 0: 
            # יצירת מאגר ציוד בסיסי
            default_equipment = [
                'מקרן', 'הגברה', 'מיקרופון', 'מסך/פליפ-צ\'ארט', 'שולחן', 'כסאות', 
                'מתנות לאורחים', 'רול-אפים', 'באנרים', 'פלטה/במה', 'תאורה', 'שולחן עגול'
            ]
            save_equipment_bank(default_equipment)
            return default_equipment
        with open(EQUIPMENT_BANK_FILE, 'r', encoding='utf-8') as f: return json.load(f)

    def save_equipment_bank(equipment):
        with open(EQUIPMENT_BANK_FILE, 'w', encoding='utf-8') as f: json.dump(equipment, f, ensure_ascii=False, indent=4)

    def load_checklist_templates():
        """טעינת תבניות צ'ק-ליסט מקובץ JSON"""
        if not os.path.exists(CHECKLIST_TEMPLATES_FILE) or os.stat(CHECKLIST_TEMPLATES_FILE).st_size == 0:
            # יצירת תבניות ברירת מחדל
            default_templates = {
                'כנס': [
                    'הזמנת קייטרינג',
                    'עיצוב רול-אפים',
                    'שליחת Save the date',
                    'הזמנת הגברה ותאורה',
                    'הזמנת מקרן ומסך',
                    'הזמנת מקומות ישיבה',
                    'אישור מיקום',
                    'הזמנת צלמים/וידאו',
                    'הכנת מצגות',
                    'הזמנת מתנות למשתתפים'
                ],
                'חתונה': [
                    'אישור אולם',
                    'הזמנת קייטרינג',
                    'הזמנת הגברה ודי.ג\'יי',
                    'הזמנת צלמים/וידאו',
                    'הזמנת פרחים ועיצוב',
                    'הזמנת בוקונז\'ה/מתנות לאורחים',
                    'הזמנת שולחנות וכסאות',
                    'הזמנת מתנות לחתן וכלה',
                    'אישור תאריכים עם כל הספקים',
                    'שליחת הזמנות'
                ],
                'השקה': [
                    'אישור מיקום',
                    'הזמנת קייטרינג/קפה',
                    'עיצוב חומרי שיווק',
                    'הזמנת הגברה',
                    'הזמנת צלמים/וידאו',
                    'שליחת הזמנות',
                    'הכנת מצגת/סרטון',
                    'הזמנת מתנות למשתתפים',
                    'הזמנת פרחים/עיצוב',
                    'אישור תאריכים'
                ]
            }
            save_checklist_templates(default_templates)
            return default_templates
        with open(CHECKLIST_TEMPLATES_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)

    def save_checklist_templates(templates):
        """שמירת תבניות צ'ק-ליסט לקובץ JSON"""
        with open(CHECKLIST_TEMPLATES_FILE, 'w', encoding='utf-8') as f:
            json.dump(templates, f, ensure_ascii=False, indent=4)

    def load_forms():
        """טעינת טפסים מקובץ JSON"""
        if not os.path.exists(FORMS_FILE) or os.stat(FORMS_FILE).st_size == 0:
            return []
        with open(FORMS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)

    def save_forms(forms):
        """שמירת טפסים לקובץ JSON"""
        with open(FORMS_FILE, 'w', encoding='utf-8') as f:
            json.dump(forms, f, ensure_ascii=False, indent=4)

def send_form_email(form_title, client_name, form_submission, uploaded_files, form_token):
    """
    שליחת מייל עם פרטי הטופס
    המייל יישלח מהמייל של המשתמש המשויך ללקוח (או אלנתן אם לא נמצא)
    מחזיר True אם הצליח, False אחרת
    """
    print("\n" + "="*80)
    print("[EMAIL] ========== Starting Email Send Process ==========")
    print(f"[EMAIL] Form: {form_title}")
    print(f"[EMAIL] Client: {client_name}")
    print("="*80 + "\n")
    
    try:
        # חיפוש המשתמש המשויך ללקוח (או אלנתן כברירת מחדל)
        forms_list = load_forms()
        form = next((f for f in forms_list if f.get('token') == form_token), None)
        
        if not form:
            print("[ERROR] Form not found!")
            return False
        
        client_id = form.get('client_id')
        users = load_users()
        data = load_data()
        
        # חיפוש משתמש משויך ללקוח
        sender_user_id = None
        sender_email = None
        sender_password = None
        
        if client_id:
            for client in data:
                if client.get('id') == client_id:
                    assigned_users = client.get('assigned_user', [])
                    if isinstance(assigned_users, str):
                        assigned_users = [assigned_users]
                    
                    # נחפש משתמש משויך עם מייל וסיסמת מייל
                    for user_id in assigned_users:
                        if user_id in users:
                            user_email = users[user_id].get('email', '')
                            user_email_password_encoded = users[user_id].get('email_password', '')
                            if user_email and user_email_password_encoded:
                                sender_user_id = user_id
                                sender_email = user_email
                                sender_password_encoded = user_email_password_encoded
                                sender_password = base64.b64decode(sender_password_encoded).decode('utf-8')
                                print(f"[EMAIL] Found assigned user: {user_id} ({sender_email})")
                                break
                    break
        
        # אם לא נמצא משתמש משויך, נשתמש באלנתן (נחפש גם "elnatan" וגם "אלנתן")
        if not sender_email:
            # נחפש משתמש עם email של ELNATAN@VATKIN.CO.IL
            elnatan_user_id = None
            for uid, user_data in users.items():
                user_email = user_data.get('email', '').upper()
                if user_email == 'ELNATAN@VATKIN.CO.IL':
                    user_email_password_encoded = user_data.get('email_password', '')
                    if user_email_password_encoded:
                        elnatan_user_id = uid
                        sender_email = user_data.get('email', '')
                        sender_password = base64.b64decode(user_email_password_encoded).decode('utf-8')
                        sender_user_id = uid
                        print(f"[EMAIL] Using default sender: {uid} ({sender_email})")
                        break
            
            if not sender_email:
                print("[ERROR] לא נמצא משתמש משויך ואלנתן לא מוגדר עם מייל וסיסמת מייל!")
                return False
        
        # קביעת כתובת המייל - נשלח ל-studio@vatkin.co.il + משתמשים משויכים
        EMAIL_TO = os.environ.get('FORM_EMAIL_TO', 'studio@vatkin.co.il')
        
        # הוספת מיילים של משתמשים משויכים (אם יש)
        if client_id:
            assigned_emails = []
            for client in data:
                if client.get('id') == client_id:
                    assigned_users = client.get('assigned_user', [])
                    if isinstance(assigned_users, str):
                        assigned_users = [assigned_users]
                    for user_id in assigned_users:
                        if user_id in users:
                            user_email = users[user_id].get('email', '')
                            if user_email and user_email not in assigned_emails and user_email != sender_email:
                                assigned_emails.append(user_email)
                    break
            
            if assigned_emails:
                EMAIL_TO = f"{EMAIL_TO}, {', '.join(assigned_emails)}"
        
        print(f"[EMAIL] Sending FROM: {sender_email}")
        print(f"[EMAIL] Sending TO: {EMAIL_TO}")
        
        # קביעת שרת SMTP לפי הדומיין של המייל
        # עבור Gmail (כולל דומיינים מותאמים אישית שמשתמשים ב-Gmail)
        if '@gmail.com' in sender_email.lower() or '@googlemail.com' in sender_email.lower() or '@vatkin.co.il' in sender_email.lower():
            SMTP_SERVER = 'smtp.gmail.com'
            SMTP_PORT = 587
        elif '@outlook.com' in sender_email.lower() or '@hotmail.com' in sender_email.lower():
            SMTP_SERVER = 'smtp-mail.outlook.com'
            SMTP_PORT = 587
        else:
            # עבור דומיינים אחרים, נשתמש בהגדרות מ-.env או ברירת מחדל
            SMTP_SERVER = os.environ.get('SMTP_SERVER', 'smtp.gmail.com')
            SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))
        
        print(f"[EMAIL] SMTP Server: {SMTP_SERVER}:{SMTP_PORT}")
        
        # בניית תוכן המייל - תבנית HTML מושקעת עם אימוג'ים
        from html import escape
        current_date = datetime.now().strftime('%d/%m/%Y %H:%M')
        email_body_parts = []
        
        # Header עם רקע צבעוני
        email_body_parts.append("""
<!DOCTYPE html>
<html dir="rtl" lang="he">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
</head>
<body style="margin: 0; padding: 0; background-color: #f5f7fa; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; direction: rtl; text-align: right;">
    <table width="100%" cellpadding="0" cellspacing="0" style="background-color: #f5f7fa; padding: 40px 20px; direction: rtl;">
        <tr>
            <td align="center" style="direction: rtl;">
                <table width="600" cellpadding="0" cellspacing="0" style="background-color: #ffffff; border-radius: 16px; box-shadow: 0 4px 20px rgba(0,0,0,0.1); overflow: hidden; max-width: 100%; direction: rtl;">
                    <tr>
                        <td style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 40px 30px; text-align: center; direction: rtl;">
                            <div style="font-size: 48px; margin-bottom: 15px;">📝</div>
                            <h1 style="color: #ffffff; margin: 0; font-size: 28px; font-weight: bold; text-shadow: 0 2px 4px rgba(0,0,0,0.2); direction: rtl;">
                                טופס חדש התקבל!
                            </h1>
                            <p style="color: #f0f0f0; margin: 10px 0 0 0; font-size: 16px; direction: rtl;">היי, יש לך טופס חדש שממתין לטיפול</p>
                        </td>
                    </tr>
                    <tr>
                        <td style="padding: 30px; direction: rtl; text-align: right;">
                            <div style="background-color: #f8f9ff; border-right: 4px solid #667eea; padding: 20px; border-radius: 8px; margin-bottom: 25px; direction: rtl; text-align: right;">
                                <table width="100%" cellpadding="0" cellspacing="0" style="direction: rtl; text-align: right;">
                                    <tr>
                                        <td style="padding: 8px 0; direction: rtl; text-align: right;">
                                            <span style="font-size: 20px; margin-left: 10px;">📋</span>
                                            <strong style="color: #2d3748; font-size: 16px;">שם הטופס:</strong>
                                            <span style="color: #4a5568; font-size: 16px; margin-right: 8px;">""" + escape(str(form_title)) + """</span>
                                        </td>
                                    </tr>
                                    <tr>
                                        <td style="padding: 8px 0; direction: rtl; text-align: right;">
                                            <span style="font-size: 20px; margin-left: 10px;">👤</span>
                                            <strong style="color: #2d3748; font-size: 16px;">לקוח:</strong>
                                            <span style="color: #4a5568; font-size: 16px; margin-right: 8px;">""" + escape(str(client_name)) + """</span>
                                        </td>
                                    </tr>
                                    <tr>
                                        <td style="padding: 8px 0; direction: rtl; text-align: right;">
                                            <span style="font-size: 20px; margin-left: 10px;">🕐</span>
                                            <strong style="color: #2d3748; font-size: 16px;">תאריך ושעה:</strong>
                                            <span style="color: #4a5568; font-size: 16px; margin-right: 8px;">""" + current_date + """</span>
                                        </td>
                                    </tr>
                                </table>
                            </div>
                            
                            <div style="margin-bottom: 25px; direction: rtl; text-align: right;">
                                <h2 style="color: #2d3748; font-size: 22px; margin: 0 0 20px 0; padding-bottom: 10px; border-bottom: 2px solid #e2e8f0; direction: rtl; text-align: right;">
                                    <span style="font-size: 24px; margin-left: 8px;">📝</span>
                                    פרטי הטופס
                                </h2>
                                <table width="100%" cellpadding="0" cellspacing="0" style="border-collapse: separate; border-spacing: 0; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 8px rgba(0,0,0,0.05); direction: rtl; text-align: right;">
        """)
        
        # הוספת שדות הטופס
        field_count = 0
        for field_id, field_data in form_submission.items():
            field_count += 1
            bg_color = "#ffffff" if field_count % 2 == 0 else "#f8f9fa"
            # Escape HTML כדי למנוע בעיות אבטחה ותצוגה
            field_label = escape(str(field_data['label']))
            field_value_raw = field_data['value'] if field_data['value'] else ''
            if field_value_raw:
                # Escape HTML אבל שמור שורות חדשות
                field_value = escape(str(field_value_raw)).replace('\n', '<br>')
            else:
                field_value = '<span style="color: #a0aec0; font-style: italic;">לא הוזן</span>'
            email_body_parts.append(f"""
                                    <tr style="background-color: {bg_color}; direction: rtl;">
                                        <td style="padding: 16px 20px; border-bottom: 1px solid #e2e8f0; direction: rtl; text-align: right;">
                                            <strong style="color: #4a5568; font-size: 15px; display: block; margin-bottom: 6px; direction: rtl; text-align: right;">{field_label}</strong>
                                            <span style="color: #2d3748; font-size: 15px; line-height: 1.6; direction: rtl; text-align: right;">{field_value}</span>
                                        </td>
                                    </tr>
            """)
        
        # הוספת קבצים עם קישורי הורדה
        if uploaded_files:
            email_body_parts.append("""
                                    <tr style="background-color: #fff5e6; direction: rtl;">
                                        <td style="padding: 20px; direction: rtl; text-align: right;">
                                            <strong style="color: #2d3748; font-size: 16px; display: block; margin-bottom: 15px; direction: rtl; text-align: right;">
                                                <span style="font-size: 20px; margin-left: 8px;">📎</span>
                                                קבצים שהועלו:
                                            </strong>
            """)
            for field_id, file_data in uploaded_files.items():
                with app.app_context():
                    download_url = url_for('download_form_file_public', filename=file_data['saved_filename'], _external=True)
                email_body_parts.append(f"""
                                            <div style="margin-bottom: 12px; padding: 12px; background-color: #ffffff; border-radius: 6px; border: 1px solid #e2e8f0; direction: rtl; text-align: right;">
                                                <div style="margin-bottom: 6px; direction: rtl; text-align: right;">
                                                    <strong style="color: #4a5568; font-size: 14px;">{escape(str(file_data['label']))}:</strong>
                                                </div>
                                                <a href="{download_url}" style="display: inline-block; padding: 10px 20px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: #ffffff; text-decoration: none; border-radius: 6px; font-weight: bold; font-size: 14px; box-shadow: 0 2px 8px rgba(102, 126, 234, 0.3); direction: rtl;">
                                                    📥 הורד קובץ: {escape(str(file_data['filename']))}
                                                </a>
                                            </div>
                """)
            email_body_parts.append("""
                                        </td>
                                    </tr>
            """)
        
        email_body_parts.append("""
                                </table>
                            </div>
                        </td>
                    </tr>
                    <tr>
                        <td style="background-color: #f8f9fa; padding: 25px 30px; text-align: center; border-top: 1px solid #e2e8f0; direction: rtl;">
                            <p style="color: #718096; font-size: 13px; margin: 0; line-height: 1.6; direction: rtl;">
                                <span style="font-size: 16px;">✨</span><br>
                                נוצר אוטומטית ממערכת <strong style="color: #4a5568;">ותקין בוטיק</strong><br>
                                <span style="font-size: 12px; color: #a0aec0;">המייל נשלח ב-""" + current_date + """</span>
                            </p>
                        </td>
                    </tr>
                </table>
            </td>
        </tr>
    </table>
</body>
</html>
        """)
        
        email_body = "".join(email_body_parts)
        
        # יצירת הודעת מייל (alternative - HTML + טקסט, ללא קבצים מצורפים)
        msg = MIMEMultipart('alternative')
        msg['From'] = sender_email
        msg['To'] = EMAIL_TO
        msg['Subject'] = f"טופס חדש התקבל: {form_title} - {client_name}"
        
        # הוספת תוכן טקסט פשוט (fallback) - צריך להיות ראשון
        text_body = f"טופס חדש התקבל\n\nשם הטופס: {form_title}\nלקוח: {client_name}\nתאריך: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n\nפרטים:\n"
        for field_id, field_data in form_submission.items():
            text_body += f"{field_data['label']}: {field_data['value']}\n"
        if uploaded_files:
            text_body += "\nקבצים (להורדה דרך הקישורים במייל):\n"
            for field_id, file_data in uploaded_files.items():
                with app.app_context():
                    download_url = url_for('download_form_file_public', filename=file_data['saved_filename'], _external=True)
                text_body += f"{file_data['label']}: {file_data['filename']} - {download_url}\n"
        
        text_part = MIMEText(text_body, 'plain', 'utf-8')
        msg.attach(text_part)
        
        # הוספת תוכן HTML - צריך להיות אחרון כדי ש-HTML ייבחר
        html_part = MIMEText(email_body, 'html', 'utf-8')
        msg.attach(html_part)
        
        # שליחת המייל
        print(f"[EMAIL] Connecting to SMTP server...")
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()  # הצפנה
        print(f"[EMAIL] Logging in as {sender_email}...")
        server.login(sender_email, sender_password)
        print(f"[EMAIL] Sending email...")
        server.send_message(msg)
        server.quit()
        
        sent_at = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        
        print("\n" + "="*80)
        print(f"[SUCCESS] ✅ מייל נשלח בהצלחה!")
        print(f"[SUCCESS] מ: {sender_email}")
        print(f"[SUCCESS] ל: {EMAIL_TO}")
        print(f"[SUCCESS] נושא: טופס חדש התקבל: {form_title} - {client_name}")
        print(f"[SUCCESS] תאריך: {sent_at}")
        print("="*80 + "\n")
        
        # החזרת פרטי המייל לאישור
        return {
            'success': True,
            'from_email': sender_email,
            'to_email': EMAIL_TO,
            'subject': f"טופס חדש התקבל: {form_title} - {client_name}",
            'sent_at': sent_at
        }
        
    except Exception as e:
        error_msg = str(e).encode('ascii', errors='replace').decode('ascii')
        print(f"[ERROR] שגיאה בשליחת מייל: {error_msg}")
        import traceback
        print(traceback.format_exc())
        return False

def get_event_checklist_template(event_type):
    """מחזיר תבנית צ'ק-ליסט לפי סוג אירוע"""
    templates = load_checklist_templates()
    return templates.get(event_type, [])

class User(UserMixin):
    def __init__(self, id):
        self.id = id
        u = load_users()
        self.name = u[id]['name'] if id in u else "Unknown"

@login_manager.user_loader
def load_user(user_id):
    u = load_users()
    return User(user_id) if user_id in u else None

@app.context_processor
def inject_sidebar_data():
    users = load_users()
    sorted_users = dict(sorted(users.items(), key=lambda item: item[1]['name']))
    return dict(sidebar_users=sorted_users)

# --- Routes ---

def load_user_activity():
    """טעינת נתוני פעילות משתמשים"""
    if not os.path.exists(USER_ACTIVITY_FILE):
        return {}
    try:
        with open(USER_ACTIVITY_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        return {}

def save_user_activity(activity):
    """שמירת נתוני פעילות משתמשים"""
    with open(USER_ACTIVITY_FILE, 'w', encoding='utf-8') as f:
        json.dump(activity, f, ensure_ascii=False, indent=4)

def update_user_activity(user_id):
    """עדכון זמן פעילות אחרון של משתמש"""
    activity = load_user_activity()
    activity[user_id] = datetime.now().isoformat()
    save_user_activity(activity)

def is_user_active(user_id):
    """בודק אם משתמש פעיל (פעיל ב-60 דקות האחרונות)"""
    activity = load_user_activity()
    if user_id not in activity:
        return False
    try:
        last_activity = datetime.fromisoformat(activity[user_id])
        time_diff = datetime.now() - last_activity
        return time_diff.total_seconds() < 3600  # שעה
    except:
        return False

def load_activity_logs():
    """טעינת לוגי פעילות"""
    if not os.path.exists(ACTIVITY_LOGS_FILE):
        return []
    try:
        with open(ACTIVITY_LOGS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        return []

def save_activity_logs(logs):
    """שמירת לוגי פעילות"""
    with open(ACTIVITY_LOGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(logs, f, ensure_ascii=False, indent=4)

@app.before_request
def track_activity():
    """עקוב אחר פעילות משתמשים לפני כל בקשת"""
    if current_user.is_authenticated:
        update_user_activity(current_user.id)

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("20 per minute")  # מקסימום 20 ניסיונות התחברות לדקה (5 היה מעט מדי)
@csrf.exempt  # פטור זמני מ-CSRF עד שנוסיף tokens לכל הטפסים
def login():
    # אם המשתמש כבר מחובר, העבר אותו ישר לדשבורד
    if current_user.is_authenticated:
        return redirect(url_for('home'))
    
    if request.method == 'POST':
        u = load_users(); uid, pwd = request.form.get('username'), request.form.get('password')
        email_match_user = None
        if uid:
            normalized_uid = uid.strip().lower()
            for user_id, info in u.items():
                if (info.get('email') or '').strip().lower() == normalized_uid:
                    email_match_user = user_id
                    break
        resolved_user_id = uid if uid in u else email_match_user
        # בדיקת סיסמה - תומך גם בסיסמאות מוצפנות וגם בטקסט פשוט (להתאימות לאחור)
        if resolved_user_id and resolved_user_id in u:
            stored_password = u[resolved_user_id].get('password', '')
            # בדיקה אם הסיסמה מוצפנת (תומך ב-pbkdf2 ו-scrypt)
            if stored_password.startswith('pbkdf2:sha256:') or stored_password.startswith('scrypt:'):
                password_valid = check_password_hash(stored_password, pwd)
            else:
                # תמיכה בסיסמאות ישנות (לא מוצפנות) - רק למטרות migration
                password_valid = (stored_password == pwd)
            
            if password_valid:
                user = User(resolved_user_id)
                login_user(user, remember=True)
                update_user_activity(resolved_user_id)
                wants_json = request.headers.get('Accept', '').find('application/json') != -1 or \
                            request.headers.get('X-Requested-With') == 'XMLHttpRequest'
                if wants_json:
                    return jsonify({'status': 'success'}), 200
                return redirect(url_for('home'))
        
        wants_json = request.headers.get('Accept', '').find('application/json') != -1 or \
                    request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if wants_json:
            return jsonify({'status': 'error', 'error': 'שם משתמש או סיסמה שגויים'}), 401
        from flask import flash
        flash('שם משתמש או סיסמה שגויים', 'error')
    # Redirect to React login
    return redirect('/app/login')


def send_password_reset_email(user_email, reset_token):
    """שליחת מייל לאיפוס סיסמה"""
    try:
        SMTP_SERVER = os.environ.get('SMTP_SERVER', 'smtp.gmail.com')
        SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))
        # פרטי SMTP מ-environment variables בלבד (לא hardcoded)
        SMTP_USERNAME = os.environ.get('SMTP_USERNAME')
        SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD')
        
        if not SMTP_USERNAME or not SMTP_PASSWORD:
            print("[WARNING] שליחת מייל מושבתת - אין הגדרות SMTP")
            return False
        
        reset_url = f"http://127.0.0.1:5000/reset_password/{reset_token}"
        
        email_body = f"""
        <html dir='rtl'>
        <body style='font-family: Heebo, sans-serif;'>
            <h2 style='color: #0073ea;'>איפוס סיסמה</h2>
            <p>שלום,</p>
            <p>התקבלה בקשה לאיפוס הסיסמה שלך במערכת.</p>
            <p>לחץ על הקישור הבא כדי לאפס את הסיסמה:</p>
            <p><a href='{reset_url}' style='background: #0073ea; color: white; padding: 12px 24px; text-decoration: none; border-radius: 8px; display: inline-block; margin: 15px 0;'>איפוס סיסמה</a></p>
            <p>או העתק את הקישור הבא לדפדפן:</p>
            <p style='background: #f8fafc; padding: 10px; border-radius: 5px; word-break: break-all;'>{reset_url}</p>
            <p>אם לא ביקשת לאפס את הסיסמה, אנא התעלם ממייל זה.</p>
            <p>תוקף הקישור: 24 שעות</p>
        </body>
        </html>
        """
        
        msg = MIMEMultipart('alternative')
        msg['From'] = SMTP_USERNAME
        msg['To'] = user_email
        msg['Subject'] = 'איפוס סיסמה - מערכת ניהול'
        
        html_part = MIMEText(email_body, 'html', 'utf-8')
        msg.attach(html_part)
        
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(SMTP_USERNAME, SMTP_PASSWORD)
        server.send_message(msg)
        server.quit()
        
        print(f"[SUCCESS] מייל איפוס סיסמה נשלח ל-{user_email}")
        return True
    except Exception as e:
        print(f"[ERROR] שגיאה בשליחת מייל איפוס סיסמה: {e}")
        return False

@app.route('/reset_password_request', methods=['POST'])
@csrf.exempt  # פטור מ-CSRF כי זה נקרא מ-modal ולא צריך token
def reset_password_request():
    """בקשת איפוס סיסמה - שולח מייל"""
    from flask import flash
    username = request.form.get('username', '').strip()
    
    if not username:
        flash('נא להזין שם משתמש', 'error')
        return redirect(url_for('login'))
    
    users = load_users()
    
    # חיפוש משתמש לפי שם משתמש
    user = None
    for uid, user_data in users.items():
        if uid == username:
            user = user_data
            user_id = uid
            break
    
    if not user:
        # לא נגלה למשתמש שהמשתמש לא קיים (אבטחה)
        flash('אם המשתמש קיים במערכת, קישור איפוס סיסמה נשלח למייל', 'success')
        return redirect(url_for('login'))
    
    user_email = user.get('email', '')
    if not user_email:
        flash('למשתמש זה לא רשום מייל במערכת. אנא פנה למנהל המערכת.', 'error')
        return redirect(url_for('login'))
    
    # יצירת טוקן איפוס סיסמה
    reset_token = str(uuid.uuid4())
    
    # שמירת טוקן איפוס סיסמה (ניתן לשמור בקובץ נפרד או ב-JSON)
    RESET_TOKENS_FILE = os.path.join(BASE_DIR, 'reset_tokens.json')
    reset_tokens = {}
    if os.path.exists(RESET_TOKENS_FILE):
        with open(RESET_TOKENS_FILE, 'r', encoding='utf-8') as f:
            reset_tokens = json.load(f)
    
    reset_tokens[reset_token] = {
        'user_id': user_id,
        'created': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'used': False
    }
    
    with open(RESET_TOKENS_FILE, 'w', encoding='utf-8') as f:
        json.dump(reset_tokens, f, ensure_ascii=False, indent=4)
    
    # שליחת מייל
    email_sent = send_password_reset_email(user_email, reset_token)
    
    if email_sent:
        flash('קישור איפוס סיסמה נשלח למייל שלך', 'success')
    else:
        flash('שגיאה בשליחת המייל. אנא פנה למנהל המערכת.', 'error')
    
    return redirect(url_for('login'))

@app.route('/reset_password/<token>', methods=['GET', 'POST'])
@csrf.exempt  # פטור מ-CSRF כי יש token מיוחד משלו
def reset_password(token):
    """איפוס סיסמה עם טוקן"""
    from flask import flash
    
    RESET_TOKENS_FILE = os.path.join(BASE_DIR, 'reset_tokens.json')
    
    # טעינת טוקנים
    if not os.path.exists(RESET_TOKENS_FILE):
        return "קישור לא תקין או פג תוקף", 400
    
    with open(RESET_TOKENS_FILE, 'r', encoding='utf-8') as f:
        reset_tokens = json.load(f)
    
    if token not in reset_tokens:
        return "קישור לא תקין או פג תוקף", 400
    
    token_data = reset_tokens[token]
    
    # בדיקת תוקף (24 שעות)
    created_time = datetime.strptime(token_data['created'], '%Y-%m-%d %H:%M:%S')
    if (datetime.now() - created_time).total_seconds() > 24 * 3600:
        del reset_tokens[token]
        with open(RESET_TOKENS_FILE, 'w', encoding='utf-8') as f:
            json.dump(reset_tokens, f, ensure_ascii=False, indent=4)
        return "קישור פג תוקף. אנא בקש קישור חדש.", 400
    
    if token_data.get('used', False):
        return "קישור זה כבר נעשה בו שימוש", 400
    
    if request.method == 'POST':
        new_password = request.form.get('password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        
        if not new_password or len(new_password) < 4:
            flash('סיסמה חייבת להכיל לפחות 4 תווים', 'error')
            return render_template('reset_password.html', token=token)
        
        if new_password != confirm_password:
            flash('הסיסמאות לא תואמות', 'error')
            return render_template('reset_password.html', token=token)
        
        # עדכון סיסמה
        users = load_users()
        user_id = token_data['user_id']
        
        if user_id in users:
            users[user_id]['password'] = generate_password_hash(new_password)
            with open(USERS_FILE, 'w', encoding='utf-8') as f:
                json.dump(users, f, ensure_ascii=False, indent=4)
            
            # סימון טוקן כמשומש
            reset_tokens[token]['used'] = True
            with open(RESET_TOKENS_FILE, 'w', encoding='utf-8') as f:
                json.dump(reset_tokens, f, ensure_ascii=False, indent=4)
            
            flash('הסיסמה עודכנה בהצלחה! ניתן להתחבר עם הסיסמה החדשה.', 'success')
            return redirect(url_for('login'))
        else:
            flash('שגיאה בעדכון הסיסמה', 'error')
            return render_template('reset_password.html', token=token)
    
    return render_template('reset_password.html', token=token)

@app.route('/')
@login_required
def home():
    # Redirect to React app
    return redirect('/app')

@app.route('/api/current_user')
@limiter.exempt  # פטור מ-rate limiting כי זה נקרא הרבה פעמים
def api_current_user():
    """API endpoint להחזרת המשתמש הנוכחי"""
    try:
        # Check if user is authenticated without redirecting
        if not current_user.is_authenticated:
            return jsonify({'success': False, 'error': 'Not authenticated'}), 401
        
        users = load_users()
        user_data = users.get(current_user.id, {})
        return jsonify({
            'success': True,
            'user': {
                'id': current_user.id,
                'name': user_data.get('name', current_user.id),
                'email': user_data.get('email', ''),
                'role': get_user_role(current_user.id)
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/sidebar_users')
@login_required
def api_sidebar_users():
    """API endpoint להחזרת משתמשים לסיידבר"""
    try:
        users = load_users()
        sidebar_users = {uid: {'name': info.get('name', '')} for uid, info in users.items() if uid != 'admin'}
        # Return as list for React
        users_list = [
            {'id': uid, 'name': info.get('name', '')}
            for uid, info in users.items() if uid != 'admin'
        ]
        return jsonify({
            'success': True,
            'users': users_list,
            'users_dict': sidebar_users
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== NOTIFICATIONS API ====================

from backend.utils.notifications import (
    get_user_notifications, get_unread_count,
    mark_notifications_read, get_new_notifications_since
)

@app.route('/api/notifications')
@login_required
def api_get_notifications():
    """Get notifications for the current user"""
    try:
        unread_only = request.args.get('unread_only', 'false').lower() == 'true'
        limit = int(request.args.get('limit', 50))
        
        notifications = get_user_notifications(
            current_user.id, 
            unread_only=unread_only,
            limit=limit
        )
        
        return jsonify({
            'success': True,
            'notifications': notifications
        })
    except Exception as e:
        print(f"Error in api_get_notifications: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/notifications/unread-count')
@login_required
@limiter.exempt
def api_unread_notifications_count():
    """Get count of unread notifications for the current user"""
    try:
        count = get_unread_count(current_user.id)
        
        return jsonify({
            'success': True,
            'count': count
        })
    except Exception as e:
        print(f"Error in api_unread_notifications_count: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/notifications/new')
@login_required
@limiter.exempt
def api_new_notifications():
    """Get new notifications since a timestamp (for polling)"""
    try:
        since = request.args.get('since', '')
        
        if not since:
            # If no timestamp provided, return unread notifications
            notifications = get_user_notifications(
                current_user.id, 
                unread_only=True,
                limit=10
            )
        else:
            notifications = get_new_notifications_since(current_user.id, since)
        
        return jsonify({
            'success': True,
            'notifications': notifications,
            'count': len(notifications)
        })
    except Exception as e:
        print(f"Error in api_new_notifications: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/notifications/mark-read', methods=['POST'])
@login_required
@csrf.exempt
def api_mark_notifications_read():
    """Mark notifications as read"""
    try:
        data = request.get_json() if request.is_json else {}
        notification_ids = data.get('notification_ids', [])
        mark_all = data.get('mark_all', False)
        
        if mark_all:
            count = mark_notifications_read('all', user_id=current_user.id)
        elif notification_ids:
            count = mark_notifications_read(notification_ids)
        else:
            return jsonify({'success': False, 'error': 'No notifications specified'}), 400
        
        return jsonify({
            'success': True,
            'marked_count': count
        })
    except Exception as e:
        print(f"Error in api_mark_notifications_read: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/clients')
@login_required
def api_clients():
    """API endpoint להחזרת לקוחות"""
    try:
        user_role = get_user_role(current_user.id)
        all_c = load_data()
        all_c = filter_active_clients(all_c)
        if is_manager_or_admin(current_user.id, user_role):
            display = all_c
        else:
            display = [c for c in all_c if can_user_access_client(current_user.id, user_role, c)]
        display = sorted(display, key=lambda x: x.get('name', '').lower())
        return jsonify({
            'success': True,
            'clients': [{'id': c['id'], 'name': c.get('name', '')} for c in display]
        })
    except Exception as e:
        print(f"Error in api_clients: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/get_client_projects/<client_id>')
@login_required
def get_client_projects(client_id):
    """Route להחזרת פרויקטים של לקוח"""
    try:
        data = load_data()
        for c in data:
            if c['id'] == client_id:
                projects = c.get('projects', [])
                return jsonify({'success': True, 'projects': [{'id': p['id'], 'title': p['title']} for p in projects]})
        return jsonify({'success': False, 'error': 'לקוח לא נמצא'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/tasks/calendar')
@login_required
def get_tasks_for_calendar():
    """API endpoint להחזרת משימות עם deadline ללוח שנה"""
    try:
        data = load_data()
        users = load_users()
        user_role = get_user_role(current_user.id)
        
        tasks = []
        for client in data:
            # בדיקת הרשאות
            if not is_manager_or_admin(current_user.id, user_role):
                if not can_user_access_client(current_user.id, user_role, client):
                    continue
            
            client_name = client.get('name', '')
            for project in client.get('projects', []):
                project_title = project.get('title', '')
                for task in project.get('tasks', []):
                    deadline = task.get('deadline', '')
                    if not deadline:
                        continue
                    
                    # סינון משימות שהושלמו (למעט משימות יומיות)
                    task_status = task.get('status', 'לביצוע')
                    is_daily_task = task.get('is_daily_task', False)
                    if task_status == 'הושלם' and not is_daily_task:
                        continue
                    
                    # המרת תאריך לפורמט ISO
                    try:
                        if 'T' in deadline:
                            deadline_date = deadline.split('T')[0]
                        else:
                            deadline_date = deadline
                        
                        # קבלת שם המשתמש האחראי
                        assignee_id = task.get('assignee', '')
                        assignee_name = users.get(assignee_id, {}).get('name', assignee_id) if assignee_id else 'ללא אחראי'
                        
                        task_title = task.get('title', 'ללא כותרת')
                        
                        tasks.append({
                            'id': task.get('id', ''),
                            'title': task_title,
                            'start': deadline_date,
                            'client_name': client_name,
                            'project_title': project_title,
                            'assignee_name': assignee_name,
                            'assignee_id': assignee_id,
                            'status': task_status,
                            'client_id': client.get('id', ''),
                            'project_id': project.get('id', ''),
                            'task_id': task.get('id', '')
                        })
                    except Exception as e:
                        print(f"Error parsing deadline for task: {e}")
                        continue
        
        return jsonify({'success': True, 'tasks': tasks})
    except Exception as e:
        print(f"Error in get_tasks_for_calendar: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/quick_update_tasks')
@login_required
def api_quick_update_tasks():
    """API endpoint להחזרת כל המשימות לעדכון מהיר"""
    try:
        data = load_data()
        users = load_users()
        user_role = get_user_role(current_user.id)
        
        tasks = []
        for client in data:
            # בדיקת הרשאות
            if not is_manager_or_admin(current_user.id, user_role):
                if not can_user_access_client(current_user.id, user_role, client):
                    continue
            
            client_name = client.get('name', '')
            for project in client.get('projects', []):
                project_title = project.get('title', '')
                for task in project.get('tasks', []):
                    # רק משימות שלא הושלמו
                    task_status = task.get('status', 'ממתין')
                    if task_status not in ['ממתין', 'בביצוע']:
                        continue
                    
                    # קבלת שם המשתמש האחראי
                    assignee_id = task.get('assignee', '') or task.get('assigned_to', '')
                    assignee_name = users.get(assignee_id, {}).get('name', assignee_id) if assignee_id else 'ללא אחראי'
                    
                    tasks.append({
                        'client_id': client.get('id', ''),
                        'client_name': client_name,
                        'project_id': project.get('id', ''),
                        'project_title': project_title,
                        'task': {
                            'id': task.get('id', ''),
                            'desc': task.get('desc', '') or task.get('title', 'ללא כותרת'),
                            'status': task_status,
                            'notes': task.get('notes', '') or task.get('note', ''),
                            'assigned_to_name': assignee_name
                        }
                    })
        
        return jsonify({'success': True, 'tasks': tasks})
    except Exception as e:
        print(f"Error in api_quick_update_tasks: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/quick_add_task', methods=['POST'])
@login_required
@csrf.exempt
def quick_add_task():
    """Route להוספת משימה מהירה מהדשבורד"""
    try:
        # תמיכה גם ב-JSON וגם ב-form data
        if request.is_json:
            data_json = request.get_json()
            client_id = data_json.get('client_id')
            project_id = data_json.get('project_id')
            task_title = data_json.get('task_title')
            task_status = data_json.get('task_status', 'לביצוע')
            task_note = data_json.get('task_note', '')
            task_deadline = data_json.get('task_deadline', '')
            is_daily_task = data_json.get('is_daily_task', False)
            assigned_to = data_json.get('assigned_to', '')  # NEW: Support task assignment
        else:
            client_id = request.form.get('client_id')
            project_id = request.form.get('project_id')
            task_title = request.form.get('task_title')
            task_status = request.form.get('task_status', 'לביצוע')
            task_note = request.form.get('task_note', '')
            task_deadline = request.form.get('task_deadline', '')
            is_daily_task = request.form.get('is_daily_task', 'false').lower() == 'true'
            assigned_to = request.form.get('assigned_to', '')  # NEW: Support task assignment
        
        if not client_id or not project_id or not task_title:
            wants_json = request.headers.get('Accept', '').find('application/json') != -1 or \
                        request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            if wants_json:
                return jsonify({'success': False, 'error': 'חסרים פרמטרים נדרשים'}), 400
            return redirect(url_for('home'))
        
        # Load users for notification
        users = load_users()
        created_by = current_user.id
        creator_name = users.get(created_by, {}).get('name', created_by)
        
        # If no assignee specified, assign to current user
        if not assigned_to:
            assigned_to = created_by
        
        data = load_data()
        for c in data:
            if c['id'] == client_id:
                for p in c.get('projects', []):
                    if p['id'] == project_id:
                        task = {
                            'id': str(uuid.uuid4()),
                            'title': task_title,
                            'status': task_status,
                            'note': task_note,
                            'created_date': datetime.now().strftime('%Y-%m-%d'),
                            'created_at': datetime.now().isoformat(),
                            'created_by': created_by,  # NEW: Track who created the task
                            'assigned_user': [assigned_to] if assigned_to else [created_by],  # NEW: Track assignee
                            'done': False
                        }
                        # הוסף deadline אם קיים
                        if task_deadline:
                            try:
                                # המרת תאריך לפורמט ISO
                                deadline = datetime.strptime(task_deadline, '%Y-%m-%d').isoformat()
                                task['deadline'] = deadline
                            except Exception as e:
                                print(f"Error parsing deadline: {e}")
                        
                        # משימה יומית - קבע ליום העבודה הקרוב
                        if is_daily_task:
                            task['is_daily_task'] = True
                            # מצא את יום העבודה הקרוב (א'-ה')
                            next_workday = get_next_workday()
                            task['deadline'] = next_workday.isoformat()
                        
                        task_number = get_next_task_number(c, p)
                        task['task_number'] = task_number
                        p.setdefault('tasks', []).append(task)
                        save_data(data)
                        
                        # NEW: Create notification if assigned to someone else
                        if assigned_to and assigned_to != created_by:
                            create_notification(
                                user_id=assigned_to,
                                notification_type='task_assigned',
                                data={
                                    'task_id': task['id'],
                                    'client_id': client_id,
                                    'project_id': project_id,
                                    'from_user_id': created_by,
                                    'from_user_name': creator_name,
                                    'task_title': task_title,
                                    'client_name': c.get('name', '')
                                }
                            )
                        
                        # בדיקה אם זה AJAX request
                        wants_json = request.headers.get('Accept', '').find('application/json') != -1 or \
                                    request.headers.get('X-Requested-With') == 'XMLHttpRequest'
                        if wants_json:
                            return jsonify({
                                'success': True,
                                'message': 'המשימה נוספה בהצלחה',
                                'task': task
                            })
                        return redirect(url_for('client_page', client_id=client_id))
        
        wants_json = request.headers.get('Accept', '').find('application/json') != -1 or \
                    request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if wants_json:
            return jsonify({'success': False, 'error': 'לקוח או פרויקט לא נמצאו'}), 404
        return redirect(url_for('home'))
    except Exception as e:
        print(f"Error in quick_add_task: {e}")
        import traceback
        traceback.print_exc()
        wants_json = request.headers.get('Accept', '').find('application/json') != -1 or \
                    request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if wants_json:
            return jsonify({'success': False, 'error': str(e)}), 500
        return redirect(url_for('home'))

@app.route('/quick_add_charge', methods=['POST'])
@login_required
@csrf.exempt  # פטור מ-CSRF כי זה API call מ-JavaScript
def quick_add_charge():
    """Route להוספת חיוב מהיר מהדשבורד"""
    try:
        client_id = request.form.get('client_id')
        charge_title = request.form.get('charge_title')
        charge_amount = request.form.get('charge_amount')
        
        if not client_id or not charge_title or not charge_amount:
            wants_json = request.headers.get('Accept', '').find('application/json') != -1 or \
                        request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            if wants_json:
                return jsonify({'success': False, 'error': 'חסרים שדות נדרשים'}), 400
            return redirect(url_for('home'))
        
        data = load_data()
        for c in data:
            if c['id'] == client_id:
                if 'extra_charges' not in c:
                    c['extra_charges'] = []
                charge_number = get_next_charge_number(c)
                our_cost = float(request.form.get('charge_our_cost', 0) or 0)
                description = request.form.get('charge_description', '')
                c['extra_charges'].append({
                    'id': str(uuid.uuid4()),
                    'title': charge_title,
                    'description': description,
                    'amount': int(float(charge_amount)),
                    'our_cost': our_cost,
                    'date': datetime.now().strftime("%d/%m/%y"),
                    'completed': False,
                    'charge_number': charge_number
                })
                save_data(data)
                # בדיקה אם זה AJAX request
                wants_json = request.headers.get('Accept', '').find('application/json') != -1 or \
                            request.headers.get('X-Requested-With') == 'XMLHttpRequest'
                if wants_json:
                    return jsonify({'success': True, 'message': 'החיוב נוסף בהצלחה'})
                return redirect(url_for('client_page', client_id=client_id))
        
        wants_json = request.headers.get('Accept', '').find('application/json') != -1 or \
                    request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if wants_json:
            return jsonify({'success': False, 'error': 'לקוח לא נמצא'}), 404
        return redirect(url_for('home'))
    except Exception as e:
        print(f"Error in quick_add_charge: {e}")
        wants_json = request.headers.get('Accept', '').find('application/json') != -1 or \
                    request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if wants_json:
            return jsonify({'success': False, 'error': str(e)}), 500
        return redirect(url_for('home'))

@app.route('/api/all_clients')
@login_required
def api_all_clients():
    """API endpoint להחזרת לקוחות"""
    try:
        user_role = get_user_role(current_user.id)
        if not check_permission('/all_clients', user_role):
            return jsonify({'success': False, 'error': 'גישה חסומה'}), 403
        
        all_clients_data = load_data()
        all_clients_data = filter_active_clients(all_clients_data)
        users = load_users()
        filter_user = request.args.get('user')
        
        # ארגון לקוחות לפי משתמשים
        clients_by_user = {}
        for uid in users.keys():
            clients_by_user[uid] = []
        
        for c in all_clients_data:
            assigned = normalize_assigned_user(c.get('assigned_user', []))
            for assigned_uid in assigned:
                assigned_uid_lower = assigned_uid.lower() if isinstance(assigned_uid, str) else str(assigned_uid).lower()
                for uid in users.keys():
                    uid_lower = uid.lower() if isinstance(uid, str) else str(uid).lower()
                    if assigned_uid == uid or assigned_uid_lower == uid_lower:
                        if c not in clients_by_user[uid]:
                            clients_by_user[uid].append(c)
                        break
        
        if filter_user:
            filter_user_lower = filter_user.lower()
            filtered_clients = []
            for uid, client_list in clients_by_user.items():
                uid_lower = uid.lower() if isinstance(uid, str) else str(uid).lower()
                if uid == filter_user or uid_lower == filter_user_lower:
                    filtered_clients = client_list
                    break
            filtered_clients.sort(key=lambda x: x.get('name', '').lower())
            clients = filtered_clients
        else:
            clients = all_clients_data
            clients.sort(key=lambda x: x.get('name', '').lower())
        
        # Convert to simple format
        clients_list = [{'id': c['id'], 'name': c.get('name', '')} for c in clients]
        clients_by_user_dict = {
            uid: [{'id': c['id'], 'name': c.get('name', '')} for c in client_list]
            for uid, client_list in clients_by_user.items()
        }
        users_dict = {uid: {'name': info.get('name', '')} for uid, info in users.items()}
        
        return jsonify({
            'success': True,
            'clients': clients_list,
            'clients_by_user': clients_by_user_dict,
            'users': users_dict,
            'filter_user': filter_user,
            'user_role': user_role
        })
    except Exception as e:
        print(f"Error in api_all_clients: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/all_clients')
@login_required
def all_clients():
    user_role = get_user_role(current_user.id)
    if not check_permission('/all_clients', user_role):
        return "גישה חסומה - אין לך הרשאה לגשת לדף זה", 403
    all_clients_data = load_data()
    # סינון לקוחות מאוישים
    all_clients_data = filter_active_clients(all_clients_data)
    users = load_users()
    filter_user = request.args.get('user')
    
    # ארגון לקוחות לפי משתמשים
    clients_by_user = {}
    for uid in users.keys():
        clients_by_user[uid] = []
    
    for c in all_clients_data:
        assigned = normalize_assigned_user(c.get('assigned_user', []))
        
        # הוסף את הלקוח לכל המשתמשים המשויכים (case-insensitive)
        for assigned_uid in assigned:
            assigned_uid_lower = assigned_uid.lower() if isinstance(assigned_uid, str) else str(assigned_uid).lower()
            # מצא את המשתמש הנכון (case-insensitive)
            for uid in users.keys():
                uid_lower = uid.lower() if isinstance(uid, str) else str(uid).lower()
                if assigned_uid == uid or assigned_uid_lower == uid_lower:
                    if c not in clients_by_user[uid]:
                        clients_by_user[uid].append(c)
                    break
    
    # אם יש סינון לפי משתמש ספציפי, החזר רק את הלקוחות שלו
    if filter_user:
        filter_user_lower = filter_user.lower()
        filtered_clients = []
        for uid, client_list in clients_by_user.items():
            uid_lower = uid.lower() if isinstance(uid, str) else str(uid).lower()
            if uid == filter_user or uid_lower == filter_user_lower:
                filtered_clients = client_list
                break
        # מיון לפי שם
        filtered_clients.sort(key=lambda x: x.get('name', '').lower())
        clients = filtered_clients
    else:
        clients = all_clients_data
        clients.sort(key=lambda x: x.get('name', '').lower())
    
    # Redirect to React
    return redirect('/app/clients')

@app.route('/project/<client_id>/<project_id>/gantt')
@login_required
def project_gantt(client_id, project_id):
    """עמוד תרשים גאנט לפרויקט"""
    user_role = get_user_role(current_user.id)
    if not check_permission('/client/', user_role):
        return "גישה חסומה - אין לך הרשאה לגשת לדף זה", 403
    data = load_data()
    users = load_users()
    client = next((c for c in data if c['id'] == client_id), None)
    if not client:
        return "לקוח לא נמצא", 404
    
    # בדיקת הרשאות גישה ללקוח
    if not can_user_access_client(current_user.id, user_role, client):
        return "גישה חסומה - אין לך הרשאה ללקוח זה", 403
    
    # מצא את הפרויקט
    project = None
    for p in client.get('projects', []):
        if p['id'] == project_id:
            project = p
            break
    
    if not project:
        return "פרויקט לא נמצא", 404
    
    # חישוב תאריכי יעד אוטומטיים על בסיס תלויות
    calculate_dependent_deadlines_for_project(project)
    
    # הכנת נתוני משימות לתרשים גאנט
    tasks_data = []
    for task in project.get('tasks', []):
        deadline = task.get('deadline', '')
        start_date = task.get('created_at', datetime.now().isoformat())
        if 'T' in start_date:
            start_date = start_date.split('T')[0]
        if 'T' in deadline:
            deadline = deadline.split('T')[0]
        
        tasks_data.append({
            'id': task.get('id'),
            'title': task.get('title', 'ללא שם'),
            'start_date': start_date,
            'deadline': deadline,
            'created_at': task.get('created_at', ''),
            'status': task.get('status', 'לביצוע'),
            'priority': task.get('priority', 'medium'),
            'assignee': task.get('assignee', ''),
            'dependencies': task.get('dependencies', []),
            'estimated_hours': task.get('estimated_hours', 8),
            'progress': 100 if task.get('status') == 'הושלם' else 0
        })
    
    # Redirect to React client page
    return redirect(f'/app/client/{client_id}')

def calculate_dependent_deadlines_for_project(project):
    """מחשב תאריכי יעד אוטומטיים למשימות בפרויקט על בסיס תלויות"""
    tasks = project.get('tasks', [])
    task_dict = {t['id']: t for t in tasks}
    
    # עבור כל משימה, בדוק אם יש תלויות וחשב תאריך יעד
    for task in tasks:
        dependencies = task.get('dependencies', [])
        if dependencies and not task.get('deadline'):
            # מצא את התאריך המקסימלי מתוך התלויות
            max_deadline = None
            for dep_id in dependencies:
                dep_task = task_dict.get(dep_id)
                if dep_task:
                    dep_deadline = dep_task.get('deadline') or dep_task.get('created_at')
                    if dep_deadline:
                        try:
                            if 'T' in dep_deadline:
                                dep_date = datetime.fromisoformat(dep_deadline.split('T')[0])
                            else:
                                dep_date = datetime.fromisoformat(dep_deadline)
                            
                            if not max_deadline or dep_date > max_deadline:
                                max_deadline = dep_date
                        except:
                            continue
            
            # אם יש תאריך מקסימלי, הוסף ימים משוערים
            if max_deadline:
                estimated_hours = task.get('estimated_hours', 8)
                # הנחה: 8 שעות = יום עבודה אחד
                estimated_days = max(1, int(estimated_hours / 8))
                new_deadline = max_deadline + timedelta(days=estimated_days)
                task['deadline'] = new_deadline.isoformat().split('T')[0]

@app.route('/api/client/<client_id>')
@login_required
def api_client(client_id):
    """API endpoint להחזרת פרטי לקוח"""
    try:
        user_role = get_user_role(current_user.id)
        if not check_permission('/client/', user_role):
            return jsonify({'success': False, 'error': 'גישה חסומה'}), 403
        
        data = load_data()
        client = next((c for c in data if c['id'] == client_id), None)
        if not client:
            return jsonify({'success': False, 'error': 'לקוח לא נמצא'}), 404
        
        # בדיקת הרשאות גישה ללקוח
        if not can_user_access_client(current_user.id, user_role, client):
            return jsonify({'success': False, 'error': 'גישה חסומה ללקוח זה'}), 403
        
        # טעינת היסטוריית אינטראקציות
        activity_logs = load_activity_logs()
        client_activities = [log for log in activity_logs if log.get('client_id') == client_id]
        client_activities.sort(key=lambda x: x.get('timestamp', ''), reverse=True)
        
        # Add activities to client
        client_copy = client.copy()
        client_copy['activities'] = client_activities
        
        return jsonify({
            'success': True,
            'client': client_copy
        })
    except Exception as e:
        print(f"Error in api_client: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/client/<client_id>')
@login_required
def client_page(client_id):
    user_role = get_user_role(current_user.id)
    if not check_permission('/client/', user_role):
        return "גישה חסומה - אין לך הרשאה לגשת לדף זה", 403
    data = load_data(); users = load_users()
    client = next((c for c in data if c['id'] == client_id), None)
    if not client:
        return "לקוח לא נמצא", 404
    
    # בדיקת הרשאות גישה ללקוח
    if not can_user_access_client(current_user.id, user_role, client):
        return "גישה חסומה - אין לך הרשאה ללקוח זה", 403
    
    # סינון פרויקטים לפי הרשאות
    # מנהל ואדמין רואים הכל
    filtered_projects = []
    if is_manager_or_admin(current_user.id, user_role):
        filtered_projects = client.get('projects', [])
    else:
        # עובד רואה רק פרויקטים משותפים או פרויקטים שהוא יצר
        for project in client.get('projects', []):
            is_shared = project.get('is_shared', False)
            created_by = project.get('created_by', '')
            if is_shared or created_by == current_user.id:
                filtered_projects.append(project)
    
    # צור העתק של הלקוח עם פרויקטים מסוננים
    client_copy = client.copy()
    client_copy['projects'] = filtered_projects
    
    # וידוא שיש רשימת documents
    if 'documents' not in client_copy:
        client_copy['documents'] = []
    
    # Debug - הדפסת המסמכים
    print(f"Client {client_id} documents: {client_copy.get('documents', [])}")
    print(f"Documents count: {len(client_copy.get('documents', []))}")
    
    # עבור assigned_name - אם יש כמה משתמשים, הצג את כולם
    assigned = normalize_assigned_user(client.get('assigned_user', []))
    assigned_names = [users.get(uid, {}).get('name', uid) for uid in assigned if uid]
    assigned_name = ', '.join(assigned_names) if assigned_names else 'לא שויך'
    
    # טעינת היסטוריית אינטראקציות
    activity_logs = load_activity_logs()
    client_activities = [log for log in activity_logs if log.get('client_id') == client_id]
    client_activities.sort(key=lambda x: x.get('timestamp', ''), reverse=True)
    
    # Cache busting עבור לוגו - הוספת timestamp
    logo_cache_bust = int(datetime.now().timestamp())
    
    # Redirect to React client page
    return redirect(f'/app/client/{client_id}')

@app.route('/upload_logo/<client_id>', methods=['POST'])
@login_required
@csrf.exempt
def upload_logo(client_id):
    """Route להעלאת לוגו - תמיכה ב-AJAX וטופס רגיל"""
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or \
              request.headers.get('Accept', '').find('application/json') != -1
    
    try:
        print(f"\n{'='*50}")
        print(f"=== Logo upload started for client {client_id} ===")
        print(f"Request method: {request.method}")
        print(f"Request content type: {request.content_type}")
        print(f"Request files keys: {list(request.files.keys())}")
        print(f"Request form keys: {list(request.form.keys())}")
        print(f"Is AJAX: {is_ajax}")
        
        if 'logo' not in request.files:
            print("ERROR: 'logo' not in request.files")
            if is_ajax:
                return jsonify({'success': False, 'error': 'לא נבחר קובץ'}), 400
            return redirect(request.referrer or url_for('client_page', client_id=client_id))
        
        file = request.files['logo']
        print("DEBUG: Received file", file.filename if file else 'None')
        print(f"File object: {file}")
        print(f"File filename: {file.filename if file else 'None'}")
        
        if not file or not file.filename:
            print("ERROR: No file or filename")
            if is_ajax:
                return jsonify({'success': False, 'error': 'לא נבחר קובץ'}), 400
            return redirect(request.referrer or url_for('client_page', client_id=client_id))
        
        # בדיקה שהקובץ הוא תמונה
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
        filename_parts = file.filename.rsplit('.', 1)
        if len(filename_parts) < 2:
            print(f"ERROR: No extension in filename: {file.filename}")
            if is_ajax:
                return jsonify({'success': False, 'error': 'סוג קובץ לא תקין'}), 400
            return redirect(request.referrer or url_for('client_page', client_id=client_id))
        
        ext = filename_parts[1].lower()
        print(f"File extension: {ext}")
        if ext not in allowed_extensions:
            print(f"ERROR: Extension not allowed: {ext}")
            if is_ajax:
                return jsonify({'success': False, 'error': f'סוג קובץ {ext} לא נתמך. השתמש ב: png, jpg, jpeg, gif, webp'}), 400
            return redirect(request.referrer or url_for('client_page', client_id=client_id))
        
        # יצירת שם קובץ בטוח
        filename = f"logo_{client_id}.{ext}"
        print(f"Target filename: {filename}")
        
        # שימוש ב-app.root_path במקום BASE_DIR
        logos_path = os.path.join(app.root_path, 'static', 'logos')
        if not os.path.exists(logos_path):
            os.makedirs(logos_path, exist_ok=True)
        
        filepath = os.path.join(logos_path, filename)
        print(f"Filepath: {filepath}")
        
        # שמירת הקובץ
        file.save(filepath)
        print(f"File saved to: {filepath}")
        
        # בדיקה שהקובץ נשמר
        if not os.path.exists(filepath):
            print(f"ERROR: File not found after save: {filepath}")
            if is_ajax:
                return jsonify({'success': False, 'error': 'שגיאה בשמירת הקובץ'}), 500
            return redirect(request.referrer or url_for('client_page', client_id=client_id))
        
        print(f"File exists: {os.path.exists(filepath)}")
        print(f"File size: {os.path.getsize(filepath)} bytes")
        
        # עדכון בבסיס הנתונים
        data = load_data()
        client_found = False
        for c in data:
            if c['id'] == client_id:
                c['logo_url'] = filename
                client_found = True
                print(f"Logo saved for client {client_id}: {filename}")
                break
        
        if not client_found:
            print(f"ERROR: Client not found: {client_id}")
            if is_ajax:
                return jsonify({'success': False, 'error': 'לקוח לא נמצא'}), 404
        
        save_data(data)
        
        # Return JSON for AJAX, redirect for form
        if is_ajax:
            return jsonify({
                'success': True,
                'message': 'הלוגו הועלה בהצלחה',
                'logo_url': filename
            })
        
        # Redirect עם cache busting
        base_url = url_for('client_page', client_id=client_id)
        redirect_url = f"{base_url}?logo_updated={int(datetime.now().timestamp())}"
        print(f"Redirecting to: {redirect_url}")
        return redirect(redirect_url)
    
    except Exception as e:
        import traceback
        print(f"EXCEPTION in upload_logo: {str(e)}")
        print(traceback.format_exc())
        if is_ajax:
            return jsonify({'success': False, 'error': str(e)}), 500
        return redirect(request.referrer or url_for('client_page', client_id=client_id))

@app.route('/add_project/<client_id>', methods=['POST'])
@login_required
@csrf.exempt
def add_project(client_id):
    try:
        data = load_data()
        project_id = str(uuid.uuid4())
        
        # תמיכה ב-JSON וב-form
        if request.is_json:
            project_title = request.json.get('title')
        else:
            project_title = request.form.get('title')
        
        if not project_title:
            wants_json = request.headers.get('Accept', '').find('application/json') != -1 or \
                        request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            if wants_json:
                return jsonify({'status': 'error', 'error': 'נא להזין שם פרויקט'}), 400
            return "נא להזין שם פרויקט", 400
        
        for c in data:
            if c['id'] == client_id:
                try:
                    project_number = get_next_project_number(c)
                except Exception as e:
                    print(f"Error generating project number: {e}")
                    import traceback
                    traceback.print_exc()
                    # נשתמש במספר ברירת מחדל
                    client_num = c.get('client_number', 1)
                    try:
                        client_num = int(client_num)
                    except (ValueError, TypeError):
                        client_num = 1
                    project_number = f"{client_num:03d}0001"
                
                # קבל is_shared מ-request (JSON או form)
                is_shared = False
                if request.is_json:
                    is_shared = request.json.get('is_shared', False)
                else:
                    is_shared = request.form.get('is_shared') == 'true' or request.form.get('is_shared') == 'on'
                
                c.setdefault('projects', []).append({
                    'id': project_id, 
                    'title': project_title, 
                    'tasks': [],
                    'project_number': project_number,
                    'is_shared': is_shared,
                    'created_by': current_user.id
                })
                save_data(data)
                
                # בדיקה אם זה AJAX request (בדרך כלל יש Accept: application/json או X-Requested-With)
                wants_json = request.headers.get('Accept', '').find('application/json') != -1 or \
                            request.headers.get('X-Requested-With') == 'XMLHttpRequest'
                
                if wants_json:
                    return jsonify({
                        'status': 'success',
                        'data': {
                            'project': {
                                'id': project_id,
                                'title': project_title,
                                'project_number': project_number,
                                'tasks': []
                            }
                        }
                    })
                
                return redirect(request.referrer or url_for('client_page', client_id=client_id))
        
        # אם הלקוח לא נמצא, נחזיר שגיאה בפורמט אחיד
        wants_json = request.headers.get('Accept', '').find('application/json') != -1 or \
                    request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if wants_json:
            return jsonify({'status': 'error', 'error': 'לקוח לא נמצא'}), 404
        return "לקוח לא נמצא", 404
    except Exception as e:
        print(f"Error in add_project: {e}")
        import traceback
        traceback.print_exc()
        wants_json = request.headers.get('Accept', '').find('application/json') != -1 or \
                    request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if wants_json:
            return jsonify({'status': 'error', 'error': str(e)}), 500
        return f"שגיאה: {str(e)}", 500

@app.route('/add_task/<client_id>/<project_id>', methods=['POST'])
@login_required
@csrf.exempt
def add_task(client_id, project_id):
    try:
        data = load_data()
        # Support both form and JSON
        if request.is_json:
            req_data = request.get_json()
            title = req_data.get('title')
            status = req_data.get('status', 'לביצוע')
            note = req_data.get('note', '')
        else:
            title = request.form.get('title')
            status = request.form.get('status', 'לביצוע')
            note = request.form.get('note', '')
        
        if not title:
            return jsonify({'status': 'error', 'error': 'כותרת משימה נדרשת'}), 400
        
        # קבל שדות חדשים (אם קיימים)
        deadline = None
        priority = req_data.get('priority', 'medium') if request.is_json else request.form.get('priority', 'medium')
        estimated_hours = req_data.get('estimated_hours') if request.is_json else request.form.get('estimated_hours')
        assignee = req_data.get('assignee', '') if request.is_json else request.form.get('assignee', '')
        dependencies = req_data.get('dependencies', []) if request.is_json else request.form.getlist('dependencies') or []
        is_daily_task = (req_data.get('is_daily_task', False) if request.is_json 
                         else request.form.get('is_daily_task', 'false').lower() == 'true')
        
        deadline_str = req_data.get('deadline') if request.is_json else request.form.get('deadline')
        if deadline_str:
            try:
                deadline = datetime.strptime(deadline_str, '%Y-%m-%d').isoformat()
            except:
                pass
        
        task = {
            'id': str(uuid.uuid4()),
            'title': title,
            'status': status,
            'note': note,
            'created_date': datetime.now().strftime('%d/%m/%y'),
            'created_at': datetime.now().isoformat(),
            'done': (status == 'הושלם'),
            'priority': priority,
            'assignee': assignee or current_user.id,
            'dependencies': dependencies if isinstance(dependencies, list) else []
        }
        if deadline:
            task['deadline'] = deadline
        if estimated_hours:
            try:
                task['estimated_hours'] = float(estimated_hours)
            except:
                pass
        
        # משימה יומית - קבע ליום העבודה הקרוב
        if is_daily_task:
            task['is_daily_task'] = True
            next_workday = get_next_workday()
            task['deadline'] = next_workday.isoformat()
        
        if status == 'הושלם':
            task['completed_at'] = datetime.now().isoformat()
            task['actual_hours'] = task.get('estimated_hours', 0)
        
        for c in data:
            if c['id'] == client_id:
                for p in c.get('projects', []):
                    if p['id'] == project_id:
                        try:
                            task_number = get_next_task_number(c, p)
                            task['task_number'] = task_number
                        except Exception as e:
                            print(f"Error generating task number: {e}")
                            # נשתמש במספר ברירת מחדל
                            project_number = p.get('project_number', f"{c.get('client_number', 1):03d}0001")
                            task_number = f"{project_number}001"
                            task['task_number'] = task_number
                        
                        p.setdefault('tasks', []).append(task)
                        save_data(data)
                        return jsonify({
                            'status': 'success',
                            'message': 'המשימה נוספה בהצלחה',
                            'data': {'task': task, 'client': c, 'project': p}
                        })
        return jsonify({'status': 'error', 'error': 'לקוח או פרויקט לא נמצאו'}), 404
    except Exception as e:
        print(f"Error in add_task: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/update_task_status/<client_id>/<project_id>/<task_id>', methods=['POST'])
@login_required
@csrf.exempt  # פטור מ-CSRF כי זה API call מ-JavaScript
def update_task_status(client_id, project_id, task_id):
    try:
        data = load_data()
        # Support both form and JSON
        if request.is_json:
            new_status = request.json.get('status', 'לביצוע')
            new_deadline = request.json.get('deadline', None)
        else:
            new_status = request.form.get('status', 'לביצוע')
            new_deadline = request.form.get('deadline', None)
        
        for c in data:
            if c['id'] == client_id:
                for p in c.get('projects', []):
                    if p['id'] == project_id:
                        for t in p.get('tasks', []):
                            if t['id'] == task_id:
                                # בדיקת תלויות - אם מנסים להשלים משימה
                                if new_status == 'הושלם':
                                    dependencies = t.get('dependencies', [])
                                    if dependencies:
                                        # בדוק אם כל התלויות הושלמו
                                        for dep_id in dependencies:
                                            dep_task = next((dt for dt in p.get('tasks', []) if dt.get('id') == dep_id), None)
                                            if dep_task and dep_task.get('status') != 'הושלם':
                                                dep_title = dep_task.get('title', 'לא ידוע')
                                                return jsonify({
                                                    'status': 'error',
                                                    'error': f'לא ניתן להשלים משימה. יש להשלים קודם את המשימה: "{dep_title}"',
                                                    'blocked_by': dep_id
                                                }), 400
                                
                                old_status = t.get('status', 'לביצוע')
                                t['status'] = new_status
                                t['done'] = (new_status == 'הושלם')
                                
                                # עדכון deadline אם הועבר
                                if new_deadline:
                                    try:
                                        # המרת תאריך לפורמט ISO
                                        if 'T' in new_deadline:
                                            deadline_date = new_deadline.split('T')[0]
                                        else:
                                            deadline_date = new_deadline
                                        t['deadline'] = deadline_date
                                    except Exception as e:
                                        print(f"Error parsing deadline: {e}")
                                
                                # עדכון תאריכים
                                if 'created_at' not in t and t.get('created_date'):
                                    # המרת created_date ל-created_at אם לא קיים
                                    try:
                                        date_str = t['created_date']
                                        if '/' in date_str:
                                            # פורמט dd/mm/yy
                                            parts = date_str.split('/')
                                            if len(parts) == 3:
                                                day, month, year = parts
                                                year = '20' + year if len(year) == 2 else year
                                                created_dt = datetime(int(year), int(month), int(day))
                                                t['created_at'] = created_dt.isoformat()
                                        else:
                                            # פורמט YYYY-MM-DD
                                            created_dt = datetime.strptime(date_str, '%Y-%m-%d')
                                            t['created_at'] = created_dt.isoformat()
                                    except:
                                        t['created_at'] = datetime.now().isoformat()
                                elif 'created_at' not in t:
                                    t['created_at'] = datetime.now().isoformat()
                                
                                if new_status == 'הושלם' and old_status != 'הושלם':
                                    t['completed_at'] = datetime.now().isoformat()
                                    
                                    # משימה יומית - העבר ליום העבודה הבא
                                    if t.get('is_daily_task'):
                                        # מצא את יום העבודה הבא (מחר או ראשון אם היום חמישי)
                                        tomorrow = datetime.now() + timedelta(days=1)
                                        next_workday = get_next_workday(tomorrow)
                                        t['deadline'] = next_workday.isoformat()
                                        t['status'] = 'לביצוע'  # החזר למצב לביצוע
                                        t['done'] = False
                                        del t['completed_at']  # הסר את תאריך ההשלמה
                                
                                # אם עבר למצב הושלם, הוסף חיוב של מחיאות כפיים (רק למשימות לא יומיות)
                                if new_status == 'הושלם' and old_status != 'הושלם' and not t.get('is_daily_task'):
                                    if 'extra_charges' not in c:
                                        c['extra_charges'] = []
                                    charge_number = get_next_charge_number(c)
                                    c['extra_charges'].append({
                                        'id': str(uuid.uuid4()),
                                        'title': 'מחיאות כפיים - ' + t.get('title', 'משימה הושלמה'),
                                        'amount': 50,
                                        'date': datetime.now().strftime('%d/%m/%y'),
                                        'charge_number': charge_number,
                                        'completed': False
                                    })
                                save_data(data)
                                return jsonify({
                                    'status': 'success',
                                    'message': 'סטטוס המשימה עודכן',
                                    'data': {'task': t, 'client': c}
                                })
        return jsonify({'status': 'error', 'error': 'משימה לא נמצאה'}), 404
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/get_task/<client_id>/<project_id>/<task_id>')
@login_required
def get_task(client_id, project_id, task_id):
    """Route להחזרת פרטי משימה"""
    try:
        data = load_data()
        for c in data:
            if c['id'] == client_id:
                for p in c.get('projects', []):
                    if p['id'] == project_id:
                        for t in p.get('tasks', []):
                            if t['id'] == task_id:
                                return jsonify({
                                    'success': True,
                                    'task': {
                                        'id': t.get('id'),
                                        'title': t.get('title'),
                                        'status': t.get('status'),
                                        'note': t.get('note', ''),
                                        'manager_note': t.get('manager_note', '')
                                    }
                                })
        return jsonify({'success': False, 'error': 'משימה לא נמצאה'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/task/update_dates/<client_id>/<project_id>/<task_id>', methods=['POST'])
@login_required
def update_task_dates(client_id, project_id, task_id):
    """עדכון תאריכי משימה (מ-Gantt)"""
    try:
        data = load_data()
        req_data = request.get_json()
        start_date = req_data.get('start_date')
        deadline = req_data.get('deadline')
        
        if not start_date or not deadline:
            return jsonify({'status': 'error', 'error': 'תאריכים נדרשים'}), 400
        
        for c in data:
            if c['id'] == client_id:
                for p in c.get('projects', []):
                    if p['id'] == project_id:
                        for t in p.get('tasks', []):
                            if t['id'] == task_id:
                                # עדכן תאריכים
                                if start_date:
                                    t['created_at'] = datetime.fromisoformat(start_date.split('T')[0] + 'T00:00:00').isoformat() if 'T' not in start_date else datetime.fromisoformat(start_date).isoformat()
                                if deadline:
                                    t['deadline'] = deadline.split('T')[0] if 'T' in deadline else deadline
                                
                                save_data(data)
                                return jsonify({'status': 'success', 'message': 'תאריכים עודכנו בהצלחה'})
        
        return jsonify({'status': 'error', 'error': 'משימה לא נמצאה'}), 404
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/update_task/<client_id>/<project_id>/<task_id>', methods=['POST'])
@login_required
def update_task(client_id, project_id, task_id):
    try:
        data = load_data()
        
        # תמיכה ב-JSON וב-form
        if request.is_json:
            req_data = request.get_json()
            status = req_data.get('status')
            notes = req_data.get('notes')
        else:
            status = request.form.get('status')
            notes = request.form.get('notes')
        
        for c in data:
            if c['id'] == client_id:
                for p in c.get('projects', []):
                    if p['id'] == project_id:
                        for t in p.get('tasks', []):
                            if t['id'] == task_id:
                                if status:
                                    old_status = t.get('status', 'pending')
                                    t['status'] = status
                                    t['done'] = (status == 'completed' or status == 'הושלם')
                                    
                                    # עדכון תאריכים
                                    if 'created_at' not in t and t.get('created_date'):
                                        try:
                                            date_str = t['created_date']
                                            if '/' in date_str:
                                                parts = date_str.split('/')
                                                if len(parts) == 3:
                                                    day, month, year = parts
                                                    year = '20' + year if len(year) == 2 else year
                                                    created_dt = datetime(int(year), int(month), int(day))
                                                    t['created_at'] = created_dt.isoformat()
                                            else:
                                                created_dt = datetime.strptime(date_str, '%Y-%m-%d')
                                                t['created_at'] = created_dt.isoformat()
                                        except:
                                            t['created_at'] = datetime.now().isoformat()
                                    elif 'created_at' not in t:
                                        t['created_at'] = datetime.now().isoformat()
                                    
                                    if (status == 'completed' or status == 'הושלם') and old_status not in ['completed', 'הושלם']:
                                        t['completed_at'] = datetime.now().isoformat()
                                    
                                    # אם עבר למצב הושלם, הוסף חיוב של מחיאות כפיים
                                    if (status == 'completed' or status == 'הושלם') and old_status not in ['completed', 'הושלם']:
                                        if 'extra_charges' not in c:
                                            c['extra_charges'] = []
                                        charge_number = get_next_charge_number(c)
                                        c['extra_charges'].append({
                                            'id': str(uuid.uuid4()),
                                            'title': '👏 ' + t.get('title', 'משימה הושלמה'),
                                            'amount': 0,
                                            'date': datetime.now().strftime('%Y-%m-%d'),
                                            'charge_number': charge_number,
                                            'completed': False
                                        })
                                
                                if notes is not None:
                                    t['note'] = notes
                                
                                save_data(data)
                                
                                # בדיקה אם זה AJAX request
                                wants_json = request.headers.get('Accept', '').find('application/json') != -1 or \
                                            request.headers.get('X-Requested-With') == 'XMLHttpRequest'
                                
                                if wants_json:
                                    return jsonify({
                                        'status': 'success',
                                        'message': 'המשימה עודכנה בהצלחה'
                                    })
                                
                                return redirect(request.referrer or url_for('client_page', client_id=client_id))
        
        wants_json = request.headers.get('Accept', '').find('application/json') != -1 or \
                    request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if wants_json:
            return jsonify({'status': 'error', 'error': 'משימה לא נמצאה'}), 404
        return "משימה לא נמצאה", 404
    except Exception as e:
        print(f"Error in update_task: {e}")
        import traceback
        traceback.print_exc()
        wants_json = request.headers.get('Accept', '').find('application/json') != -1 or \
                    request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if wants_json:
            return jsonify({'status': 'error', 'error': str(e)}), 500
        return f"שגיאה: {str(e)}", 500

@app.route('/update_task_note/<client_id>/<project_id>/<task_id>', methods=['POST'])
@login_required
def update_task_note(client_id, project_id, task_id):
    try:
        data = load_data()
        note = request.json.get('note', '')
        for c in data:
            if c['id'] == client_id:
                for p in c.get('projects', []):
                    if p['id'] == project_id:
                        for t in p.get('tasks', []):
                            if t['id'] == task_id:
                                t['note'] = note
                                save_data(data)
                                return jsonify({'success': True})
        return jsonify({'success': False, 'error': 'משימה לא נמצאה'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/delete_project/<client_id>/<project_id>', methods=['POST'])
@login_required
def delete_project(client_id, project_id):
    try:
        print(f"Delete project called: client_id={client_id}, project_id={project_id}")
        data = load_data()
        for c in data:
            if c['id'] == client_id:
                # שמור את הפרויקט בארכיון במקום למחוק אותו
                project_to_delete = None
                for p in c.get('projects', []):
                    if p['id'] == project_id:
                        project_to_delete = p
                        break
                
                if project_to_delete:
                    # הוסף לארכיון
                    if 'archived_projects' not in c:
                        c['archived_projects'] = []
                    project_to_delete['archived_date'] = datetime.now().strftime('%d/%m/%y %H:%M')
                    c['archived_projects'].append(project_to_delete)
                
                # מחק מהרשימה הפעילה
                c['projects'] = [p for p in c.get('projects', []) if p['id'] != project_id]
                save_data(data)
                
                # בדיקה אם זה AJAX request
                wants_json = request.headers.get('Accept', '').find('application/json') != -1 or \
                            request.headers.get('X-Requested-With') == 'XMLHttpRequest'
                
                if wants_json:
                    return jsonify({
                        'status': 'success',
                        'message': 'הפרויקט נמחק ונשמר בארכיון'
                    })
                
                return redirect(request.referrer or url_for('client_page', client_id=client_id))
        return jsonify({'status': 'error', 'error': 'פרויקט לא נמצא'}), 404
    except Exception as e:
        print(f"Error in delete_project: {e}")
        import traceback
        traceback.print_exc()
        wants_json = request.headers.get('Accept', '').find('application/json') != -1 or \
                    request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if wants_json:
            return jsonify({'status': 'error', 'error': str(e)}), 500
        return f"שגיאה: {str(e)}", 500

@app.route('/delete_task/<client_id>/<project_id>/<task_id>', methods=['POST'])
@login_required
@csrf.exempt  # פטור מ-CSRF כי זה API call מ-JavaScript
def delete_task(client_id, project_id, task_id):
    try:
        data = load_data()
        task_found = False
        task_title = None
        
        for c in data:
            if c['id'] != client_id:
                continue
            for p in c.get('projects', []):
                if p['id'] != project_id:
                    continue
                tasks = p.get('tasks', [])
                
                # Find the task to be deleted to get its title and check if it exists
                task_to_delete = next((t for t in tasks if t['id'] == task_id), None)
                if not task_to_delete:
                    # Task already missing; treat as success to keep API simple
                    return jsonify({'success': True})
                
                task_found = True
                task_title = task_to_delete.get('title', 'לא ידוע')
                
                # Remove the task from dependencies of other tasks in the same project
                for t in tasks:
                    if t['id'] != task_id:
                        dependencies = t.get('dependencies', [])
                        if isinstance(dependencies, list) and task_id in dependencies:
                            dependencies.remove(task_id)
                            t['dependencies'] = dependencies
                
                # Cancel any active time tracking sessions for this task
                try:
                    time_data = load_time_tracking()
                    active_sessions = time_data.get('active_sessions', {})
                    sessions_to_cancel = []
                    
                    for user_id, session in active_sessions.items():
                        if (session.get('client_id') == client_id and 
                            session.get('project_id') == project_id and 
                            session.get('task_id') == task_id):
                            sessions_to_cancel.append(user_id)
                    
                    for user_id in sessions_to_cancel:
                        del active_sessions[user_id]
                        print(f"Cancelled active time tracking session for user {user_id} on task {task_id}")
                    
                    if sessions_to_cancel:
                        save_time_tracking(time_data)
                except Exception as time_error:
                    # Don't fail task deletion if time tracking cleanup fails
                    print(f"Warning: Could not clean up time tracking for task {task_id}: {time_error}")
                
                # Remove the task itself
                filtered_tasks = [t for t in tasks if t['id'] != task_id]
                p['tasks'] = filtered_tasks
                
                save_data(data)
                print(f"Task '{task_title}' (ID: {task_id}) deleted successfully from project {project_id}")
                return jsonify({'success': True, 'message': f'המשימה "{task_title}" נמחקה בהצלחה'})
        
        if not task_found:
            return jsonify({'success': False, 'error': 'לקוח או פרויקט לא נמצאו'}), 404
        return jsonify({'success': False, 'error': 'משימה לא נמצאה'}), 404
    except Exception as e:
        print(f"Error in delete_task: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/upload_document/<client_id>', methods=['POST'])
@login_required
def upload_document(client_id):
    try:
        if 'document' not in request.files:
            return redirect(request.referrer or url_for('client_page', client_id=client_id))
        
        file = request.files['document']
        doc_name = request.form.get('doc_name', '')  # שם המסמך האופציונלי מהטופס
        
        if file and file.filename != '':
            # יצירת שם קובץ בטוח ומניעת דריסה על ידי הוספת מזהה ייחודי
            filename = secure_filename(f"{uuid.uuid4().hex}_{file.filename}")
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            
            # וידוא שהתיקייה קיימת
            if not os.path.exists(app.config['UPLOAD_FOLDER']):
                os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
            
            # שמירת הקובץ
            file.save(filepath)
            
            # עדכון הנתונים ב-JSON
            data = load_data()
            for c in data:
                if c['id'] == client_id:
                    # וודא שקיימת רשימת מסמכים
                    if 'documents' not in c:
                        c['documents'] = []
                    
                    c['documents'].append({
                        'id': str(uuid.uuid4()),
                        'display_name': doc_name if doc_name else file.filename,
                        'file_path': filename,
                        'upload_date': datetime.now().strftime("%d/%m/%y %H:%M")
                    })
                    save_data(data)
                    print(f"Document saved: {filename} for client {client_id}")  # Debug
                    break
        
        return redirect(request.referrer or url_for('client_page', client_id=client_id))
    
    except Exception as e:
        import traceback
        return jsonify({'success': False, 'error': f'שגיאה: {str(e)}'}), 500

@app.route('/add_activity/<client_id>', methods=['POST'])
@login_required
def add_activity(client_id):
    """הוספת פעילות חדשה (שיחה, פגישה, הערה)"""
    try:
        activity_type = request.form.get('activity_type')  # 'call', 'meeting', 'note', 'email'
        title = request.form.get('title', '')
        content = request.form.get('content', '')
        duration = request.form.get('duration', '')  # לשיחות/פגישות
        participants = request.form.get('participants', '')  # לוג טלפון/פגישה
        follow_up_required = request.form.get('follow_up_required') == 'true'
        follow_up_date = request.form.get('follow_up_date', '')
        tags = request.form.get('tags', '')
        
        logs = load_activity_logs()
        new_log = {
            'id': str(uuid.uuid4()),
            'client_id': client_id,
            'user_id': current_user.id,
            'activity_type': activity_type,
            'timestamp': datetime.now().isoformat(),
            'title': title,
            'content': content,
            'duration': duration,
            'participants': participants,
            'follow_up_required': follow_up_required,
            'follow_up_date': follow_up_date,
            'tags': tags
        }
        logs.append(new_log)
        save_activity_logs(logs)
        
        if request.is_json or request.headers.get('Accept') == 'application/json':
            return jsonify({'success': True, 'activity': new_log})
        return redirect(request.referrer or url_for('client_page', client_id=client_id))
    except Exception as e:
        return f"שגיאה בהוספת פעילות: {str(e)}", 500

@app.route('/delete_activity/<activity_id>', methods=['POST'])
@login_required
def delete_activity(activity_id):
    """מחיקת פעילות"""
    try:
        logs = load_activity_logs()
        logs = [log for log in logs if log.get('id') != activity_id]
        save_activity_logs(logs)
        
        if request.is_json or request.headers.get('Accept') == 'application/json':
            return jsonify({'success': True})
        return redirect(request.referrer or url_for('client_page', client_id=request.form.get('client_id', '')))
    except Exception as e:
        return f"שגיאה במחיקת פעילות: {str(e)}", 500

@app.route('/add_contact/<client_id>', methods=['POST'])
@login_required
def add_contact(client_id):
    """הוספת איש קשר ללקוח"""
    try:
        data = load_data()
        for c in data:
            if c['id'] == client_id:
                if 'contacts' not in c:
                    c['contacts'] = []
                
                c['contacts'].append({
                    'id': str(uuid.uuid4()),
                    'name': request.form.get('name', ''),
                    'phone': request.form.get('phone', ''),
                    'email': request.form.get('email', ''),
                    'created_date': datetime.now().strftime("%d/%m/%y %H:%M")
                })
                save_data(data)
                break
        return redirect(request.referrer or url_for('client_page', client_id=client_id))
    except Exception as e:
        return f"שגיאה בהוספת איש קשר: {str(e)}", 500

@app.route('/delete_contact/<client_id>/<contact_id>', methods=['POST'])
@login_required
def delete_contact(client_id, contact_id):
    """מחיקת איש קשר משל לקוח"""
    try:
        data = load_data()
        for c in data:
            if c['id'] == client_id:
                if 'contacts' in c:
                    c['contacts'] = [contact for contact in c['contacts'] if contact.get('id') != contact_id]
                    save_data(data)
                break
        return redirect(request.referrer or url_for('client_page', client_id=client_id))
    except Exception as e:
        return f"שגיאה במחיקת איש קשר: {str(e)}", 500

@app.route('/add_client', methods=['POST'])
@login_required
@csrf.exempt  # פטור מ-CSRF כי זה API call מ-JavaScript
def add_client():
    """הוספת לקוח חדש"""
    try:
        user_role = get_user_role(current_user.id)
        if not is_manager_or_admin(current_user.id, user_role):
            return jsonify({'success': False, 'error': 'גישה חסומה - רק מנהלים יכולים להוסיף לקוחות'}), 403
        
        data = load_data()
        client_number = get_next_client_number()
        
        name = request.form.get('name', '').strip()
        
        if not name:
            return jsonify({'success': False, 'error': 'שם לקוח חסר'}), 400
        
        # Create new client with full structure matching existing clients
        new_client = {
            'id': str(uuid.uuid4()),
            'name': name,
            'client_number': client_number,
            'retainer': 0,
            'extra_charges': [],
            'projects': [],
            'assigned_user': [current_user.id],  # Assign to current user
            'files': [],
            'contacts': [],
            'documents': [],
            'calculated_extra': 0,
            'calculated_retainer': 0,
            'calculated_total': 0,
            'calculated_open_charges': 0,
            'calculated_monthly_revenue': 0,
            'archived': False
        }
        
        data.append(new_client)
        save_data(data)
        
        return jsonify({'success': True, 'client_id': new_client['id']})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/delete_document/<client_id>/<doc_id>', methods=['POST'])
@login_required
def delete_document(client_id, doc_id):
    try:
        print(f"DEBUG: delete_document called with client_id={client_id}, doc_id={doc_id}")
        data = load_data()
        for c in data:
            if c['id'] == client_id:
                documents = c.get('documents', [])
                print(f"DEBUG: Found client, documents count: {len(documents)}")
                print(f"DEBUG: Looking for doc_id: {doc_id} (type: {type(doc_id)})")
                
                # מצא את המסמך - השוואה כפי string
                doc_to_remove = None
                for doc in documents:
                    doc_id_str = str(doc.get('id', ''))
                    print(f"DEBUG: Comparing doc id: {doc_id_str} (type: {type(doc_id_str)}) with {doc_id}")
                    if doc_id_str == str(doc_id):
                        doc_to_remove = doc
                        break
                
                if doc_to_remove:
                    print(f"DEBUG: Found document to delete: {doc_to_remove}")
                    # מחיקת הקובץ הפיזי - תמיכה גם במבנה הישן וגם החדש
                    file_path = doc_to_remove.get('file_path') or doc_to_remove.get('filename', '')
                    print(f"DEBUG: File path: {file_path}")
                    
                    if file_path:
                        # נסה את הנתיב החדש (UPLOAD_FOLDER)
                        filepath_new = os.path.join(app.config['UPLOAD_FOLDER'], file_path)
                        print(f"DEBUG: Trying new path: {filepath_new}")
                        if os.path.exists(filepath_new):
                            os.remove(filepath_new)
                            print(f"DEBUG: File deleted from new path")
                        else:
                            # נסה את הנתיב הישן (DOCUMENTS_FOLDER/client_id)
                            filepath_old = os.path.join(DOCUMENTS_FOLDER, client_id, file_path)
                            print(f"DEBUG: Trying old path: {filepath_old}")
                            if os.path.exists(filepath_old):
                                os.remove(filepath_old)
                                print(f"DEBUG: File deleted from old path")
                            else:
                                print(f"DEBUG: Warning: File not found in either path")
                    
                    # מחיקת הרשומה מהנתונים
                    documents.remove(doc_to_remove)
                    c['documents'] = documents
                    save_data(data)
                    print(f"DEBUG: Document removed from database, returning success")
                    return jsonify({'status': 'success', 'message': 'המסמך נמחק בהצלחה'})
                else:
                    print(f"DEBUG: Document not found in list")
                    return "מסמך לא נמצא", 404
        print(f"DEBUG: Client not found")
        return "לקוח לא נמצא", 404
    except Exception as e:
        import traceback
        error_msg = f"שגיאה: {str(e)}\n{traceback.format_exc()}"
        print(f"DEBUG: Exception in delete_document: {error_msg}")
        return error_msg, 500

@app.route('/download_doc/<filename>')
@login_required
def download_doc(filename):
    """Route להורדת/צפייה במסמך"""
    try:
        # וידוא שהתיקייה קיימת
        upload_folder = app.config.get('UPLOAD_FOLDER')
        if not upload_folder or not os.path.exists(upload_folder):
            print(f"Upload folder not found: {upload_folder}")  # Debug
            return "תיקיית מסמכים לא נמצאה", 404
        
        # בדיקה שהקובץ קיים
        filepath = os.path.join(upload_folder, filename)
        print(f"Looking for file: {filepath}")  # Debug
        if not os.path.exists(filepath):
            print(f"File not found: {filepath}")  # Debug
            # נסה גם עם נתיב אבסולוטי
            filepath_abs = os.path.abspath(filepath)
            if not os.path.exists(filepath_abs):
                return f"קובץ לא נמצא: {filename}", 404
            filepath = filepath_abs
            upload_folder = os.path.dirname(filepath_abs)
        
        print(f"Serving file from: {upload_folder}, filename: {filename}")  # Debug
        return send_from_directory(upload_folder, filename, as_attachment=False)
    except Exception as e:
        import traceback
        print(f"Error in download_doc: {str(e)}")  # Debug
        print(traceback.format_exc())  # Debug
        return f"שגיאה: {str(e)}", 500

@app.route('/static/documents/<client_id>/<filename>')
@login_required
def serve_document(client_id, filename):
    """Route לשרת מסמכים (legacy - תמיכה במבנה הישן)"""
    return send_from_directory(os.path.join(DOCUMENTS_FOLDER, client_id), filename)

@app.route('/static/logos/<filename>')
def serve_logo(filename):
    """Route לשרת לוגואים"""
    return send_from_directory(LOGOS_FOLDER, filename)

@app.route('/api/finance')
@login_required
def api_finance():
    """API endpoint להחזרת נתוני כספים"""
    try:
        user_role = get_user_role(current_user.id)
        if not check_permission('/finance', user_role):
            return jsonify({'success': False, 'error': 'גישה חסומה'}), 403
        
        clients = load_data()
        clients = filter_active_clients(clients)
        current_month = datetime.now().strftime('%m')
        current_year = datetime.now().strftime('%Y')
        
        selected_month = request.args.get('month', '')
        
        total_open_charges = 0
        total_monthly_revenue = 0
        
        clients_data = []
        for c in clients:
            extra_charges = c.get('extra_charges', [])
            for ch in extra_charges:
                if 'completed' not in ch:
                    ch['completed'] = False
                if 'our_cost' not in ch:
                    ch['our_cost'] = 0
            
            calculated_extra = sum(ch.get('amount', 0) for ch in extra_charges)
            calculated_retainer = c.get('retainer', 0)
            calculated_total = calculated_retainer + calculated_extra
            calculated_open_charges = sum(
                ch.get('amount', 0) for ch in extra_charges 
                if not ch.get('completed', False)
            )
            total_open_charges += calculated_open_charges
            
            monthly_revenue = 0
            for ch in extra_charges:
                charge_date = ch.get('date', '')
                if charge_date:
                    date_parts = charge_date.split('/')
                    if len(date_parts) >= 3:
                        month = date_parts[1].zfill(2)
                        year_str = date_parts[2]
                        if len(year_str) == 2:
                            year = '20' + year_str
                        else:
                            year = year_str
                        
                        filter_month = selected_month if selected_month else current_month
                        filter_year = current_year
                        
                        if month == filter_month and year == filter_year:
                            monthly_revenue += ch.get('amount', 0)
            
            total_monthly_revenue += monthly_revenue
            
            clients_data.append({
                'id': c['id'],
                'name': c.get('name', ''),
                'retainer': calculated_retainer,
                'extra_charges': extra_charges,
                'calculated_extra': calculated_extra,
                'calculated_retainer': calculated_retainer,
                'calculated_total': calculated_total,
                'calculated_open_charges': calculated_open_charges,
                'calculated_monthly_revenue': monthly_revenue,
                'retainer_payments': c.get('retainer_payments', {}),
            })
        
        return jsonify({
            'success': True,
            'clients': clients_data,
            'total_open_charges': total_open_charges,
            'total_monthly_revenue': total_monthly_revenue,
            'current_month': current_month,
            'current_year': current_year,
        })
    except Exception as e:
        print(f"Error in api_finance: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/finance')
@login_required
def finance():
    user_role = get_user_role(current_user.id)
    if not check_permission('/finance', user_role):
        return "גישה חסומה - אין לך הרשאה לגשת לדף זה", 403
    clients = load_data()
    # סינון לקוחות מאוישים
    clients = filter_active_clients(clients)
    current_month = datetime.now().strftime('%m')  # חודש נוכחי (01-12)
    current_year = datetime.now().strftime('%Y')   # שנה נוכחית
    
    # Calculate totals for each client
    total_open_charges = 0
    total_monthly_revenue = 0
    
    for c in clients:
        extra_charges = c.get('extra_charges', [])
        # לוודא שלכל חיוב יש completed (ברירת מחדל: False) ו-our_cost
        for ch in extra_charges:
            if 'completed' not in ch:
                ch['completed'] = False
            if 'our_cost' not in ch:
                ch['our_cost'] = 0
        
        c['calculated_extra'] = sum(ch.get('amount', 0) for ch in extra_charges)
        c['calculated_retainer'] = c.get('retainer', 0)
        c['calculated_total'] = c['calculated_retainer'] + c['calculated_extra']
        
        # חישוב חיובים פתוחים (לא הושלמו)
        c['calculated_open_charges'] = sum(
            ch.get('amount', 0) for ch in extra_charges 
            if not ch.get('completed', False)
        )
        total_open_charges += c['calculated_open_charges']
        
        # חישוב הכנסות לחודש הנוכחי (כל החיובים שנכנסו החודש)
        monthly_revenue = 0
        for ch in extra_charges:
            charge_date = ch.get('date', '')
            if charge_date:
                date_parts = charge_date.split('/')
                if len(date_parts) >= 3:
                    month = date_parts[1].zfill(2)
                    year_str = date_parts[2]
                    # אם השנה היא 2 ספרות, הוסף 2000
                    if len(year_str) == 2:
                        year = '20' + year_str
                    else:
                        year = year_str
                    
                    if month == current_month and year == current_year:
                        monthly_revenue += ch.get('amount', 0)
        
        c['calculated_monthly_revenue'] = monthly_revenue
        total_monthly_revenue += monthly_revenue
    
    # שמירה חזרה (אם הוספנו completed לחיובים ישנים)
    save_data(clients)
    
    # מיון לפי סה"כ בסדר יורד (הגבוה ביותר ראשון)
    clients.sort(key=lambda x: x.get('calculated_total', 0), reverse=True)
    
    # Redirect to React finance page
    return redirect('/app/finance')

@app.route('/update_finance/<client_id>', methods=['POST'])
@login_required
@csrf.exempt  # פטור מ-CSRF כי זה API call מ-JavaScript
def update_finance(client_id):
    action, data = request.form.get('action'), load_data()
    wants_json = request.headers.get('Accept', '').find('application/json') != -1 or \
                request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    for c in data:
        if c['id'] == client_id:
            if action == 'retainer': c['retainer'] = int(request.form.get('amount', 0))
            elif action == 'extra':
                charge_number = get_next_charge_number(c)
                our_cost = float(request.form.get('our_cost', 0) or 0)
                description = request.form.get('description', '')
                new_charge = {
                    'id': str(uuid.uuid4()),
                    'title': request.form.get('title'),
                    'description': description,
                    'amount': int(request.form.get('amount', 0)),
                    'our_cost': our_cost,
                    'date': datetime.now().strftime("%d/%m/%y"),
                    'completed': False,
                    'charge_number': charge_number
                }
                c.setdefault('extra_charges', []).append(new_charge)
                
                # Send email notification for new charge
                try:
                    send_charge_notification_email(c.get('name', ''), new_charge)
                except Exception as e:
                    print(f"[WARNING] Failed to send charge notification email: {e}")
    save_data(data)
    
    if wants_json:
        return jsonify({'success': True, 'message': 'עודכן בהצלחה'})
    return redirect(request.referrer)

@app.route('/generate_invoice/<client_id>')
@login_required
def generate_invoice(client_id):
    """Route ליצירת דו"ח חיוב ב-Excel"""
    try:
        data = load_data()
        client = next((c for c in data if c['id'] == client_id), None)
        if not client:
            return "לקוח לא נמצא", 404
        
        # קבלת החודש מהפילטר (אם יש)
        selected_month = request.args.get('month', '')
        
        # חישוב חיובים לפי החודש
        all_charges = client.get('extra_charges', [])
        if selected_month:
            # סינון חיובים לפי חודש
            filtered_charges = []
            for charge in all_charges:
                charge_date = charge.get('date', '')
                if charge_date:
                    date_parts = charge_date.split('/')
                    if len(date_parts) >= 2:
                        month_str = date_parts[1]
                        # אם זה בפורמט yy או yyyy, קח רק את החלק הראשון
                        if len(month_str) > 2:
                            month_str = month_str[:2]
                        month = month_str.zfill(2)
                        if month == selected_month:
                            filtered_charges.append(charge)
        else:
            filtered_charges = all_charges
        
        # חישוב סכומים
        retainer = client.get('retainer', 0)
        extra_total = sum(ch.get('amount', 0) for ch in filtered_charges)
        total = retainer + extra_total
        
        # יצירת Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "דו\"ח חיוב"
        
        # הגדרת כותרת
        ws['A1'] = f"דו\"ח חיוב - {client.get('name', 'לקוח')}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='right', vertical='center')
        ws.merge_cells('A1:D1')
        
        # חודש (אם נבחר)
        if selected_month:
            month_names = {
                '01': 'ינואר', '02': 'פברואר', '03': 'מרץ', '04': 'אפריל',
                '05': 'מאי', '06': 'יוני', '07': 'יולי', '08': 'אוגוסט',
                '09': 'ספטמבר', '10': 'אוקטובר', '11': 'נובמבר', '12': 'דצמבר'
            }
            ws['A2'] = f"חודש: {month_names.get(selected_month, selected_month)}"
            ws['A2'].font = Font(bold=True, size=12)
            ws['A2'].alignment = Alignment(horizontal='right', vertical='center')
            ws.merge_cells('A2:D2')
            start_row = 4
        else:
            start_row = 3
        
        # כותרות טבלה
        headers = ['תיאור', 'תאריך', 'סכום']
        header_row = start_row
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='0073ea', end_color='0073ea', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # ריטיינר
        data_row = start_row + 1
        ws.cell(row=data_row, column=1, value='ריטיינר')
        ws.cell(row=data_row, column=2, value='')
        ws.cell(row=data_row, column=3, value=f'₪{retainer}')
        
        # חיובים נוספים
        for charge in filtered_charges:
            data_row += 1
            ws.cell(row=data_row, column=1, value=charge.get('title', 'ללא תיאור'))
            ws.cell(row=data_row, column=2, value=charge.get('date', 'ללא תאריך'))
            ws.cell(row=data_row, column=3, value=f'₪{charge.get("amount", 0)}')
        
        # עיצוב שורות הנתונים
        for row in range(start_row + 1, data_row + 1):
            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='right' if col == 1 else 'center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # שורת סיכום
        summary_row = data_row + 2
        ws.cell(row=summary_row, column=1, value='סה"כ חיוב').font = Font(bold=True, size=12)
        ws.cell(row=summary_row, column=2, value='')
        ws.cell(row=summary_row, column=3, value=f'₪{total}').font = Font(bold=True, size=12)
        
        # עיצוב שורת סיכום
        for col in range(1, 4):
            cell = ws.cell(row=summary_row, column=col)
            cell.fill = PatternFill(start_color='e1e6ff', end_color='e1e6ff', fill_type='solid')
            cell.alignment = Alignment(horizontal='right' if col == 1 else 'center', vertical='center')
            cell.border = Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium')
            )
        
        # התאמת רוחב עמודות
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        
        # שמירת הקובץ
        filename = f"invoice_{client_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(app.root_path, 'static', filename)
        wb.save(filepath)
        
        # שליחת הקובץ
        return send_file(filepath, as_attachment=True, download_name=f"דוח_חיוב_{client.get('name', 'לקוח')}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
    except Exception as e:
        import traceback
        print(f"Error generating invoice: {e}")
        print(traceback.format_exc())
        return f"שגיאה ביצירת הדו\"ח: {str(e)}", 500

@app.route('/toggle_charge_status/<client_id>/<charge_id>', methods=['POST'])
@login_required
@csrf.exempt
def toggle_charge_status(client_id, charge_id):
    """עדכון סטטוס חיוב (completed/paid)"""
    try:
        data = load_data()
        for c in data:
            if c['id'] == client_id:
                for charge in c.get('extra_charges', []):
                    if charge.get('id') == charge_id:
                        current_status = charge.get('paid', False) or charge.get('completed', False)
                        new_status = not current_status
                        charge['completed'] = new_status
                        charge['paid'] = new_status
                        save_data(data)
                        return jsonify({'success': True, 'completed': new_status, 'paid': new_status})
        return jsonify({'success': False, 'error': 'חיוב לא נמצא'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/toggle_retainer_status/<client_id>/<month>', methods=['POST'])
@login_required
@csrf.exempt
def toggle_retainer_status(client_id, month):
    """עדכון סטטוס תשלום ריטיינר חודשי"""
    try:
        data = load_data()
        for c in data:
            if c['id'] == client_id:
                if 'retainer_payments' not in c:
                    c['retainer_payments'] = {}
                
                current_status = c['retainer_payments'].get(month, False)
                new_status = not current_status
                c['retainer_payments'][month] = new_status
                save_data(data)
                return jsonify({'success': True, 'paid': new_status, 'month': month})
        return jsonify({'success': False, 'error': 'לקוח לא נמצא'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/update_charge_our_cost/<client_id>/<charge_id>', methods=['POST'])
@login_required
def update_charge_our_cost(client_id, charge_id):
    """עדכון עלות פנימית של חיוב"""
    try:
        data = load_data()
        if request.is_json:
            our_cost = float(request.json.get('our_cost', 0) or 0)
        else:
            our_cost = float(request.form.get('our_cost', 0) or 0)
        
        for c in data:
            if c['id'] == client_id:
                for charge in c.get('extra_charges', []):
                    if charge.get('id') == charge_id:
                        charge['our_cost'] = our_cost
                        save_data(data)
                        return jsonify({'success': True})
        
        return jsonify({'success': False, 'error': 'חיוב לא נמצא'}), 404
    except Exception as e:
        print(f"Error updating charge our_cost: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/delete_charge/<client_id>/<charge_id>', methods=['POST'])
@login_required
@csrf.exempt  # פטור מ-CSRF כי זה API call מ-JavaScript (דף לקוח)
@limiter.exempt  # פטור מ-rate limiting (פעולה רגילה של משתמש)
def delete_charge(client_id, charge_id):
    """מחיקת חיוב מלקוח. מחזיר JSON להתאמה ל-SPA."""
    try:
        data = load_data()
        for c in data:
            if c['id'] == client_id:
                charges = c.get('extra_charges', [])
                c['extra_charges'] = [ch for ch in charges if ch.get('id') != charge_id]
                save_data(data)
                return jsonify({'success': True})
        return jsonify({'success': False, 'error': 'לקוח לא נמצא'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/archive_client/<client_id>', methods=['POST'])
@login_required
def archive_client(client_id):
    """מעביר לקוח לארכיון (לא מוחק אותו). מחזיר JSON כדי למנוע 'Pending' על ה-toggle."""
    try:
        user_role = get_user_role(current_user.id)
        if not is_manager_or_admin(current_user.id, user_role):
            return jsonify({'success': False, 'error': 'גישה חסומה - אין לך הרשאה לבצע פעולה זו'}), 403

        data = load_data()
        for c in data:
            if c['id'] == client_id:
                c['archived'] = True
                c['archived_at'] = datetime.now().isoformat()
                save_data(data)
                return jsonify({'success': True})
        return jsonify({'success': False, 'error': 'לקוח לא נמצא'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/toggle_client_active/<client_id>', methods=['POST'])
@login_required
@csrf.exempt
def toggle_client_active(client_id):
    """מעדכן את סטטוס הפעיל/לא פעיל של לקוח"""
    try:
        user_role = get_user_role(current_user.id)
        # רק מנהל ואדמין יכולים לעדכן את הסטטוס
        if not is_manager_or_admin(current_user.id, user_role):
            return jsonify({'success': False, 'error': 'גישה חסומה - אין לך הרשאה לבצע פעולה זו'}), 403
        
        data = request.get_json()
        is_active = data.get('active', True)
        
        data_clients = load_data()
        for c in data_clients:
            if c['id'] == client_id:
                if is_active:
                    # הפעל את הלקוח - הסר את הסטטוס מאויש
                    c['archived'] = False
                    if 'archived_at' in c:
                        del c['archived_at']
                else:
                    # העבר לארכיון
                    c['archived'] = True
                    c['archived_at'] = datetime.now().isoformat()
                
                save_data(data_clients)
                return jsonify({'success': True})
        
        return jsonify({'success': False, 'error': 'לקוח לא נמצא'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/archive')
@login_required
def api_archive():
    """API endpoint להחזרת לקוחות מאוישים"""
    try:
        user_role = get_user_role(current_user.id)
        if not check_permission('/archive', user_role):
            return jsonify({'success': False, 'error': 'גישה חסומה'}), 403
        
        if not is_manager_or_admin(current_user.id, user_role):
            return jsonify({'success': False, 'error': 'גישה חסומה'}), 403
        
        all_clients = load_data()
        archived_clients = filter_archived_clients(all_clients)
        archived_clients.sort(key=lambda x: x.get('archived_at', ''), reverse=True)
        
        return jsonify({
            'success': True,
            'clients': archived_clients
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/archive')
@login_required
def archive():
    """דף ארכיון לקוחות"""
    user_role = get_user_role(current_user.id)
    if not check_permission('/archive', user_role):
        return "גישה חסומה - אין לך הרשאה לגשת לדף זה", 403
    
    # רק מנהל ואדמין יכולים לראות את הארכיון
    if not is_manager_or_admin(current_user.id, user_role):
        return "גישה חסומה - אין לך הרשאה לגשת לדף זה", 403
    
    all_clients = load_data()
    # סינון לקוחות מאוישים בלבד
    archived_clients = filter_archived_clients(all_clients)
    archived_clients.sort(key=lambda x: x.get('archived_at', ''), reverse=True)  # מיון לפי תאריך ארכיון (החדש ביותר ראשון)
    
    # Redirect to React archive page
    return redirect('/app/archive')

@app.route('/events_archive')
@login_required
def events_archive():
    """דף ארכיון אירועים"""
    user_role = get_user_role(current_user.id)
    if not check_permission('/events', user_role):
        return "גישה חסומה - אין לך הרשאה לגשת לדף זה", 403
    
    # רק מנהל ואדמין יכולים לראות את הארכיון
    if not is_manager_or_admin(current_user.id, user_role):
        return "גישה חסומה - אין לך הרשאה לגשת לדף זה", 403
    
    events_list = load_events()
    clients = load_data()
    
    # סינון אירועים מאוישים בלבד
    archived_events = filter_archived_events(events_list)
    archived_events.sort(key=lambda x: x.get('archived_at', ''), reverse=True)  # מיון לפי תאריך ארכיון (החדש ביותר ראשון)
    
    # חיבור לקוחות לאירועים
    for event in archived_events:
        client_id = event.get('client_id', '')
        event['client_name'] = next((c.get('name', '') for c in clients if c.get('id') == client_id), 'לא צוין')
    
    # Redirect to React archive page
    return redirect('/app/archive')

@app.route('/export_open_charges')
@login_required
def export_open_charges():
    """ייצוא חיובים פתוחים לאקסל"""
    try:
        data = load_data()
        
        # איסוף כל החיובים הפתוחים
        open_charges_list = []
        for client in data:
            client_name = client.get('name', 'לא צוין')
            extra_charges = client.get('extra_charges', [])
            
            for charge in extra_charges:
                if not charge.get('completed', False):
                    open_charges_list.append({
                        'client_name': client_name,
                        'title': charge.get('title', 'ללא תיאור'),
                        'date': charge.get('date', 'ללא תאריך'),
                        'amount': charge.get('amount', 0)
                    })
        
        # יצירת Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "חיובים פתוחים"
        
        # הגדרת כותרת
        ws['A1'] = "דו\"ח חיובים פתוחים"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='right', vertical='center')
        ws.merge_cells('A1:D1')
        
        # חישוב סה"כ
        total_open = sum(ch['amount'] for ch in open_charges_list)
        ws['A2'] = f"סה\"כ חיובים פתוחים: ₪{total_open:,.0f}"
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].alignment = Alignment(horizontal='right')
        
        # כותרת הטבלה
        header_row = 4
        ws.cell(row=header_row, column=1, value='לקוח').font = Font(bold=True)
        ws.cell(row=header_row, column=2, value='תיאור').font = Font(bold=True)
        ws.cell(row=header_row, column=3, value='תאריך').font = Font(bold=True)
        ws.cell(row=header_row, column=4, value='סכום').font = Font(bold=True)
        
        # עיצוב שורת כותרת
        for col in range(1, 5):
            cell = ws.cell(row=header_row, column=col)
            cell.fill = PatternFill(start_color='e1e6ff', end_color='e1e6ff', fill_type='solid')
            cell.alignment = Alignment(horizontal='right' if col == 1 or col == 2 else 'center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # הוספת חיובים
        data_row = header_row + 1
        for charge in open_charges_list:
            ws.cell(row=data_row, column=1, value=charge['client_name'])
            ws.cell(row=data_row, column=2, value=charge['title'])
            ws.cell(row=data_row, column=3, value=charge['date'])
            ws.cell(row=data_row, column=4, value=f"₪{charge['amount']:,.0f}")
            
            # עיצוב שורות הנתונים
            for col in range(1, 5):
                cell = ws.cell(row=data_row, column=col)
                cell.alignment = Alignment(horizontal='right' if col == 1 or col == 2 else 'center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            data_row += 1
        
        # שורת סיכום
        summary_row = data_row + 1
        ws.cell(row=summary_row, column=1, value='סה"כ').font = Font(bold=True, size=12)
        ws.cell(row=summary_row, column=2, value='')
        ws.cell(row=summary_row, column=3, value='')
        ws.cell(row=summary_row, column=4, value=f'₪{total_open:,.0f}').font = Font(bold=True, size=12)
        
        # עיצוב שורת סיכום
        for col in range(1, 5):
            cell = ws.cell(row=summary_row, column=col)
            cell.fill = PatternFill(start_color='e1e6ff', end_color='e1e6ff', fill_type='solid')
            cell.alignment = Alignment(horizontal='right' if col == 1 else 'center', vertical='center')
            cell.border = Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium')
            )
        
        # התאמת רוחב עמודות
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        # שמירת הקובץ
        filename = f"open_charges_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(app.root_path, 'static', filename)
        wb.save(filepath)
        
        # שליחת הקובץ
        return send_file(filepath, as_attachment=True, download_name="חיובים_פתוחים.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
    except Exception as e:
        import traceback
        print(f"Error exporting open charges: {e}")
        print(traceback.format_exc())
        return f"שגיאה בייצוא החיובים הפתוחים: {str(e)}", 500

# --- Suppliers Routes ---
@app.route('/api/suppliers')
@login_required
def api_suppliers():
    """API endpoint להחזרת ספקים"""
    try:
        user_role = get_user_role(current_user.id)
        if not check_permission('/suppliers', user_role):
            return jsonify({'success': False, 'error': 'גישה חסומה'}), 403
        
        suppliers_list = load_suppliers()
        return jsonify({
            'success': True,
            'suppliers': suppliers_list
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/suppliers')
@login_required
def suppliers():
    user_role = get_user_role(current_user.id)
    if not check_permission('/suppliers', user_role):
        return "גישה חסומה - אין לך הרשאה לגשת לדף זה", 403
    suppliers_list = load_suppliers()
    users = load_users()
    sidebar_users = {uid: {'name': info.get('name', '')} for uid, info in users.items() if uid != 'admin'}
    # Redirect to React suppliers page
    return redirect('/app/suppliers')

@app.route('/add_supplier', methods=['POST'])
@login_required
def add_supplier():
    try:
        suppliers_list = load_suppliers()
        supplier = {
            'id': str(uuid.uuid4()),
            'name': request.form.get('name', ''),
            'phone': request.form.get('phone', ''),
            'email': request.form.get('email', ''),
            'supplier_type': request.form.get('supplier_type', ''),
            'category': request.form.get('category', ''),
            'notes': request.form.get('notes', ''),
            'created_date': datetime.now().strftime('%d/%m/%y')
        }
        suppliers_list.append(supplier)
        save_suppliers(suppliers_list)
        
        if request.is_json or request.headers.get('Accept') == 'application/json':
            return jsonify({'success': True, 'supplier': supplier})
        return redirect(url_for('suppliers'))
    except Exception as e:
        if request.is_json or request.headers.get('Accept') == 'application/json':
            return jsonify({'success': False, 'error': str(e)}), 500
        return f"שגיאה בהוספת הספק: {str(e)}", 500

@app.route('/edit_supplier/<supplier_id>', methods=['POST'])
@login_required
def edit_supplier(supplier_id):
    try:
        suppliers_list = load_suppliers()
        for s in suppliers_list:
            if s['id'] == supplier_id:
                s['name'] = request.form.get('name', '')
                s['phone'] = request.form.get('phone', '')
                s['email'] = request.form.get('email', '')
                s['supplier_type'] = request.form.get('supplier_type', '')
                s['category'] = request.form.get('category', '')
                s['notes'] = request.form.get('notes', '')
                save_suppliers(suppliers_list)
                return redirect(url_for('suppliers'))
        return "ספק לא נמצא", 404
    except Exception as e:
        return f"שגיאה בעדכון הספק: {str(e)}", 500

@app.route('/delete_supplier/<supplier_id>', methods=['POST'])
@login_required
def delete_supplier(supplier_id):
    try:
        suppliers_list = load_suppliers()
        suppliers_list = [s for s in suppliers_list if s['id'] != supplier_id]
        save_suppliers(suppliers_list)
        return redirect(url_for('suppliers'))
    except Exception as e:
        return f"שגיאה במחיקת הספק: {str(e)}", 500

@app.route('/supplier/<supplier_id>')
@login_required
def supplier_page(supplier_id):
    """עמוד ספק בודד - תיק ספק"""
    user_role = get_user_role(current_user.id)
    if not check_permission('/suppliers', user_role):
        return "גישה חסומה - אין לך הרשאה לגשת לדף זה", 403
    
    suppliers_list = load_suppliers()
    supplier = next((s for s in suppliers_list if s['id'] == supplier_id), None)
    
    if not supplier:
        return "ספק לא נמצא", 404
    
    # וידוא שיש רשימת files ו-notes
    if 'files' not in supplier:
        supplier['files'] = []
    if 'notes_list' not in supplier:
        supplier['notes_list'] = []
    
    # Redirect to React suppliers page
    return redirect('/app/suppliers')

@app.route('/upload_supplier_file/<supplier_id>', methods=['POST'])
@login_required
def upload_supplier_file(supplier_id):
    """העלאת קובץ לספק"""
    try:
        if 'file' not in request.files:
            return redirect(request.referrer or url_for('supplier_page', supplier_id=supplier_id))
        
        file = request.files['file']
        file_name = request.form.get('file_name', '')  # שם הקובץ האופציונלי
        
        if file and file.filename != '':
            # יצירת שם קובץ בטוח
            filename = secure_filename(f"{uuid.uuid4().hex}_{file.filename}")
            filepath = os.path.join(SUPPLIER_FILES_FOLDER, filename)
            
            # וידוא שהתיקייה קיימת
            if not os.path.exists(SUPPLIER_FILES_FOLDER):
                os.makedirs(SUPPLIER_FILES_FOLDER, exist_ok=True)
            
            file.save(filepath)
            
            # עדכון הנתונים
            suppliers_list = load_suppliers()
            for s in suppliers_list:
                if s['id'] == supplier_id:
                    if 'files' not in s:
                        s['files'] = []
                    
                    file_doc = {
                        'id': str(uuid.uuid4()),
                        'filename': filename,
                        'original_name': file_name if file_name else file.filename,
                        'upload_date': datetime.now().strftime('%d/%m/%Y %H:%M'),
                        'uploaded_by': current_user.id
                    }
                    s['files'].append(file_doc)
                    save_suppliers(suppliers_list)
                    return redirect(url_for('supplier_page', supplier_id=supplier_id))
            
            return "ספק לא נמצא", 404
        
        return redirect(request.referrer or url_for('supplier_page', supplier_id=supplier_id))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"שגיאה בהעלאת הקובץ: {str(e)}", 500

@app.route('/delete_supplier_file/<supplier_id>/<file_id>', methods=['POST'])
@login_required
def delete_supplier_file(supplier_id, file_id):
    """מחיקת קובץ מספק"""
    try:
        suppliers_list = load_suppliers()
        for s in suppliers_list:
            if s['id'] == supplier_id:
                if 'files' not in s:
                    s['files'] = []
                
                # מצא את הקובץ
                file_to_remove = None
                for f in s['files']:
                    if str(f.get('id', '')) == str(file_id):
                        file_to_remove = f
                        break
                
                if file_to_remove:
                    # מחיקת הקובץ הפיזי
                    filepath = os.path.join(SUPPLIER_FILES_FOLDER, file_to_remove['filename'])
                    if os.path.exists(filepath):
                        os.remove(filepath)
                    
                    # מחיקת הרשומה
                    s['files'].remove(file_to_remove)
                    save_suppliers(suppliers_list)
                    return jsonify({'status': 'success', 'message': 'הקובץ נמחק בהצלחה'})
                else:
                    return "קובץ לא נמצא", 404
        
        return "ספק לא נמצא", 404
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': f'שגיאה: {str(e)}'}), 500

@app.route('/add_supplier_note/<supplier_id>', methods=['POST'])
@login_required
def add_supplier_note(supplier_id):
    """הוספת הערה לספק"""
    try:
        note_text = request.form.get('note', '').strip()
        if not note_text:
            return redirect(request.referrer or url_for('supplier_page', supplier_id=supplier_id))
        
        suppliers_list = load_suppliers()
        for s in suppliers_list:
            if s['id'] == supplier_id:
                if 'notes_list' not in s:
                    s['notes_list'] = []
                
                note_doc = {
                    'id': str(uuid.uuid4()),
                    'text': note_text,
                    'created_date': datetime.now().strftime('%d/%m/%Y %H:%M'),
                    'created_by': current_user.id
                }
                s['notes_list'].append(note_doc)
                save_suppliers(suppliers_list)
                return redirect(url_for('supplier_page', supplier_id=supplier_id))
        
        return "ספק לא נמצא", 404
    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"שגיאה בהוספת ההערה: {str(e)}", 500

@app.route('/delete_supplier_note/<supplier_id>/<note_id>', methods=['POST'])
@login_required
def delete_supplier_note(supplier_id, note_id):
    """מחיקת הערה מספק"""
    try:
        suppliers_list = load_suppliers()
        for s in suppliers_list:
            if s['id'] == supplier_id:
                if 'notes_list' not in s:
                    s['notes_list'] = []
                
                # מצא את ההערה
                note_to_remove = None
                for note in s['notes_list']:
                    if str(note.get('id', '')) == str(note_id):
                        note_to_remove = note
                        break
                
                if note_to_remove:
                    s['notes_list'].remove(note_to_remove)
                    save_suppliers(suppliers_list)
                    return jsonify({'status': 'success', 'message': 'ההערה נמחקה בהצלחה'})
                else:
                    return "הערה לא נמצאה", 404
        
        return "ספק לא נמצא", 404
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': f'שגיאה: {str(e)}'}), 500

@app.route('/supplier_files/<filename>')
@login_required
def get_supplier_file(filename):
    """שירות קובץ של ספק"""
    try:
        # בדיקת הרשאות - רק משתמשים מחוברים
        return send_from_directory(SUPPLIER_FILES_FOLDER, filename)
    except Exception as e:
        return f"שגיאה בטעינת הקובץ: {str(e)}", 404

@app.route('/import_suppliers_excel', methods=['POST'])
@login_required
def import_suppliers_excel():
    """Route לייבוא ספקים מקובץ Excel"""
    try:
        if 'excel_file' not in request.files:
            return jsonify({'success': False, 'error': 'לא נבחר קובץ'}), 400
        
        file = request.files['excel_file']
        if not file or not file.filename:
            return jsonify({'success': False, 'error': 'קובץ לא תקין'}), 400
        
        # בדיקת סוג קובץ
        if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            return jsonify({'success': False, 'error': 'קובץ Excel בלבד (xlsx, xls)'}), 400
        
        # קריאת קובץ Excel
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        
        # קריאת הנתונים (מניחים שהשורה הראשונה היא כותרות)
        # פורמט צפוי: שם | טלפון | אימייל | אתר | תחום | הערות
        suppliers_list = load_suppliers()
        imported_count = 0
        
        # דילוג על שורת הכותרת
        for row in ws.iter_rows(min_row=2, values_only=True):
            # בדיקה שיש לפחות שם
            if not row[0] or str(row[0]).strip() == '':
                continue
            
            supplier = {
                'id': str(uuid.uuid4()),
                'name': str(row[0]).strip() if row[0] else '',
                'phone': str(row[1]).strip() if len(row) > 1 and row[1] else '',
                'email': str(row[2]).strip() if len(row) > 2 and row[2] else '',
                'supplier_type': str(row[3]).strip() if len(row) > 3 and row[3] else '',
                'category': str(row[4]).strip() if len(row) > 4 and row[4] else '',
                'notes': str(row[5]).strip() if len(row) > 5 and row[5] else '',
                'created_date': datetime.now().strftime('%d/%m/%y')
            }
            
            suppliers_list.append(supplier)
            imported_count += 1
        
        save_suppliers(suppliers_list)
        return jsonify({'success': True, 'count': imported_count})
        
    except Exception as e:
        import traceback
        print(f"Error importing Excel: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'error': f'שגיאה: {str(e)}'}), 500

# --- Quotes Routes ---
@app.route('/api/quotes')
@login_required
def api_quotes():
    """API endpoint להחזרת הצעות מחיר"""
    try:
        user_role = get_user_role(current_user.id)
        if not check_permission('/quotes', user_role):
            return jsonify({'success': False, 'error': 'גישה חסומה'}), 403
        
        quotes_list = load_quotes()
        clients = load_data()
        clients_dict = {c['id']: c['name'] for c in clients}
        
        # Add client names to quotes
        for quote in quotes_list:
            client_id = quote.get('client_id', '')
            quote['client_name'] = clients_dict.get(client_id, '')
            # Calculate total
            items = quote.get('items', [])
            quote['total'] = sum(item.get('quantity', 0) * item.get('price', 0) for item in items)
        
        return jsonify({
            'success': True,
            'quotes': quotes_list
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/quotes')
@login_required
def quotes():
    user_role = get_user_role(current_user.id)
    if not check_permission('/quotes', user_role):
        return "גישה חסומה - אין לך הרשאה לגשת לדף זה", 403
    quotes_list = load_quotes()
    clients = load_data()
    # Redirect to React quotes page
    return redirect('/app/quotes')

@app.route('/add_quote', methods=['POST'])
@login_required
def add_quote():
    try:
        quotes_list = load_quotes()
        items = []
        # קבלת פריטים מהטופס
        item_titles = request.form.getlist('item_title[]')
        item_quantities = request.form.getlist('item_quantity[]')
        item_prices = request.form.getlist('item_price[]')
        
        for i in range(len(item_titles)):
            if item_titles[i] and item_quantities[i] and item_prices[i]:
                items.append({
                    'title': item_titles[i],
                    'quantity': int(item_quantities[i]) if item_quantities[i] else 1,
                    'price': float(item_prices[i]) if item_prices[i] else 0
                })
        
        total = sum(item['quantity'] * item['price'] for item in items)
        
        quote = {
            'id': str(uuid.uuid4()),
            'name': request.form.get('name', ''),
            'client_id': request.form.get('client_id', ''),
            'items': items,
            'total': total,
            'notes': request.form.get('notes', ''),
            'created_date': datetime.now().strftime('%d/%m/%y'),
            'status': request.form.get('status', 'draft')  # draft, sent, approved, rejected
        }
        quotes_list.append(quote)
        save_quotes(quotes_list)
        
        if request.is_json or request.headers.get('Accept') == 'application/json':
            return jsonify({'success': True, 'quote': quote})
        return redirect(url_for('quotes'))
    except Exception as e:
        if request.is_json or request.headers.get('Accept') == 'application/json':
            return jsonify({'success': False, 'error': str(e)}), 500
        return f"שגיאה ביצירת הצעת מחיר: {str(e)}", 500

@app.route('/edit_quote/<quote_id>', methods=['POST'])
@login_required
def edit_quote(quote_id):
    try:
        quotes_list = load_quotes()
        for q in quotes_list:
            if q['id'] == quote_id:
                items = []
                item_titles = request.form.getlist('item_title[]')
                item_quantities = request.form.getlist('item_quantity[]')
                item_prices = request.form.getlist('item_price[]')
                
                for i in range(len(item_titles)):
                    if item_titles[i] and item_quantities[i] and item_prices[i]:
                        items.append({
                            'title': item_titles[i],
                            'quantity': int(item_quantities[i]) if item_quantities[i] else 1,
                            'price': float(item_prices[i]) if item_prices[i] else 0
                        })
                
                q['name'] = request.form.get('name', '')
                q['client_id'] = request.form.get('client_id', '')
                q['items'] = items
                q['total'] = sum(item['quantity'] * item['price'] for item in items)
                q['notes'] = request.form.get('notes', '')
                q['status'] = request.form.get('status', 'draft')
                save_quotes(quotes_list)
                return redirect(url_for('quotes'))
        return "הצעת מחיר לא נמצאה", 404
    except Exception as e:
        return f"שגיאה בעדכון הצעת המחיר: {str(e)}", 500

@app.route('/delete_quote/<quote_id>', methods=['POST'])
@login_required
def delete_quote(quote_id):
    try:
        quotes_list = load_quotes()
        quotes_list = [q for q in quotes_list if q['id'] != quote_id]
        save_quotes(quotes_list)
        return redirect(url_for('quotes'))
    except Exception as e:
        return f"שגיאה במחיקת הצעת המחיר: {str(e)}", 500

# --- Chat Routes (New Chat System) ---
@app.route('/api/chat/conversations')
@limiter.exempt  # פטור מ-rate limiting כי זה auto-refresh
@login_required
def get_chat_conversations():
    """מחזיר רשימת שיחות עם משתמשים אחרים"""
    try:
        messages_list = load_messages()
        users = load_users()
        current_id = current_user.id
        
        # קבל כל השיחות הרלוונטיות למשתמש הנוכחי
        user_conversations = {}
        for msg in messages_list:
            from_user = msg.get('from_user')
            to_user = msg.get('to_user')
            
            # קבע את המשתמש השני בשיחה
            other_user = None
            if from_user == current_id and to_user != current_id:
                other_user = to_user
            elif to_user == current_id and from_user != current_id:
                other_user = from_user
            
            if other_user and other_user in users:
                if other_user not in user_conversations:
                    user_conversations[other_user] = {
                        'user_id': other_user,
                        'user_name': users[other_user].get('name', other_user),
                        'last_message': '',
                        'last_message_time': '',
                        'unread_count': 0,
                        'is_active': is_user_active(other_user)
                    }
                
                # עדכן הודעה אחרונה
                msg_time = msg.get('created_date', '')
                if not user_conversations[other_user]['last_message_time'] or msg_time > user_conversations[other_user]['last_message_time']:
                    user_conversations[other_user]['last_message'] = msg.get('content', '')[:50]
                    user_conversations[other_user]['last_message_time'] = msg_time
                
                # ספור הודעות שלא נקראו
                if to_user == current_id and not msg.get('read', False):
                    user_conversations[other_user]['unread_count'] += 1
        
        # המר לרשימה ומיין לפי זמן
        conversations = list(user_conversations.values())
        conversations.sort(key=lambda x: x['last_message_time'], reverse=True)
        
        return jsonify({
            'status': 'success',
            'conversations': conversations
        })
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/chat/messages/<user_id>')
@limiter.exempt  # פטור מ-rate limiting כי זה auto-refresh
@login_required
def get_chat_messages(user_id):
    """מחזיר את כל ההודעות בשיחה עם משתמש מסוים"""
    try:
        messages_list = load_messages()
        current_id = current_user.id
        
        # סנן הודעות רלוונטיות לשיחה
        conversation_messages = []
        for msg in messages_list:
            from_user = msg.get('from_user')
            to_user = msg.get('to_user')
            
            if (from_user == current_id and to_user == user_id) or \
               (from_user == user_id and to_user == current_id):
                conversation_messages.append({
                    'id': msg.get('id'),
                    'from_user': from_user,
                    'to_user': to_user,
                    'content': msg.get('content', ''),
                    'created_date': msg.get('created_date', ''),
                    'read': msg.get('read', False),
                    'files': msg.get('files', []),
                    'is_manager_note': msg.get('is_manager_note', False)
                })
        
        # מיון לפי זמן
        conversation_messages.sort(key=lambda x: x['created_date'])
        
        return jsonify({
            'status': 'success',
            'messages': conversation_messages
        })
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/chat/send', methods=['POST'])
@csrf.exempt  # פטור מ-CSRF כי זה API call מ-JavaScript
@login_required
def send_chat_message():
    """שולח הודעה חדשה עם תמיכה בקבצים"""
    try:
        to_user = request.form.get('to_user')
        content = request.form.get('content', '').strip()
        files = request.files.getlist('files')
        
        if not to_user:
            return jsonify({'status': 'error', 'error': 'חסר נמען'}), 400
        
        # בדיקה אם יש תוכן או קבצים
        has_content = content and content.strip()
        has_files = files and len([f for f in files if f and f.filename]) > 0
        
        if not has_content and not has_files:
            return jsonify({'status': 'error', 'error': 'יש להזין תוכן או לצרף קובץ'}), 400
        
        # שמירת קבצים
        saved_files = []
        for file in files:
            if file and file.filename:
                try:
                    filename = secure_filename(file.filename)
                    # יצירת שם ייחודי
                    file_id = str(uuid.uuid4())
                    file_ext = os.path.splitext(filename)[1] or '.bin'
                    unique_filename = f"{file_id}{file_ext}"
                    file_path = os.path.join(CHAT_FILES_FOLDER, unique_filename)
                    
                    # וודא שהתיקייה קיימת
                    os.makedirs(CHAT_FILES_FOLDER, exist_ok=True)
                    
                    file.save(file_path)
                    saved_files.append({
                        'filename': unique_filename,
                        'original_name': filename
                    })
                except Exception as file_error:
                    print(f"[ERROR] Error saving file {file.filename}: {str(file_error)}")
                    import traceback
                    traceback.print_exc()
                    # המשך עם שאר הקבצים
        
        messages_list = load_messages()
        message = {
            'id': str(uuid.uuid4()),
            'from_user': current_user.id,
            'to_user': to_user,
            'subject': '',
            'content': content,
            'project_id': '',
            'task_id': '',
            'client_id': '',
            'created_date': datetime.now().strftime('%d/%m/%y %H:%M'),
            'read': False,
            'files': saved_files if saved_files else []
        }
        messages_list.append(message)
        save_messages(messages_list)
        
        return jsonify({
            'status': 'success',
            'message': message
        })
    except Exception as e:
        import traceback
        error_msg = str(e)
        print(f"[ERROR] Error in send_chat_message: {error_msg}")
        print(traceback.format_exc())
        return jsonify({'status': 'error', 'error': f'שגיאה בשליחת ההודעה: {error_msg}'}), 500

@app.route('/api/chat/mark-read/<user_id>', methods=['POST'])
@limiter.exempt  # פטור מ-rate limiting
@login_required
def mark_chat_read(user_id):
    """מסמן את כל ההודעות ממשתמש מסוים כנקראות"""
    try:
        messages_list = load_messages()
        updated = False
        
        for msg in messages_list:
            if msg.get('from_user') == user_id and msg.get('to_user') == current_user.id:
                if not msg.get('read', False):
                    msg['read'] = True
                    updated = True
        
        if updated:
            save_messages(messages_list)
        
        return jsonify({'status': 'success'})
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/chat/users')
@login_required
def get_chat_users():
    """מחזיר רשימת משתמשים זמינים לשיחה"""
    try:
        users = load_users()
        current_id = current_user.id
        
        # החזר את כל המשתמשים למעט המשתמש הנוכחי
        available_users = []
        for user_id, user_data in users.items():
            if user_id != current_id:
                available_users.append({
                    'id': user_id,
                    'name': user_data.get('name', user_id),
                    'is_active': is_user_active(user_id)
                })
        
        return jsonify({
            'status': 'success',
            'users': available_users
        })
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/static/chat_files/<filename>')
@login_required
def get_chat_file(filename):
    """הורדת/הצגת קובץ מצ'אט (תמונות יוצגו ישירות)"""
    try:
        file_path = os.path.join(CHAT_FILES_FOLDER, filename)
        if os.path.exists(file_path):
            # בדיקה אם זה קובץ תמונה
            ext = os.path.splitext(filename)[1].lower()
            image_extensions = {'.jpg': 'image/jpeg', '.jpeg': 'image/jpeg', '.png': 'image/png', 
                               '.gif': 'image/gif', '.webp': 'image/webp', '.bmp': 'image/bmp'}
            if ext in image_extensions:
                # החזר תמונה להצגה ישירה
                return send_file(file_path, mimetype=image_extensions[ext])
            else:
                # החזר קובץ להורדה
                return send_file(file_path, as_attachment=True)
        else:
            return "קובץ לא נמצא", 404
    except Exception as e:
        return f"שגיאה: {str(e)}", 500

# --- Old Messages Routes (Kept for backward compatibility) ---
@app.route('/messages')
@login_required
def messages():
    messages_list = load_messages()
    users = load_users()
    clients = load_data()
    # סנן רק הודעות רלוונטיות למשתמש הנוכחי
    user_messages = [m for m in messages_list if m.get('to_user') == current_user.id or m.get('from_user') == current_user.id]
    # מיון לפי תאריך - החדשים למעלה
    user_messages.sort(key=lambda x: x.get('created_date', ''), reverse=True)
    # Redirect to React app
    return redirect('/app')

@app.route('/send_message', methods=['POST'])
@login_required
def send_message():
    try:
        messages_list = load_messages()
        message = {
            'id': str(uuid.uuid4()),
            'from_user': current_user.id,
            'to_user': request.form.get('to_user', ''),
            'subject': request.form.get('subject', ''),
            'content': request.form.get('content', ''),
            'project_id': request.form.get('project_id', ''),
            'task_id': request.form.get('task_id', ''),
            'client_id': request.form.get('client_id', ''),
            'created_date': datetime.now().strftime('%d/%m/%y %H:%M'),
            'read': False
        }
        messages_list.append(message)
        save_messages(messages_list)
        return redirect(url_for('messages'))
    except Exception as e:
        return f"שגיאה בשליחת ההודעה: {str(e)}", 500

@app.route('/mark_message_read/<message_id>', methods=['POST'])
@login_required
def mark_message_read(message_id):
    try:
        messages_list = load_messages()
        for m in messages_list:
            if m['id'] == message_id and m.get('to_user') == current_user.id:
                m['read'] = True
                save_messages(messages_list)
                break
        return redirect(url_for('messages'))
    except Exception as e:
        return f"שגיאה: {str(e)}", 500

# --- Events Routes ---
@app.route('/api/events')
@login_required
def api_events():
    """API endpoint להחזרת אירועים"""
    try:
        user_role = get_user_role(current_user.id)
        if not check_permission('/events', user_role):
            return jsonify({'success': False, 'error': 'גישה חסומה'}), 403
        
        events_list = load_events()
        clients = load_data()
        events_list = filter_active_events(events_list)
        
        today = datetime.now().date()
        open_events = []
        for event in events_list:
            event_date_str = event.get('date', '')
            if event_date_str:
                try:
                    if '/' in event_date_str:
                        parts = event_date_str.split('/')
                        if len(parts) == 3:
                            event_date = datetime.strptime(event_date_str, '%d/%m/%Y').date()
                        else:
                            event_date = datetime.strptime(event_date_str, '%d/%m/%y').date()
                    else:
                        event_date = datetime.strptime(event_date_str, '%Y-%m-%d').date()
                    
                    if event_date >= today:
                        open_events.append(event)
                except:
                    open_events.append(event)
            else:
                open_events.append(event)
        
        # Add client names
        for event in open_events:
            client_id = event.get('client_id', '')
            event['client_name'] = next((c.get('name', '') for c in clients if c.get('id') == client_id), '')
        
        return jsonify({
            'success': True,
            'events': open_events
        })
    except Exception as e:
        print(f"Error in api_events: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/event/<event_id>')
@login_required
def api_event_details(event_id):
    """API endpoint להחזרת פרטי אירוע בודד"""
    try:
        user_role = get_user_role(current_user.id)
        if not check_permission('/events', user_role):
            return jsonify({'success': False, 'error': 'גישה חסומה'}), 403
        
        events_list = load_events()
        event = next((e for e in events_list if e.get('id') == event_id), None)
        
        if not event:
            return jsonify({'success': False, 'error': 'אירוע לא נמצא'}), 404
        
        clients = load_data()
        suppliers_list = load_suppliers()
        equipment_bank = load_equipment_bank()
        
        # חיבור לקוח לאירוע
        client_id = event.get('client_id', '')
        client = next((c for c in clients if c.get('id') == client_id), None)
        event['client'] = client
        event['client_name'] = client.get('name', '') if client else ''
        
        # טעינת צ'ק-ליסט לפי סוג האירוע
        event_type = event.get('event_type', '')
        checklist_template = get_event_checklist_template(event_type)
        
        # אם אין צ'ק-ליסט לאירוע, טען מהתבנית
        if 'checklist' not in event or not event.get('checklist'):
            event['checklist'] = [{'task': task, 'completed': False} for task in checklist_template]
        
        # וודא שכל הפריטים מהתבנית קיימים
        if checklist_template:
            existing_tasks = {item.get('task', '') for item in event.get('checklist', [])}
            for task in checklist_template:
                if task not in existing_tasks:
                    event['checklist'].append({'task': task, 'completed': False})
        
        # וודא ששדות קיימים
        if 'suppliers' not in event:
            event['suppliers'] = []
        if 'equipment' not in event:
            event['equipment'] = []
        if 'charges' not in event:
            event['charges'] = []
        if 'graphics_items' not in event:
            event['graphics_items'] = []
        
        # הוסף our_cost=0 לחיובים ישנים
        for charge in event.get('charges', []):
            if 'our_cost' not in charge:
                charge['our_cost'] = 0
        
        # חישוב סה"כ לתקציב
        total_budget = sum(ch.get('amount', 0) for ch in event.get('charges', []))
        total_expenses = sum(ch.get('our_cost', 0) for ch in event.get('charges', []))
        profit_margin = total_budget - total_expenses
        
        return jsonify({
            'success': True,
            'event': event,
            'suppliers': suppliers_list,
            'equipment_bank': equipment_bank,
            'total_budget': total_budget,
            'total_expenses': total_expenses,
            'profit_margin': profit_margin
        })
    except Exception as e:
        print(f"Error in api_event_details: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/events')
@login_required
def events():
    """דשבורד אירועים - רשימת אירועים פתוחים"""
    user_role = get_user_role(current_user.id)
    if not check_permission('/events', user_role):
        return "גישה חסומה - אין לך הרשאה לגשת לדף זה", 403
    events_list = load_events()
    clients = load_data()
    users = load_users()
    
    # בניית sidebar_users
    sidebar_users = {uid: {'name': info.get('name', '')} for uid, info in users.items() if uid != 'admin'}
    
    # סינון אירועים מאוישים
    events_list = filter_active_events(events_list)
    
    # סינון אירועים פתוחים (אירועים שעוד לא עברו או סטטוס פעיל)
    today = datetime.now().date()
    open_events = []
    for event in events_list:
        event_date_str = event.get('date', '')
        if event_date_str:
            try:
                # פורמט תאריך: YYYY-MM-DD או DD/MM/YYYY
                if '/' in event_date_str:
                    parts = event_date_str.split('/')
                    if len(parts) == 3:
                        event_date = datetime.strptime(event_date_str, '%d/%m/%Y').date()
                    else:
                        event_date = datetime.strptime(event_date_str, '%d/%m/%y').date()
                else:
                    event_date = datetime.strptime(event_date_str, '%Y-%m-%d').date()
                
                # הוסף אירועים עתידיים או שעברו ב-30 יום האחרונים
                if event_date >= today or (today - event_date).days <= 30:
                    open_events.append(event)
            except:
                # אם יש שגיאה בפענוח תאריך, הוסף את האירוע
                open_events.append(event)
        else:
            # אם אין תאריך, הוסף את האירוע
            open_events.append(event)
    
    # חיבור לקוחות לאירועים
    for event in open_events:
        client_id = event.get('client_id', '')
        event['client_name'] = next((c.get('name', '') for c in clients if c.get('id') == client_id), 'לא צוין')
    
    user_role = get_user_role(current_user.id)
    # Redirect to React events page
    return redirect('/app/events')

@app.route('/event/<event_id>')
@login_required
def event_page(event_id):
    """עמוד אירוע פרטי עם כל הכרטיסיות"""
    events_list = load_events()
    event = next((e for e in events_list if e.get('id') == event_id), None)
    
    if not event:
        return "אירוע לא נמצא", 404
    
    clients = load_data()
    suppliers_list = load_suppliers()
    equipment_bank = load_equipment_bank()
    users = load_users()
    
    # בניית sidebar_users
    sidebar_users = {uid: {'name': info.get('name', '')} for uid, info in users.items() if uid != 'admin'}
    
    # חיבור לקוח לאירוע
    client_id = event.get('client_id', '')
    event['client'] = next((c for c in clients if c.get('id') == client_id), None)
    
    # טעינת צ'ק-ליסט לפי סוג האירוע
    event_type = event.get('event_type', '')
    checklist_template = get_event_checklist_template(event_type)
    
    # אם אין צ'ק-ליסט לאירוע, טען מהתבנית
    if 'checklist' not in event or not event.get('checklist'):
        event['checklist'] = [{'task': task, 'completed': False} for task in checklist_template]
    
    # וודא שכל הפריטים מהתבנית קיימים (למקרה שנוספו פריטים חדשים לתבנית)
    if checklist_template:
        existing_tasks = {item.get('task', '') for item in event.get('checklist', [])}
        for task in checklist_template:
            if task not in existing_tasks:
                event['checklist'].append({'task': task, 'completed': False})
    
    # וודא ששדות קיימים
    if 'suppliers' not in event:
        event['suppliers'] = []
    if 'equipment' not in event:
        event['equipment'] = []
    if 'charges' not in event:
        event['charges'] = []
    if 'graphics_items' not in event:
        event['graphics_items'] = []
    
    # הוסף our_cost=0 לחיובים ישנים שאין להם
    for charge in event.get('charges', []):
        if 'our_cost' not in charge:
            charge['our_cost'] = 0
    
    # חישוב סה"כ לתקציב
    total_budget = sum(ch.get('amount', 0) for ch in event.get('charges', []))
    total_expenses = sum(ch.get('our_cost', 0) for ch in event.get('charges', []))
    profit_margin = total_budget - total_expenses
    
    # Redirect to React events page
    return redirect('/app/events')

@app.route('/add_event', methods=['POST'])
@login_required
def add_event():
    """יצירת אירוע חדש"""
    try:
        events_list = load_events()
        event_type = request.form.get('event_type', '')
        
        # טעינת תבנית צ'ק-ליסט לפי סוג האירוע
        checklist_template = get_event_checklist_template(event_type)
        
        event = {
            'id': str(uuid.uuid4()),
            'title': request.form.get('title', '') or request.form.get('name', ''),
            'name': request.form.get('title', '') or request.form.get('name', ''),
            'client_id': request.form.get('client_id', ''),
            'date': request.form.get('date', ''),
            'location': request.form.get('location', ''),
            'type': request.form.get('type', '') or event_type,
            'event_type': event_type,
            'checklist': [{'task': task, 'completed': False} for task in checklist_template],
            'suppliers': [],
            'equipment': [],
            'charges': [],
            'reminders': [],
            'notes': '',
            'created_date': datetime.now().strftime('%d/%m/%y'),
            'status': 'active'  # active, completed, cancelled
        }
        
        events_list.append(event)
        save_events(events_list)
        wants_json = request.is_json or request.headers.get('Accept', '').find('application/json') != -1 or \
                    request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if wants_json:
            return jsonify({'success': True, 'event': event})
        return redirect(url_for('event_page', event_id=event['id']))
    except Exception as e:
        wants_json = request.is_json or request.headers.get('Accept', '').find('application/json') != -1 or \
                    request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if wants_json:
            return jsonify({'success': False, 'error': str(e)}), 500
        return f"שגיאה ביצירת האירוע: {str(e)}", 500

@app.route('/update_event/<event_id>', methods=['POST'])
@login_required
def update_event(event_id):
    """עדכון פרטי אירוע"""
    try:
        events_list = load_events()
        for event in events_list:
            if event['id'] == event_id:
                event['name'] = request.form.get('name', event.get('name', ''))
                event['client_id'] = request.form.get('client_id', event.get('client_id', ''))
                event['date'] = request.form.get('date', event.get('date', ''))
                event['location'] = request.form.get('location', event.get('location', ''))
                event['event_type'] = request.form.get('event_type', event.get('event_type', ''))
                event['notes'] = request.form.get('notes', event.get('notes', ''))
                save_events(events_list)
                tab = request.form.get('tab', 'details')
                return redirect(url_for('event_page', event_id=event_id, tab=tab))
        return "אירוע לא נמצא", 404
    except Exception as e:
        return f"שגיאה בעדכון האירוע: {str(e)}", 500

@app.route('/update_event_checklist/<event_id>', methods=['POST'])
@login_required
def update_event_checklist(event_id):
    """עדכון צ'ק-ליסט של אירוע"""
    try:
        events_list = load_events()
        for event in events_list:
            if event['id'] == event_id:
                checklist_data = request.get_json()
                if checklist_data and 'checklist' in checklist_data:
                    event['checklist'] = checklist_data['checklist']
                    save_events(events_list)
                    return jsonify({'success': True})
        return jsonify({'success': False, 'error': 'אירוע לא נמצא'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/add_checklist_item/<event_id>', methods=['POST'])
@login_required
def add_checklist_item(event_id):
    """הוספת פריט חדש לצ'ק-ליסט של אירוע ולתבנית הכללית"""
    try:
        events_list = load_events()
        event = next((e for e in events_list if e.get('id') == event_id), None)
        if not event:
            return "אירוע לא נמצא", 404
        
        new_item = request.form.get('new_item', '').strip()
        if not new_item:
            tab = request.form.get('tab', 'checklist')
            return redirect(url_for('event_page', event_id=event_id, tab=tab))
        
        # הוספה לאירוע הספציפי
        if 'checklist' not in event:
            event['checklist'] = []
        event['checklist'].append({'task': new_item, 'completed': False})
        
        # הוספה לתבנית הכללית לפי סוג האירוע
        event_type = event.get('event_type', '')
        if event_type:
            templates = load_checklist_templates()
            if event_type not in templates:
                templates[event_type] = []
            if new_item not in templates[event_type]:
                templates[event_type].append(new_item)
                save_checklist_templates(templates)
        
        save_events(events_list)
        tab = request.form.get('tab', 'checklist')
        return redirect(url_for('event_page', event_id=event_id, tab=tab))
    except Exception as e:
        return f"שגיאה: {str(e)}", 500

@app.route('/remove_checklist_item/<event_id>/<item_index>', methods=['POST'])
@login_required
def remove_checklist_item(event_id, item_index):
    """הסרת פריט מצ'ק-ליסט של אירוע ספציפי (לא מהתבנית הכללית)"""
    try:
        events_list = load_events()
        event = next((e for e in events_list if e.get('id') == event_id), None)
        if not event:
            return "אירוע לא נמצא", 404
        
        if 'checklist' in event:
            index = int(item_index)
            if 0 <= index < len(event['checklist']):
                event['checklist'].pop(index)
                save_events(events_list)
        
        return redirect(url_for('event_page', event_id=event_id))
    except Exception as e:
        return f"שגיאה: {str(e)}", 500

@app.route('/add_event_supplier/<event_id>', methods=['POST'])
@login_required
def add_event_supplier(event_id):
    """הוספת ספק לאירוע - תמיכה בספק קיים או חדש"""
    try:
        events_list = load_events()
        suppliers_list = load_suppliers()
        
        for event in events_list:
            if event['id'] == event_id:
                supplier_id = request.form.get('supplier_id', '')
                price = request.form.get('price', '0')
                payment_status = request.form.get('payment_status', 'pending')
                
                # בדיקה אם צריך ליצור ספק חדש
                if supplier_id == '__NEW__':
                    # יצירת ספק חדש
                    new_supplier = {
                        'id': str(uuid.uuid4()),
                        'name': request.form.get('new_supplier_name', ''),
                        'phone': request.form.get('new_supplier_phone', ''),
                        'email': request.form.get('new_supplier_email', ''),
                        'supplier_type': request.form.get('new_supplier_type', ''),
                        'category': request.form.get('new_supplier_category', ''),
                        'notes': request.form.get('new_supplier_notes', ''),
                        'created_date': datetime.now().strftime('%d/%m/%y')
                    }
                    suppliers_list.append(new_supplier)
                    save_suppliers(suppliers_list)
                    supplier_id = new_supplier['id']
                    supplier = new_supplier
                else:
                    # שימוש בספק קיים
                    supplier = next((s for s in suppliers_list if s.get('id') == supplier_id), None)
                
                if supplier:
                    if 'suppliers' not in event:
                        event['suppliers'] = []
                    
                    event['suppliers'].append({
                        'supplier_id': supplier_id,
                        'name': supplier.get('name', ''),
                        'phone': supplier.get('phone', ''),
                        'category': supplier.get('category', ''),
                        'price': float(price) if price else 0,
                        'payment_status': payment_status,  # pending, deposit_paid, fully_paid
                        'notes': request.form.get('notes', '')
                    })
                    save_events(events_list)
                break
        tab = request.form.get('tab', 'suppliers')
        return redirect(url_for('event_page', event_id=event_id, tab=tab))
    except Exception as e:
        return f"שגיאה בהוספת הספק: {str(e)}", 500

@app.route('/remove_event_supplier/<event_id>/<supplier_index>', methods=['POST'])
@login_required
def remove_event_supplier(event_id, supplier_index):
    """הסרת ספק מאירוע"""
    try:
        events_list = load_events()
        for event in events_list:
            if event['id'] == event_id:
                suppliers = event.get('suppliers', [])
                index = int(supplier_index)
                if 0 <= index < len(suppliers):
                    suppliers.pop(index)
                    event['suppliers'] = suppliers
                    save_events(events_list)
                break
        return redirect(url_for('event_page', event_id=event_id))
    except Exception as e:
        return f"שגיאה בהסרת הספק: {str(e)}", 500

@app.route('/update_event_equipment/<event_id>', methods=['POST'])
@login_required
def update_event_equipment(event_id):
    """עדכון רשימת ציוד לאירוע"""
    try:
        events_list = load_events()
        for event in events_list:
            if event['id'] == event_id:
                equipment_data = request.get_json()
                if equipment_data and 'equipment' in equipment_data:
                    event['equipment'] = equipment_data['equipment']
                    save_events(events_list)
                    return jsonify({'success': True})
        return jsonify({'success': False, 'error': 'אירוע לא נמצא'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/add_equipment_to_bank', methods=['POST'])
@login_required
def add_equipment_to_bank():
    """הוספת פריט חדש למאגר הציוד"""
    try:
        equipment_bank = load_equipment_bank()
        new_item = request.form.get('equipment_name', '').strip()
        if not new_item:
            return redirect(request.referrer or url_for('events'))
        if new_item not in equipment_bank:
            equipment_bank.append(new_item)
            save_equipment_bank(equipment_bank)
        return redirect(request.referrer or url_for('events'))
    except Exception as e:
        return f"שגיאה: {str(e)}", 500

@app.route('/update_event_management_table/<event_id>', methods=['POST'])
@login_required
def update_event_management_table(event_id):
    """עדכון טבלת הניהול של אירוע"""
    try:
        events_list = load_events()
        for event in events_list:
            if event['id'] == event_id:
                data = request.get_json()
                if data and 'management_table' in data:
                    event['management_table'] = data['management_table']
                    save_events(events_list)
                    return jsonify({'success': True})
        return jsonify({'success': False, 'error': 'אירוע לא נמצא'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/update_event_shopping_list/<event_id>', methods=['POST'])
@login_required
def update_event_shopping_list(event_id):
    """עדכון רשימת הקניות של אירוע"""
    try:
        events_list = load_events()
        for event in events_list:
            if event['id'] == event_id:
                data = request.get_json()
                if data and 'shopping_list' in data:
                    event['shopping_list'] = data['shopping_list']
                    save_events(events_list)
                    return jsonify({'success': True})
        return jsonify({'success': False, 'error': 'אירוע לא נמצא'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/export_event_management/<event_id>')
@login_required
def export_event_management(event_id):
    """ייצוא טבלת הניהול לאקסל"""
    try:
        events_list = load_events()
        event = next((e for e in events_list if e.get('id') == event_id), None)
        
        if not event:
            return "אירוע לא נמצא", 404
        
        management_table = event.get('management_table', [])
        
        wb = Workbook()
        ws = wb.active
        ws.title = "טבלת ניהול"
        
        # Headers
        headers = ['זמן', 'פעילות', 'מיקום', 'הפקה', 'ציוד/מזון', 'גרפיקה', 'בוצע', 'תמחור', 'עלות']
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="043841", end_color="043841", fill_type="solid")
            cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Data
        for row in management_table:
            if row.get('type') == 'day-header':
                ws.append([row.get('day', ''), '', '', '', '', '', '', '', ''])
                # Style day header
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
                    cell.font = Font(bold=True)
            else:
                ws.append([
                    row.get('time', ''),
                    row.get('activity', ''),
                    row.get('location', ''),
                    row.get('operations', ''),
                    row.get('logistics', ''),
                    row.get('graphics', ''),
                    '✓' if row.get('completed') else '',
                    row.get('budget', 0),
                    row.get('cost', 0)
                ])
        
        # Save to BytesIO
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=f'management_{event_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    except Exception as e:
        return f"שגיאה: {str(e)}", 500

@app.route('/export_event_shopping/<event_id>')
@login_required
def export_event_shopping(event_id):
    """ייצוא רשימת הקניות לאקסל"""
    try:
        events_list = load_events()
        event = next((e for e in events_list if e.get('id') == event_id), None)
        
        if not event:
            return "אירוע לא נמצא", 404
        
        shopping_list = event.get('shopping_list', [])
        
        wb = Workbook()
        ws = wb.active
        ws.title = "רשימת קניות"
        
        # Headers
        headers = ['פריט', 'כמות', 'יחידת מידה', 'ספק', 'מחיר', 'הערות', 'נרכש']
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="043841", end_color="043841", fill_type="solid")
            cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Data
        for item in shopping_list:
            ws.append([
                item.get('name', ''),
                item.get('quantity', 1),
                item.get('unit', 'יחידה'),
                item.get('supplier', ''),
                item.get('price', 0),
                item.get('notes', ''),
                '✓' if item.get('purchased') else ''
            ])
        
        # Save to BytesIO
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=f'shopping_{event_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    except Exception as e:
        return f"שגיאה: {str(e)}", 500

@app.route('/export_event_equipment/<event_id>')
@login_required
def export_event_equipment(event_id):
    """ייצוא רשימת ציוד מסומן לאקסל"""
    try:
        events_list = load_events()
        event = next((e for e in events_list if e.get('id') == event_id), None)
        
        if not event:
            return "אירוע לא נמצא", 404
        
        # קבלת רשימת פריטים מה-query string
        items_json = request.args.get('items', '[]')
        try:
            selected_items = json.loads(items_json)
        except:
            selected_items = []
        
        if not selected_items:
            return "לא נבחרו פריטים לייצוא", 400
        
        # יצירת Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "רשימת ציוד"
        
        # הגדרת כותרת
        ws['A1'] = f"רשימת ציוד - {event.get('name', 'אירוע')}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='right', vertical='center')
        ws.merge_cells('A1:B1')
        
        # פרטי האירוע
        ws['A3'] = 'לקוח:'
        ws['B3'] = event.get('client', {}).get('name', 'לא צוין') if isinstance(event.get('client'), dict) else 'לא צוין'
        ws['A4'] = 'תאריך:'
        ws['B4'] = event.get('date', 'לא צוין')
        ws['A5'] = 'מיקום:'
        ws['B5'] = event.get('location', 'לא צוין')
        
        # עיצוב שורות פרטים
        for row in range(3, 6):
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
        
        # כותרת הטבלה
        header_row = 7
        ws.cell(row=header_row, column=1, value='מס\'').font = Font(bold=True)
        ws.cell(row=header_row, column=2, value='פריט ציוד').font = Font(bold=True)
        ws.cell(row=header_row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=header_row, column=2).alignment = Alignment(horizontal='right')
        
        # עיצוב שורת כותרת
        for col in range(1, 3):
            cell = ws.cell(row=header_row, column=col)
            cell.fill = PatternFill(start_color='e1e6ff', end_color='e1e6ff', fill_type='solid')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # הוספת פריטים
        data_row = header_row + 1
        for index, item in enumerate(selected_items, 1):
            ws.cell(row=data_row, column=1, value=index)
            ws.cell(row=data_row, column=2, value=item)
            
            # עיצוב שורות הנתונים
            for col in range(1, 3):
                cell = ws.cell(row=data_row, column=col)
                cell.alignment = Alignment(horizontal='center' if col == 1 else 'right', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            data_row += 1
        
        # התאמת רוחב עמודות
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        
        # שמירת הקובץ
        filename = f"equipment_{event_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(app.root_path, 'static', filename)
        wb.save(filepath)
        
        # שליחת הקובץ
        event_name = event.get('name', 'אירוע').replace('/', '_')
        return send_file(filepath, as_attachment=True, download_name=f"ציוד_{event_name}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
    except Exception as e:
        import traceback
        print(f"Error exporting equipment: {e}")
        print(traceback.format_exc())
        return f"שגיאה בייצוא הציוד: {str(e)}", 500

@app.route('/add_event_charge/<event_id>', methods=['POST'])
@login_required
def add_event_charge(event_id):
    """הוספת חיוב לאירוע - מסנכרן עם לקוח"""
    try:
        events_list = load_events()
        data = load_data()
        
        for event in events_list:
            if event['id'] == event_id:
                client_id = event.get('client_id', '')
                if not client_id:
                    return "אין לקוח משויך לאירוע", 400
                
                our_cost = float(request.form.get('our_cost', 0) or 0)
                
                # מצא את הלקוח ליצירת מספר חיוב
                client_for_charge = next((cl for cl in data if cl['id'] == client_id), None)
                charge_number = get_next_charge_number(client_for_charge) if client_for_charge else None
                
                charge = {
                    'id': str(uuid.uuid4()),
                    'title': request.form.get('title', ''),
                    'amount': float(request.form.get('amount', 0)),
                    'our_cost': our_cost,  # עלות שלנו - לא מתווספת ללקוח
                    'date': datetime.now().strftime('%d/%m/%y'),
                    'completed': False,  # ברירת מחדל: לא הושלם
                    'charge_number': charge_number
                }
                
                # יצירת עותק לחיוב הלקוח (ללא our_cost)
                client_charge = {
                    'id': charge['id'],
                    'title': charge['title'],
                    'amount': charge['amount'],
                    'date': charge['date'],
                    'completed': charge['completed'],
                    'charge_number': charge_number
                }
                
                # הוספה לאירוע
                if 'charges' not in event:
                    event['charges'] = []
                event['charges'].append(charge)
                
                # סנכרון עם הלקוח (רק amount, ללא our_cost)
                for client in data:
                    if client['id'] == client_id:
                        if 'extra_charges' not in client:
                            client['extra_charges'] = []
                        client['extra_charges'].append(client_charge)
                        break
                
                save_events(events_list)
                save_data(data)
                break
        tab = request.form.get('tab', 'charges')
        return redirect(url_for('event_page', event_id=event_id, tab=tab))
    except Exception as e:
        return f"שגיאה בהוספת החיוב: {str(e)}", 500

@app.route('/edit_event_charge/<event_id>', methods=['POST'])
@login_required
def edit_event_charge(event_id):
    """עריכת חיוב באירוע"""
    try:
        events_list = load_events()
        data = load_data()
        charge_id = request.form.get('charge_id')
        
        if not charge_id:
            return "ID חיוב לא צוין", 400
        
        for event in events_list:
            if event['id'] == event_id:
                client_id = event.get('client_id', '')
                if not client_id:
                    return "אין לקוח משויך לאירוע", 400
                
                # עדכון החיוב באירוע
                for charge in event.get('charges', []):
                    if charge.get('id') == charge_id:
                        our_cost = float(request.form.get('our_cost', 0) or 0)
                        
                        charge['title'] = request.form.get('title', charge.get('title', ''))
                        charge['amount'] = float(request.form.get('amount', 0))
                        charge['our_cost'] = our_cost
                        
                        # עדכון החיוב גם בלקוח (רק amount, ללא our_cost)
                        for client in data:
                            if client['id'] == client_id:
                                if 'extra_charges' in client:
                                    for client_charge in client['extra_charges']:
                                        if client_charge.get('id') == charge_id:
                                            client_charge['title'] = charge['title']
                                            client_charge['amount'] = charge['amount']
                                            break
                                break
                        
                        save_events(events_list)
                        save_data(data)
                        break
                break
        
        tab = request.form.get('tab', 'charges')
        return redirect(url_for('event_page', event_id=event_id, tab=tab))
    except Exception as e:
        return f"שגיאה בעריכת החיוב: {str(e)}", 500

@app.route('/toggle_event_active/<event_id>', methods=['POST'])
@login_required
@csrf.exempt
def toggle_event_active(event_id):
    """מעדכן את סטטוס הפעיל/לא פעיל של אירוע"""
    try:
        user_role = get_user_role(current_user.id)
        # רק מנהל ואדמין יכולים לעדכן את הסטטוס
        if not is_manager_or_admin(current_user.id, user_role):
            return jsonify({'success': False, 'error': 'גישה חסומה - אין לך הרשאה לבצע פעולה זו'}), 403
        
        data = request.get_json()
        is_active = data.get('active', True)
        
        events_list = load_events()
        for event in events_list:
            if event.get('id') == event_id:
                if is_active:
                    # הפעל את האירוע - הסר את הסטטוס מאויש
                    event['archived'] = False
                    if 'archived_at' in event:
                        del event['archived_at']
                else:
                    # העבר לארכיון
                    event['archived'] = True
                    event['archived_at'] = datetime.now().isoformat()
                
                save_events(events_list)
                return jsonify({'success': True})
        
        return jsonify({'success': False, 'error': 'אירוע לא נמצא'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

def load_permissions():
    """טעינת הרשאות דפים"""
    if not os.path.exists(PERMISSIONS_FILE):
        # ברירת מחדל - כל הדפים פתוחים לכולם
        default_permissions = {
            '/': 'עובד',
            '/all_clients': 'עובד',
            '/client/': 'עובד',
            '/finance': 'עובד',
            '/events': 'עובד',
            '/suppliers': 'עובד',
            '/quotes': 'עובד',
            '/forms': 'עובד',
            '/admin/dashboard': 'מנהל',
            '/admin/users': 'אדמין'
        }
        save_permissions(default_permissions)
        return default_permissions
    with open(PERMISSIONS_FILE, 'r', encoding='utf-8') as f: 
        return json.load(f)

def save_permissions(permissions):
    with open(PERMISSIONS_FILE, 'w', encoding='utf-8') as f: 
        json.dump(permissions, f, ensure_ascii=False, indent=4)

def get_user_role(user_id):
    """מחזיר את התפקיד של המשתמש"""
    users = load_users()
    if user_id in users:
        return users[user_id].get('role', 'עובד')
    return 'עובד'

def is_manager_or_admin(user_id, user_role):
    """בודק אם המשתמש הוא מנהל או אדמין"""
    return user_id == 'admin' or user_role in ['מנהל', 'אדמין']

def normalize_assigned_user(assigned):
    """מנרמל את assigned_user לרשימה - תומך גם ב-string וגם ב-list"""
    if isinstance(assigned, str):
        return [assigned] if assigned else []
    elif isinstance(assigned, list):
        return assigned
    else:
        return []

def can_user_access_client(user_id, user_role, client):
    """בודק אם משתמש יכול לגשת ללקוח מסוים"""
    if is_manager_or_admin(user_id, user_role):
        return True
    assigned = normalize_assigned_user(client.get('assigned_user', []))
    # בדיקה case-insensitive
    user_id_lower = user_id.lower() if isinstance(user_id, str) else str(user_id).lower()
    for assigned_uid in assigned:
        assigned_uid_lower = assigned_uid.lower() if isinstance(assigned_uid, str) else str(assigned_uid).lower()
        if user_id == assigned_uid or user_id_lower == assigned_uid_lower:
            return True
    return False

def filter_active_clients(clients):
    """מסנן לקוחות פעילים (לא מאוישים)"""
    return [c for c in clients if not c.get('archived', False)]

def filter_archived_clients(clients):
    """מסנן לקוחות מאוישים בלבד"""
    return [c for c in clients if c.get('archived', False)]

def filter_active_events(events):
    """מסנן אירועים פעילים (לא מאוישים)"""
    return [e for e in events if not e.get('archived', False)]

def filter_archived_events(events):
    """מסנן אירועים מאוישים בלבד"""
    return [e for e in events if e.get('archived', False)]

def check_permission(route_path, user_role):
    """בודק אם למשתמש יש הרשאה לגשת לדף מסוים
    עובד - כולם רואים
    מנהל - רק מנהל ואדמין
    אדמין - רק אדמין
    """
    permissions = load_permissions()
    
    # מצא את ההרשאה המתאימה (תמיד בדוק גם routes שמתחילים ב-route_path)
    required_role = None
    for route, role in permissions.items():
        if route_path == route or route_path.startswith(route):
            required_role = role
            break
    
    # אם לא נמצא, ברירת מחדל - כולם יכולים
    if not required_role:
        return True
    
    # בדיקת הרשאות
    if required_role == 'עובד':
        return True  # כולם יכולים
    elif required_role == 'מנהל':
        return user_role in ['מנהל', 'אדמין']
    elif required_role == 'אדמין':
        return user_role == 'אדמין'
    
    return False


@app.route('/api/admin/users')
@login_required
def api_admin_users():
    """API endpoint להחזרת נתוני ניהול משתמשים"""
    try:
        if current_user.id != 'admin':
            return jsonify({'success': False, 'error': 'גישה חסומה'}), 403
        
        users = load_users()
        clients = load_data()
        permissions = load_permissions()
        
        all_pages = [
            {'route': '/', 'name': 'דשבורד ראשי'},
            {'route': '/all_clients', 'name': 'לוח לקוחות'},
            {'route': '/client/', 'name': 'תיק לקוח'},
            {'route': '/finance', 'name': 'כספים'},
            {'route': '/events', 'name': 'אירועים'},
            {'route': '/suppliers', 'name': 'ספקים'},
            {'route': '/quotes', 'name': 'הצעות מחיר'},
            {'route': '/forms', 'name': 'טפסים'},
            {'route': '/admin/dashboard', 'name': 'דוח מנהלים'},
            {'route': '/admin/users', 'name': 'ניהול צוות'},
        ]
        
        users_list = [
            {
                'id': uid,
                'name': info.get('name', ''),
                'email': info.get('email', ''),
                'role': info.get('role', 'עובד'),
            }
            for uid, info in users.items()
        ]
        
        clients_list = [
            {
                'id': c['id'],
                'name': c.get('name', ''),
                'assigned_user': c.get('assigned_user', []),
            }
            for c in clients
        ]
        
        return jsonify({
            'success': True,
            'users': users_list,
            'clients': clients_list,
            'permissions': permissions,
            'pages': all_pages,
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/users', methods=['GET', 'POST'])
@login_required
@csrf.exempt  # פטור זמני מ-CSRF עד שנוסיף tokens לכל הטפסים
def manage_users():
    if current_user.id != 'admin': return "גישה חסומה", 403
    users, clients = load_users(), load_data()
    permissions = load_permissions()
    
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add_user':
            username = request.form.get('username')
            email = request.form.get('email', '').strip()
            email_password = request.form.get('email_password', '').strip()
            
            # הצפנת סיסמת המייל (base64) - צריך את הסיסמה המקורית לשליחת מיילים
            email_password_encoded = base64.b64encode(email_password.encode('utf-8')).decode('utf-8') if email_password else ''
            
            users[username] = {
                'password': generate_password_hash(request.form.get('password')), 
                'name': request.form.get('name'),
                'email': email,
                'email_password': email_password_encoded,  # סיסמת מייל מוצפנת (base64)
                'role': request.form.get('role', 'עובד')
            }
            save_users(users)
        elif action == 'reset_password':
            user_id = request.form.get('user_id')
            new_password = request.form.get('new_password')
            if user_id in users and user_id != 'admin':
                users[user_id]['password'] = generate_password_hash(new_password)
                save_users(users)
                if request.is_json or request.headers.get('Accept') == 'application/json':
                    return jsonify({'success': True, 'message': 'סיסמה אופסה בהצלחה'})
        elif action == 'update_role':
            user_id = request.form.get('user_id')
            new_role = request.form.get('role')
            if user_id in users:
                users[user_id]['role'] = new_role
                save_users(users)
                if request.is_json or request.headers.get('Accept') == 'application/json':
                    return jsonify({'success': True, 'message': 'תפקיד עודכן בהצלחה'})
        elif action == 'update_email':
            user_id = request.form.get('user_id')
            new_email = request.form.get('email', '').strip()
            if user_id in users:
                if new_email:
                    users[user_id]['email'] = new_email
                else:
                    # אם ריק, מחק את השדה
                    users[user_id].pop('email', None)
                save_users(users)
                if request.is_json or request.headers.get('Accept') == 'application/json':
                    return jsonify({'success': True, 'message': 'מייל עודכן בהצלחה'})
                flash(f'מייל עודכן בהצלחה', 'success')
        elif action == 'update_email_password':
            user_id = request.form.get('user_id')
            new_email_password = request.form.get('email_password', '').strip()
            if user_id in users:
                if new_email_password:
                    # הצפנת סיסמת המייל (base64)
                    email_password_encoded = base64.b64encode(new_email_password.encode('utf-8')).decode('utf-8')
                    users[user_id]['email_password'] = email_password_encoded
                else:
                    # אם ריק, מחק את השדה
                    users[user_id].pop('email_password', None)
                save_users(users)
                if request.is_json or request.headers.get('Accept') == 'application/json':
                    return jsonify({'success': True, 'message': 'סיסמת מייל עודכנה בהצלחה'})
                flash(f'סיסמת מייל עודכנה בהצלחה', 'success')
        elif action == 'update_permission':
            route = request.form.get('route')
            required_role = request.form.get('required_role')
            permissions[route] = required_role
            save_permissions(permissions)
            if request.is_json or request.headers.get('Accept') == 'application/json':
                return jsonify({'success': True, 'message': 'הרשאה עודכנה בהצלחה'})
        elif action == 'delete_user':
            user_id = request.form.get('user_id')
            if not user_id or user_id == 'admin' or user_id not in users:
                flash('לא ניתן למחוק משתמש זה', 'error')
            else:
                user_name = users[user_id].get('name', user_id)
                deleted = False
                if USE_DATABASE:
                    deleted = delete_user_record(user_id)
                    if deleted:
                        users.pop(user_id, None)
                else:
                    deleted = True
                    users.pop(user_id, None)
                    save_users(users)

                if deleted:
                    for c in clients:
                        if 'assigned_user' in c:
                            if isinstance(c['assigned_user'], list):
                                if user_id in c['assigned_user']:
                                    c['assigned_user'] = [uid for uid in c['assigned_user'] if uid != user_id]
                            elif c['assigned_user'] == user_id:
                                c['assigned_user'] = []
                    save_data(clients)
                    flash(f'המשתמש "{user_name}" נמחק בהצלחה', 'success')
                else:
                    flash('לא ניתן למחוק משתמש זה', 'error')
        elif action == 'assign':
            user_ids = request.form.getlist('user_ids')  # מקבל רשימת משתמשים
            cid = request.form.get('client_id')
            client_found = False
            client_name = None
            for c in clients:
                if c['id'] == cid:
                    client_found = True
                    client_name = c.get('name', 'Unknown')
                    # עדכן את הרשימה - תמיד list
                    c['assigned_user'] = user_ids
                    break
            
            if client_found:
                save_data(clients)
                # בניית הודעת הצלחה
                if user_ids:
                    user_names = [users.get(uid, {}).get('name', uid) for uid in user_ids if uid in users]
                    flash(f'הלקוח "{client_name}" שויך בהצלחה לעובדים: {", ".join(user_names)}', 'success')
                else:
                    flash(f'השיוך של הלקוח "{client_name}" הוסר', 'info')
            else:
                flash('שגיאה: לקוח לא נמצא', 'error')
        return redirect(url_for('manage_users'))
    
    # רשימת כל הדפים במערכת
    all_pages = [
        {'route': '/', 'name': 'דשבורד ראשי'},
        {'route': '/all_clients', 'name': 'לוח לקוחות'},
        {'route': '/client/', 'name': 'תיק לקוח'},
        {'route': '/finance', 'name': 'כספים'},
        {'route': '/events', 'name': 'אירועים'},
        {'route': '/suppliers', 'name': 'ספקים'},
        {'route': '/quotes', 'name': 'הצעות מחיר'},
        {'route': '/forms', 'name': 'טפסים'},
        {'route': '/admin/dashboard', 'name': 'דוח מנהלים'},
        {'route': '/admin/users', 'name': 'ניהול צוות'}
    ]
    
    # Redirect to React manage users page
    return redirect('/app/admin/users')

# --- Forms Routes ---
@app.route('/api/forms')
@login_required
def api_forms():
    """API endpoint להחזרת טפסים"""
    try:
        user_role = get_user_role(current_user.id)
        if not check_permission('/forms', user_role):
            return jsonify({'success': False, 'error': 'גישה חסומה'}), 403
        
        forms_list = load_forms()
        clients = load_data()
        clients_dict = {c['id']: c['name'] for c in clients}
        
        # Add client names to forms
        for form in forms_list:
            client_id = form.get('client_id', '')
            form['client_name'] = clients_dict.get(client_id, '')
        
        return jsonify({
            'success': True,
            'forms': forms_list
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/forms')
@login_required
def forms():
    """עמוד ניהול טפסים"""
    user_role = get_user_role(current_user.id)
    if not check_permission('/forms', user_role):
        return "גישה חסומה - אין לך הרשאה לגשת לדף זה", 403
    forms_list = load_forms()
    clients = load_data()
    # Map client IDs to client names
    clients_dict = {c['id']: c['name'] for c in clients}
    # Redirect to React forms page
    return redirect('/app/forms')

@app.route('/add_form', methods=['POST'])
@login_required
def add_form():
    """יצירת טופס חדש"""
    try:
        forms_list = load_forms()
        form_data = request.get_json() if request.is_json else request.form
        
        # יצירת token ייחודי לטופס
        form_token = str(uuid.uuid4())
        
        new_form = {
            'id': str(uuid.uuid4()),
            'token': form_token,
            'client_id': form_data.get('client_id', ''),
            'title': form_data.get('title', ''),
            'fields': form_data.get('fields', []),  # רשימת שדות
            'created_date': datetime.now().strftime('%d/%m/%y %H:%M')
        }
        
        forms_list.append(new_form)
        save_forms(forms_list)
        
        if request.is_json or request.headers.get('Accept') == 'application/json':
            return jsonify({'success': True, 'form': new_form})
        return redirect(url_for('forms'))
    except Exception as e:
        if request.is_json or request.headers.get('Accept') == 'application/json':
            return jsonify({'success': False, 'error': str(e)}), 500
        return f"שגיאה ביצירת הטופס: {str(e)}", 500

@app.route('/edit_form/<form_id>', methods=['POST'])
@login_required
def edit_form(form_id):
    """עריכת טופס"""
    try:
        forms_list = load_forms()
        form_data = request.get_json() if request.is_json else request.form
        
        for form in forms_list:
            if form['id'] == form_id:
                form['title'] = form_data.get('title', form['title'])
                form['client_id'] = form_data.get('client_id', form['client_id'])
                form['fields'] = form_data.get('fields', form.get('fields', []))
                save_forms(forms_list)
                if request.is_json:
                    return jsonify({'status': 'success', 'form': form})
                return redirect(url_for('forms'))
        
        return jsonify({'status': 'error', 'error': 'טופס לא נמצא'}), 404
    except Exception as e:
        if request.is_json:
            return jsonify({'status': 'error', 'error': str(e)}), 500
        return f"שגיאה בעדכון הטופס: {str(e)}", 500

@app.route('/delete_form/<form_id>', methods=['POST'])
@login_required
def delete_form(form_id):
    """מחיקת טופס"""
    try:
        forms_list = load_forms()
        forms_list = [f for f in forms_list if f['id'] != form_id]
        save_forms(forms_list)
        return redirect(url_for('forms'))
    except Exception as e:
        return f"שגיאה במחיקת הטופס: {str(e)}", 500

@app.route('/form/<form_token>', methods=['GET'])
def public_form(form_token):
    """תצוגת טופס ציבורי (ללא login)"""
    forms_list = load_forms()
    form = next((f for f in forms_list if f.get('token') == form_token), None)
    
    if not form:
        return "טופס לא נמצא", 404
    
    clients = load_data()
    client = next((c for c in clients if c['id'] == form.get('client_id')), None)
    client_name = client['name'] if client else 'לא משויך'
    
    return render_template('public_form.html', form=form, client_name=client_name)

@app.route('/download_form_file/<filename>')
@login_required
def download_form_file(filename):
    """Route להורדת קבצים שהועלו דרך טפסים (דורש login)"""
    try:
        filepath = os.path.join(FORMS_UPLOAD_FOLDER, filename)
        if not os.path.exists(filepath):
            return "קובץ לא נמצא", 404
        return send_file(filepath, as_attachment=True)
    except Exception as e:
        return f"שגיאה: {str(e)}", 500

@app.route('/download_form_file_public/<filename>')
def download_form_file_public(filename):
    """Route ציבורי להורדת קבצים שהועלו דרך טפסים (ללא login - לשימוש במיילים)"""
    try:
        filepath = os.path.join(FORMS_UPLOAD_FOLDER, filename)
        if not os.path.exists(filepath):
            return "קובץ לא נמצא", 404
        
        # חילוץ שם הקובץ המקורי (אם נשמר)
        # הקבצים נשמרים עם UUID_originalname, אז נשמור רק את החלק האחרון
        original_name = filename.split('_', 1)[1] if '_' in filename else filename
        return send_file(filepath, as_attachment=True, download_name=original_name)
    except Exception as e:
        return f"שגיאה: {str(e)}", 500

@app.route('/submit_form/<form_token>', methods=['POST'])
@csrf.exempt  # פטור מ-CSRF כי זה טופס ציבורי (ללא login)
def submit_form(form_token):
    """שליחת טופס (ללא login)"""
    print(f"\n[DEBUG] Form submission received for token: {form_token}")
    try:
        forms_list = load_forms()
        form = next((f for f in forms_list if f.get('token') == form_token), None)
        print(f"[DEBUG] Form found: {form is not None}")
        
        if not form:
            return jsonify({'status': 'error', 'error': 'טופס לא נמצא'}), 404
        
        client_id = form.get('client_id')
        if not client_id:
            return jsonify({'status': 'error', 'error': 'טופס לא משויך ללקוח'}), 400
        
        # קבלת נתונים מהטופס
        form_submission = {}
        uploaded_files = {}
        
        # עיבוד שדות רגילים
        for field in form.get('fields', []):
            field_id = field.get('id')
            field_name = f"field_{field_id}"
            
            if field.get('type') == 'file':
                # טיפול בהעלאת קבצים
                if field_name in request.files:
                    file = request.files[field_name]
                    if file and file.filename:
                        filename = secure_filename(f"{uuid.uuid4().hex}_{file.filename}")
                        filepath = os.path.join(FORMS_UPLOAD_FOLDER, filename)
                        file.save(filepath)
                        uploaded_files[field_id] = {
                            'label': field.get('label', 'קובץ'),
                            'filename': file.filename,
                            'saved_filename': filename
                        }
            else:
                value = request.form.get(field_name, '')
                form_submission[field_id] = {
                    'label': field.get('label', ''),
                    'value': value
                }
        
        # יצירת משימה חדשה ללקוח
        data = load_data()
        task_title = f"טופס התקבל: {form.get('title', '')}"
        
        # בניית תוכן המשימה מנתוני הטופס
        task_note_parts = []
        for field_id, field_data in form_submission.items():
            task_note_parts.append(f"{field_data['label']}: {field_data['value']}")
        
        if uploaded_files:
            task_note_parts.append("\nקבצים שהועלו:")
            for field_id, file_data in uploaded_files.items():
                task_note_parts.append(f"{file_data['label']}: {file_data['filename']}")
        
        task_note = "\n".join(task_note_parts)
        
        # חיפוש הלקוח ויצירת פרויקט אם צריך
        client_name = 'לא ידוע'
        for client in data:
            if client['id'] == client_id:
                client_name = client.get('name', 'לא ידוע')
                # חיפוש פרויקט "טפסים" או יצירתו
                projects = client.get('projects', [])
                forms_project = next((p for p in projects if p.get('title') == 'טפסים'), None)
                
                if not forms_project:
                    forms_project = {
                        'id': str(uuid.uuid4()),
                        'title': 'טפסים',
                        'tasks': []
                    }
                    client.setdefault('projects', []).append(forms_project)
                
                # הוספת המשימה
                task = {
                    'id': str(uuid.uuid4()),
                    'title': task_title,
                    'status': 'לביצוע',
                    'note': task_note,
                    'created_date': datetime.now().strftime('%d/%m/%y'),
                    'done': False
                }
                forms_project['tasks'].append(task)
                save_data(data)
                
                # שליחת מייל
                email_result = send_form_email(
                    form_title=form.get('title', ''),
                    client_name=client_name,
                    form_submission=form_submission,
                    uploaded_files=uploaded_files,
                    form_token=form_token
                )
                
                # הוספת אישור שליחת מייל למשימה
                if email_result and isinstance(email_result, dict):
                    email_confirmation = f"\n\n{'='*50}\n📧 אישור שליחת מייל:\n"
                    email_confirmation += f"✅ מייל נשלח בהצלחה!\n"
                    email_confirmation += f"📤 נשלח מ: {email_result.get('from_email', 'לא ידוע')}\n"
                    email_confirmation += f"📥 נשלח ל: {email_result.get('to_email', 'לא ידוע')}\n"
                    email_confirmation += f"🕐 תאריך ושעה: {email_result.get('sent_at', 'לא ידוע')}\n"
                    email_confirmation += f"{'='*50}"
                    task_note += email_confirmation
                    task['note'] = task_note
                    save_data(data)  # שמירה מחדש עם האישור
                elif not email_result:
                    email_warning = f"\n\n{'='*50}\n⚠️ אזהרה:\n"
                    email_warning += f"המשימה נוצרה אך המייל לא נשלח - בדוק הגדרות SMTP\n"
                    email_warning += f"{'='*50}"
                    task_note += email_warning
                    task['note'] = task_note
                    save_data(data)
                
                return jsonify({
                    'status': 'success',
                    'message': 'הטופס נשלח בהצלחה! תודה.'
                })
        
        return jsonify({'status': 'error', 'error': 'לקוח לא נמצא'}), 404
        
    except Exception as e:
        import traceback
        print(f"[ERROR] Exception in submit_form: {e}")
        traceback.print_exc()
        return jsonify({'status': 'error', 'error': str(e), 'traceback': traceback.format_exc()}), 500

@app.route('/update_event_graphics/<event_id>', methods=['POST'])
@login_required
def update_event_graphics(event_id):
    """עדכון פריטי גרפיקה באירוע"""
    try:
        events_list = load_events()
        event = next((e for e in events_list if e.get('id') == event_id), None)
        
        if not event:
            return jsonify({'success': False, 'error': 'אירוע לא נמצא'}), 404
        
        if request.is_json:
            data = request.get_json()
            graphics_items = data.get('graphics_items', [])
        else:
            graphics_items = json.loads(request.form.get('graphics_items', '[]'))
        
        event['graphics_items'] = graphics_items
        save_events(events_list)
        
        return jsonify({'success': True})
    except Exception as e:
        print(f"Error updating graphics: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/export_event_graphics/<event_id>')
@login_required
def export_event_graphics(event_id):
    """ייצוא טבלת גרפיקה לאקסל"""
    try:
        events_list = load_events()
        event = next((e for e in events_list if e.get('id') == event_id), None)
        
        if not event:
            return "אירוע לא נמצא", 404
        
        graphics_items = event.get('graphics_items', [])
        
        # יצירת Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "טבלת גרפיקה"
        
        # כותרת
        ws['A1'] = f"טבלת גרפיקה - {event.get('name', 'אירוע')}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='right', vertical='center')
        ws.merge_cells('A1:G1')
        
        # כותרות
        headers = ['שם הפריט', 'מידות', 'כמות', 'הערות לסטודיו', 'סטטוס', 'הגיע?']
        header_row = 3
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='0073ea', end_color='0073ea', fill_type='solid')
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # נתונים
        data_row = header_row + 1
        for item in graphics_items:
            ws.cell(row=data_row, column=1, value=item.get('name', ''))
            ws.cell(row=data_row, column=2, value=item.get('dimensions', ''))
            ws.cell(row=data_row, column=3, value=item.get('quantity', 1))
            ws.cell(row=data_row, column=4, value=item.get('studio_notes', ''))
            ws.cell(row=data_row, column=5, value=item.get('status', ''))
            ws.cell(row=data_row, column=6, value='כן' if item.get('arrived', False) else 'לא')
            
            # עיצוב שורות
            for col in range(1, 7):
                cell = ws.cell(row=data_row, column=col)
                cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            data_row += 1
        
        # התאמת רוחב עמודות
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 10
        
        # שמירת הקובץ
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'graphics_{event_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        print(f"Error exporting graphics: {e}")
        import traceback
        traceback.print_exc()
        return f"שגיאה בייצוא: {str(e)}", 500

@app.route('/api/tasks/notifications')
@login_required
def get_task_notifications():
    """מחזיר התראות על משימות עם deadlines קרובים"""
    try:
        data = load_data()
        users = load_users()
        current_date = datetime.now().date()
        notifications = {
            'urgent': [],  # deadline עבר או היום
            'warning': [],  # 1-3 ימים
            'approaching': []  # 4-7 ימים
        }
        
        for client in data:
            client_name = client.get('name', 'ללא שם')
            for project in client.get('projects', []):
                project_title = project.get('title', 'ללא שם')
                for task in project.get('tasks', []):
                    if task.get('status') != 'הושלם' and task.get('deadline'):
                        try:
                            deadline_str = task.get('deadline')
                            if 'T' in deadline_str:
                                deadline_date = datetime.fromisoformat(deadline_str).date()
                            else:
                                deadline_date = datetime.strptime(deadline_str[:10], '%Y-%m-%d').date()
                            
                            days_diff = (deadline_date - current_date).days
                            
                            if days_diff < 0:
                                notifications['urgent'].append({
                                    'client_id': client.get('id'),
                                    'client_name': client_name,
                                    'project_id': project.get('id'),
                                    'project_title': project_title,
                                    'task_id': task.get('id'),
                                    'task_title': task.get('title', 'ללא שם'),
                                    'deadline': deadline_str,
                                    'days_overdue': abs(days_diff),
                                    'priority': task.get('priority', 'medium')
                                })
                            elif days_diff == 0:
                                notifications['urgent'].append({
                                    'client_id': client.get('id'),
                                    'client_name': client_name,
                                    'project_id': project.get('id'),
                                    'project_title': project_title,
                                    'task_id': task.get('id'),
                                    'task_title': task.get('title', 'ללא שם'),
                                    'deadline': deadline_str,
                                    'days_overdue': 0,
                                    'priority': task.get('priority', 'medium')
                                })
                            elif days_diff <= 3:
                                notifications['warning'].append({
                                    'client_id': client.get('id'),
                                    'client_name': client_name,
                                    'project_id': project.get('id'),
                                    'project_title': project_title,
                                    'task_id': task.get('id'),
                                    'task_title': task.get('title', 'ללא שם'),
                                    'deadline': deadline_str,
                                    'days_remaining': days_diff,
                                    'priority': task.get('priority', 'medium')
                                })
                            elif days_diff <= 7:
                                notifications['approaching'].append({
                                    'client_id': client.get('id'),
                                    'client_name': client_name,
                                    'project_id': project.get('id'),
                                    'project_title': project_title,
                                    'task_id': task.get('id'),
                                    'task_title': task.get('title', 'ללא שם'),
                                    'deadline': deadline_str,
                                    'days_remaining': days_diff,
                                    'priority': task.get('priority', 'medium')
                                })
                        except Exception as e:
                            print(f"Error processing deadline for task {task.get('id')}: {e}")
                            continue
        
        return jsonify({
            'status': 'success',
            'notifications': notifications,
            'total': len(notifications['urgent']) + len(notifications['warning']) + len(notifications['approaching'])
        })
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/admin/open_tasks')
@login_required
def admin_open_tasks():
    """Route להחזרת משימות פתוחות למנהל"""
    if current_user.id != 'admin':
        return jsonify({'error': 'גישה חסומה'}), 403
    
    try:
        data = load_data()
        users = load_users()
        
        open_tasks = []
        for client in data:
            # תמיכה ב-list או string ישן
            assigned_users_list = normalize_assigned_user(client.get('assigned_user', []))
            
            # צור רשימת שמות
            assigned_user_names = [users.get(uid, {}).get('name', uid) for uid in assigned_users_list if uid]
            assigned_user_name = ', '.join(assigned_user_names) if assigned_user_names else 'לא שויך'
            
            for project in client.get('projects', []):
                for task in project.get('tasks', []):
                    # משימות פתוחות (לא הושלמו)
                    if not task.get('done', False) and task.get('status') != 'הושלם':
                        # הוסף משימה אחת עם כל המשתמשים
                        open_tasks.append({
                            'task_id': task.get('id', ''),
                            'task_title': task.get('title', 'ללא שם'),
                            'task_number': task.get('task_number', ''),
                            'task_status': task.get('status', 'לביצוע'),
                            'task_note': task.get('note', ''),
                            'manager_note': task.get('manager_note', ''),
                            'created_date': task.get('created_date', ''),
                            'client_id': client.get('id', ''),
                            'client_name': client.get('name', 'ללא שם'),
                            'project_id': project.get('id', ''),
                            'project_title': project.get('title', 'ללא שם'),
                            'assigned_user': assigned_users_list[0] if assigned_users_list else 'admin',  # משתמש ראשון לתאימות
                            'assigned_user_name': assigned_user_name,
                            'assigned_users': assigned_users_list  # רשימה מלאה
                        })
        
        # מיון לפי תאריך יצירה - החדש ביותר ראשון
        open_tasks.sort(key=lambda x: x.get('created_date', ''), reverse=True)
        
        return jsonify({
            'success': True,
            'tasks': open_tasks
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/add_manager_note/<client_id>/<project_id>/<task_id>', methods=['POST'])
@login_required
def add_manager_note(client_id, project_id, task_id):
    """הוספת הערת מנהל למשימה ושילוח התראה למשתמש האחראי"""
    if current_user.id != 'admin':
        return jsonify({'error': 'גישה חסומה'}), 403
    
    try:
        data = load_data()
        users = load_users()
        messages_list = load_messages()
        
        if request.is_json:
            manager_note = request.json.get('manager_note', '').strip()
        else:
            manager_note = request.form.get('manager_note', '').strip()
        
        # מציאת המשימה ועדכון ההערה
        client_found = None
        task_found = None
        assigned_users_list = []
        
        for c in data:
            if c['id'] == client_id:
                client_found = c
                # תמיכה ב-list או string ישן
                assigned_users_list = normalize_assigned_user(c.get('assigned_user', []))
                for p in c.get('projects', []):
                    if p['id'] == project_id:
                        for t in p.get('tasks', []):
                            if t['id'] == task_id:
                                task_found = t
                                t['manager_note'] = manager_note
                                # אם יש הערה של מנהל, נוסיף גם timestamp
                                if manager_note:
                                    t['manager_note_date'] = datetime.now().isoformat()
                                break
                        break
                break
        
        if not task_found:
            return jsonify({'success': False, 'error': 'משימה לא נמצאה'}), 404
        
        save_data(data)
        
        # שליחת התראה לכל המשתמשים שויכו ללקוח (אם לא זה admin)
        if assigned_users_list and manager_note and client_found:
            client_name = client_found.get('name', 'לקוח')
            task_title = task_found.get('title', 'משימה')
            project_title = next((p.get('title', 'פרויקט') for p in client_found.get('projects', []) if p.get('id') == project_id), 'פרויקט')
            
            # יצירת הודעה בצ'אט עם קישור יפה לכל משתמש
            for assigned_user in assigned_users_list:
                if assigned_user != 'admin':
                    message_content = f"המנהל הוסיף הערה למשימה '{task_title}' בפרויקט '{project_title}' של הלקוח '{client_name}':\n\n{manager_note}\n\n👉 <a href='/client/{client_id}'>פתח תיק לקוח</a>"
                    
                    chat_message = {
                        'id': str(uuid.uuid4()),
                        'from_user': 'admin',
                        'to_user': assigned_user,
                        'subject': '',
                        'content': message_content,
                        'project_id': project_id,
                        'task_id': task_id,
                        'client_id': client_id,
                        'created_date': datetime.now().strftime('%d/%m/%y %H:%M'),
                        'read': False,
                        'files': [],
                        'is_manager_note': True
                    }
                    messages_list.append(chat_message)
            save_messages(messages_list)
        
        return jsonify({
            'success': True,
            'message': 'הערת המנהל נוספה והודעה נשלחה'
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/stats')
@login_required
def admin_stats():
    """Route להחזרת נתונים סטטיסטיים למנהל"""
    if current_user.id != 'admin':
        return jsonify({'error': 'גישה חסומה'}), 403
    
    try:
        data = load_data()
        users = load_users()
        events_list = load_events()
        
        # חישוב משימות פעילות לכל עובד
        user_task_counts = {}
        user_names = {}
        for uid, user_info in users.items():
            if uid != 'admin':
                user_task_counts[uid] = 0
                user_names[uid] = user_info.get('name', uid)
        
        # חישוב משימות שנפתחו/נסגרו ב-7 הימים האחרונים
        seven_days_ago = datetime.now() - timedelta(days=7)
        tasks_opened = {i: 0 for i in range(7)}  # 0 = היום, 6 = לפני 6 ימים
        tasks_closed = {i: 0 for i in range(7)}
        
        for client in data:
            # תמיכה ב-list או string ישן
            assigned_users_list = normalize_assigned_user(client.get('assigned_user', []))
            
            for project in client.get('projects', []):
                for task in project.get('tasks', []):
                    # משימות פעילות (לא הושלמו)
                    if not task.get('done', False) and task.get('status') != 'הושלם':
                        # הוסף את המשימה לכל משתמש שויך
                        for assigned_user in assigned_users_list:
                            if assigned_user in user_task_counts:
                                user_task_counts[assigned_user] += 1
                    
                    # תאריך יצירה
                    created_at = task.get('created_at')
                    if created_at:
                        try:
                            created_dt = datetime.fromisoformat(created_at.replace('Z', '+00:00') if 'Z' in created_at else created_at)
                            if created_dt >= seven_days_ago:
                                days_ago = (datetime.now() - created_dt).days
                                if 0 <= days_ago < 7:
                                    tasks_opened[days_ago] += 1
                        except:
                            pass
                    elif task.get('created_date'):
                        try:
                            date_str = task['created_date']
                            if '/' in date_str:
                                parts = date_str.split('/')
                                if len(parts) == 3:
                                    day, month, year = parts
                                    year = '20' + year if len(year) == 2 else year
                                    created_dt = datetime(int(year), int(month), int(day))
                                    if created_dt >= seven_days_ago:
                                        days_ago = (datetime.now() - created_dt).days
                                        if 0 <= days_ago < 7:
                                            tasks_opened[days_ago] += 1
                        except:
                            pass
                    
                    # תאריך סגירה
                    completed_at = task.get('completed_at')
                    if completed_at:
                        try:
                            completed_dt = datetime.fromisoformat(completed_at.replace('Z', '+00:00') if 'Z' in completed_at else completed_at)
                            if completed_dt >= seven_days_ago:
                                days_ago = (datetime.now() - completed_dt).days
                                if 0 <= days_ago < 7:
                                    tasks_closed[days_ago] += 1
                        except:
                            pass
        
        # בניית נתונים לעובדים
        user_loads = []
        for uid, count in user_task_counts.items():
            capacity = 20
            percentage = (count / capacity) * 100 if capacity > 0 else 0
            user_loads.append({
                'user_id': uid,
                'name': user_names[uid],
                'active_tasks': count,
                'capacity': capacity,
                'percentage': round(percentage, 1),
                'is_overloaded': percentage > 90
            })
        
        # בניית נתונים ל-line chart (7 ימים אחרונים)
        line_chart_data = {
            'labels': [],
            'opened': [],
            'closed': []
        }
        
        for i in range(6, -1, -1):  # מ-6 ימים לפני עד היום
            date = datetime.now() - timedelta(days=i)
            line_chart_data['labels'].append(date.strftime('%d/%m'))
            line_chart_data['opened'].append(tasks_opened[i])
            line_chart_data['closed'].append(tasks_closed[i])
        
        # אירועים קרובים (30 הימים הקרובים)
        upcoming_events = []
        today = datetime.now().date()
        thirty_days_later = today + timedelta(days=30)
        
        for event in events_list:
            event_date_str = event.get('date')
            if event_date_str:
                try:
                    event_date = datetime.strptime(event_date_str, '%Y-%m-%d').date()
                    if today <= event_date <= thirty_days_later:
                        # מציאת שם הלקוח
                        client_id = event.get('client_id', '')
                        client_name = next((c.get('name', 'לא צוין') for c in data if c.get('id') == client_id), 'לא צוין')
                        
                        upcoming_events.append({
                            'id': event.get('id', ''),
                            'name': event.get('name', 'ללא שם'),
                            'date': event_date_str,
                            'client_name': client_name,
                            'location': event.get('location', ''),
                            'event_type': event.get('event_type', '')
                        })
                except:
                    pass
        
        # מיון אירועים לפי תאריך
        upcoming_events.sort(key=lambda x: x['date'])
        
        return jsonify({
            'success': True,
            'user_loads': user_loads,
            'line_chart': line_chart_data,
            'upcoming_events': upcoming_events
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/dashboard')
@login_required
def api_admin_dashboard():
    """API endpoint להחזרת נתוני דשבורד מנהלים"""
    try:
        user_role = get_user_role(current_user.id)
        if not check_permission('/admin/dashboard', user_role):
            return jsonify({'success': False, 'error': 'גישה חסומה'}), 403
        
        data = load_data()
        users = load_users()
        events_list = load_events()
        
        # חישוב סטטיסטיקות משימות לפי עובד
        task_stats = []
        for uid, user_info in users.items():
            if uid == 'admin':
                continue
            
            total_tasks = 0
            completed_tasks = 0
            pending_tasks = 0
            
            for client in data:
                if not can_user_access_client(uid, user_info.get('role', 'עובד'), client):
                    continue
                
                for project in client.get('projects', []):
                    for task in project.get('tasks', []):
                        if task.get('assignee') == uid:
                            total_tasks += 1
                            status = task.get('status', '')
                            if status == 'הושלם':
                                completed_tasks += 1
                            else:
                                pending_tasks += 1
            
            task_stats.append({
                'user_id': uid,
                'user_name': user_info.get('name', uid),
                'total_tasks': total_tasks,
                'completed_tasks': completed_tasks,
                'pending_tasks': pending_tasks,
                'overloaded': pending_tasks > 10  # עומס יתר אם יש יותר מ-10 משימות ממתינות
            })
        
        # חישוב הכנסות חודשיות
        current_month = datetime.now().strftime('%m')
        current_year = datetime.now().strftime('%Y')
        monthly_revenue = 0
        for client in data:
            for charge in client.get('extra_charges', []):
                charge_date = charge.get('date', '')
                if charge_date:
                    date_parts = charge_date.split('/')
                    if len(date_parts) >= 3:
                        month = date_parts[1].zfill(2)
                        year_str = date_parts[2]
                        if len(year_str) == 2:
                            year = '20' + year_str
                        else:
                            year = year_str
                        if month == current_month and year == current_year:
                            monthly_revenue += charge.get('amount', 0)
        
        # חישוב לקוחות ופרויקטים פעילים
        active_clients = filter_active_clients(data)
        total_clients = len(active_clients)
        active_projects = sum(len(c.get('projects', [])) for c in active_clients)
        
        # אירועי לוח שנה
        calendar_events = []
        for client in active_clients:
            for project in client.get('projects', []):
                for task in project.get('tasks', []):
                    deadline = task.get('deadline', '')
                    if deadline:
                        try:
                            if '/' in deadline:
                                parts = deadline.split('/')
                                if len(parts) == 3:
                                    event_date = datetime.strptime(deadline, '%d/%m/%Y').strftime('%Y-%m-%d')
                                else:
                                    event_date = datetime.strptime(deadline, '%d/%m/%y').strftime('%Y-%m-%d')
                            else:
                                event_date = deadline
                            
                            calendar_events.append({
                                'title': task.get('title', 'ללא כותרת'),
                                'start': event_date,
                                'color': get_task_status_color(task.get('status', '')),
                                'extendedProps': {
                                    'client_name': client.get('name', ''),
                                    'project_title': project.get('title', ''),
                                }
                            })
                        except:
                            pass
        
        return jsonify({
            'success': True,
            'task_stats': task_stats,
            'calendar_events': calendar_events,
            'monthly_revenue': monthly_revenue,
            'total_clients': total_clients,
            'active_projects': active_projects,
        })
    except Exception as e:
        print(f"Error in api_admin_dashboard: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

def get_task_status_color(status: str) -> str:
    """מחזיר צבע לפי סטטוס משימה"""
    colors = {
        'לביצוע': '#bfc9f2',
        'הועבר לסטודיו': '#2b585e',
        'הועבר לדיגיטל': '#043841',
        'נשלח ללקוח': '#b8e994',
        'הושלם': '#14a675',
    }
    return colors.get(status, '#0073ea')

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    """דשבורד מנהל"""
    user_role = get_user_role(current_user.id)
    if not check_permission('/admin/dashboard', user_role):
        return "גישה חסומה - אין לך הרשאה לגשת לדף זה", 403
    
    users = load_users()
    sidebar_users = {uid: {'name': info.get('name', '')} for uid, info in users.items() if uid != 'admin'}
    
    # Redirect to React admin dashboard
    return redirect('/app/admin')

# ========== Time Tracking API Endpoints ==========

STALE_SESSION_HOURS = 2  # מדידות שלא נעצרו (דפדפן נסגר וכו') – מנוקות אחרי 2 שעות

def _parse_start_time(iso_str):
    """מפרש start_time מ-ISO. אם יש Z או +00:00 – UTC; אחרת naive (תאימות לאחור)."""
    if not iso_str:
        return None
    s = (iso_str or '').strip()
    if 'Z' in s or '+00:00' in s:
        s = s.replace('Z', '+00:00')
        return datetime.fromisoformat(s)
    return datetime.fromisoformat(s)

def _now_for_start(start_dt):
    """מחזיר 'עכשיו' באותו הקשר של start_dt (UTC אם start מוגדר כ-UTC, אחרת local)."""
    if start_dt and start_dt.tzinfo is not None:
        return datetime.now(timezone.utc)
    return datetime.now()

def _drop_stale_active_sessions(time_data):
    """מסיר מדידות פעילות ישנות (למשל לא נעצרו – דפדפן נסגר). מונע 'מדידה של שעתיים' תמידית."""
    changed = False
    sessions = time_data.get('active_sessions', {})
    for user_id, sess in list(sessions.items()):
        try:
            start = _parse_start_time(sess.get('start_time', '') or '')
            if start is None:
                continue
            now = _now_for_start(start)
            if (now - start).total_seconds() >= STALE_SESSION_HOURS * 3600:
                del sessions[user_id]
                changed = True
        except Exception:
            pass
    if changed:
        save_time_tracking(time_data)


def _enrich_time_tracking_session(session):
    """מוסיף ל-session שמות לקוח/פרויקט/משימה ו-elapsed_seconds (לפי start_time)."""
    if not session:
        return
    clients = load_data()
    clients_dict = {c['id']: c for c in clients}
    client = clients_dict.get(session.get('client_id'), {})
    session['client_name'] = client.get('name', 'לא ידוע')
    project = None
    task = None
    for p in client.get('projects', []):
        if p.get('id') == session.get('project_id'):
            project = p
            for t in p.get('tasks', []):
                if t.get('id') == session.get('task_id'):
                    task = t
                    break
            break
    session['project_title'] = project.get('title', 'לא ידוע') if project else 'לא ידוע'
    session['task_title'] = task.get('title', task.get('desc', 'לא ידוע')) if task else 'לא ידוע'
    if session.get('start_time'):
        start = _parse_start_time(session['start_time'])
        if start is not None:
            now = _now_for_start(start)
            session['elapsed_seconds'] = int((now - start).total_seconds())


@app.route('/api/time_tracking/start', methods=['POST'])
@login_required
@csrf.exempt
def api_time_tracking_start():
    """התחלת מדידת זמן עבור משימה"""
    try:
        data = request.get_json() if request.is_json else request.form.to_dict()
        client_id = data.get('client_id')
        project_id = data.get('project_id')
        task_id = data.get('task_id')
        user_id = current_user.id
        
        if not all([client_id, project_id, task_id]):
            return jsonify({'success': False, 'error': 'חסרים פרמטרים נדרשים'}), 400
        
        time_data = load_time_tracking()
        _drop_stale_active_sessions(time_data)
        
        # בדיקה אם יש מדידה פעילה למשתמש זה
        if user_id in time_data.get('active_sessions', {}):
            active_session = time_data['active_sessions'][user_id].copy()
            _enrich_time_tracking_session(active_session)
            return jsonify({
                'success': False,
                'error': 'יש מדידה פעילה אחרת',
                'active_session': active_session
            }), 400
        
        # יצירת מדידה חדשה – שמירת start_time ב-UTC עם Z כדי שהדפדפן יפרש נכון (מניעת +2h בישראל)
        session_id = str(uuid.uuid4())
        start_time = datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z')
        
        session = {
            'id': session_id,
            'user_id': user_id,
            'client_id': client_id,
            'project_id': project_id,
            'task_id': task_id,
            'start_time': start_time,
        }
        
        time_data.setdefault('active_sessions', {})[user_id] = session
        save_time_tracking(time_data)
        
        return jsonify({
            'success': True,
            'session': session
        })
    except Exception as e:
        print(f"Error in api_time_tracking_start: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/time_tracking/stop', methods=['POST'])
@login_required
@csrf.exempt
def api_time_tracking_stop():
    """עצירת מדידת זמן"""
    try:
        data = request.get_json() if request.is_json else request.form.to_dict()
        user_id = current_user.id
        note = data.get('note', '')
        
        time_data = load_time_tracking()
        
        if user_id not in time_data.get('active_sessions', {}):
            return jsonify({'success': False, 'error': 'אין מדידה פעילה'}), 400
        
        session = time_data['active_sessions'][user_id]
        start_dt = _parse_start_time(session.get('start_time', '') or '')
        if start_dt is None:
            return jsonify({'success': False, 'error': 'שגיאה בתאריך התחלה'}), 400
        end_time = _now_for_start(start_dt)
        duration_seconds = (end_time - start_dt).total_seconds()
        duration_hours = round(duration_seconds / 3600, 2)
        end_time_iso = end_time.isoformat().replace('+00:00', 'Z') if end_time.tzinfo else end_time.isoformat()
        
        # יצירת רשומה במדידות
        entry = {
            'id': session['id'],
            'user_id': user_id,
            'client_id': session['client_id'],
            'project_id': session['project_id'],
            'task_id': session['task_id'],
            'start_time': session['start_time'],
            'end_time': end_time_iso,
            'duration_hours': duration_hours,
            'note': note,
            'date': start_dt.date().isoformat(),
        }
        
        time_data.setdefault('entries', []).append(entry)
        del time_data['active_sessions'][user_id]
        save_time_tracking(time_data)
        
        return jsonify({
            'success': True,
            'entry': entry
        })
    except Exception as e:
        print(f"Error in api_time_tracking_stop: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/time_tracking/cancel', methods=['POST'])
@login_required
@csrf.exempt
def api_time_tracking_cancel():
    """ביטול מדידת זמן פעילה (ללא שמירה בהיסטוריה)"""
    try:
        user_id = current_user.id
        time_data = load_time_tracking()
        
        if user_id not in time_data.get('active_sessions', {}):
            return jsonify({'success': False, 'error': 'אין מדידה פעילה'}), 400
        
        # מחיקת המדידה הפעילה ללא שמירה
        del time_data['active_sessions'][user_id]
        save_time_tracking(time_data)
        
        return jsonify({
            'success': True,
            'message': 'המדידה בוטלה'
        })
    except Exception as e:
        print(f"Error in api_time_tracking_cancel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/time_tracking/active', methods=['GET'])
@limiter.exempt  # פטור מ-rate limiting כי זה נקרא באופן אוטומטי כל 10 שניות
@login_required
@csrf.exempt
def api_time_tracking_active():
    """קבלת מדידה פעילה של המשתמש הנוכחי"""
    try:
        user_id = current_user.id
        time_data = load_time_tracking()
        _drop_stale_active_sessions(time_data)
        
        active_session = time_data.get('active_sessions', {}).get(user_id)
        if active_session:
            active_session = dict(active_session)
            _enrich_time_tracking_session(active_session)
        
        return jsonify({
            'success': True,
            'active_session': active_session
        })
    except Exception as e:
        print(f"Error in api_time_tracking_active: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/time_tracking/entries', methods=['GET'])
@login_required
@csrf.exempt
def api_time_tracking_entries():
    """קבלת כל מדידות הזמן"""
    try:
        user_id = request.args.get('user_id')  # אופציונלי - אם לא מוגדר, מחזיר את כל המדידות
        client_id = request.args.get('client_id')  # אופציונלי
        task_id = request.args.get('task_id')  # אופציונלי
        month = request.args.get('month')  # בפורמט YYYY-MM
        
        time_data = load_time_tracking()
        entries = time_data.get('entries', [])
        
        # סינון לפי משתמש
        if user_id:
            entries = [e for e in entries if e.get('user_id') == user_id]
        
        # סינון לפי לקוח
        if client_id:
            entries = [e for e in entries if e.get('client_id') == client_id]
        
        # סינון לפי משימה
        if task_id:
            entries = [e for e in entries if e.get('task_id') == task_id]
        
        # סינון לפי חודש
        if month:
            entries = [e for e in entries if e.get('date', '').startswith(month)]
        
        # מיון לפי תאריך (החדש ביותר ראשון)
        entries.sort(key=lambda x: x.get('start_time', ''), reverse=True)
        
        # הוספת שמות לקוחות, פרויקטים ומשימות
        clients = load_data()
        users = load_users()
        clients_dict = {c['id']: c for c in clients}
        
        for entry in entries:
            client = clients_dict.get(entry['client_id'], {})
            entry['client_name'] = client.get('name', 'לא ידוע')
            
            # מציאת פרויקט ומשימה
            project = None
            task = None
            for p in client.get('projects', []):
                if p.get('id') == entry['project_id']:
                    project = p
                    for t in p.get('tasks', []):
                        if t.get('id') == entry['task_id']:
                            task = t
                            break
                    break
            
            entry['project_title'] = project.get('title', 'לא ידוע') if project else 'לא ידוע'
            entry['task_title'] = task.get('title', task.get('desc', 'לא ידוע')) if task else 'לא ידוע'
            entry['user_name'] = users.get(entry['user_id'], {}).get('name', 'לא ידוע')
        
        return jsonify({
            'success': True,
            'entries': entries
        })
    except Exception as e:
        print(f"Error in api_time_tracking_entries: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/time_tracking/report', methods=['GET'])
@login_required
@csrf.exempt
def api_time_tracking_report():
    """דוח חודשי של מדידות זמן"""
    try:
        month = request.args.get('month')  # בפורמט YYYY-MM
        user_id = request.args.get('user_id')  # אופציונלי
        client_id = request.args.get('client_id')  # אופציונלי
        
        if not month:
            # אם לא הוגדר חודש, משתמש בחודש הנוכחי
            month = datetime.now().strftime('%Y-%m')
        
        time_data = load_time_tracking()
        entries = time_data.get('entries', [])
        
        # סינון לפי חודש
        entries = [e for e in entries if e.get('date', '').startswith(month)]
        
        # סינון לפי משתמש
        if user_id:
            entries = [e for e in entries if e.get('user_id') == user_id]
        
        # סינון לפי לקוח
        if client_id:
            entries = [e for e in entries if e.get('client_id') == client_id]
        
        # חישוב סיכומים
        total_hours = sum(e.get('duration_hours', 0) for e in entries)
        
        # סיכום לפי לקוח
        by_client = {}
        for entry in entries:
            cid = entry['client_id']
            if cid not in by_client:
                by_client[cid] = {'hours': 0, 'entries': []}
            by_client[cid]['hours'] += entry.get('duration_hours', 0)
            by_client[cid]['entries'].append(entry)
        
        # סיכום לפי משתמש
        by_user = {}
        for entry in entries:
            uid = entry['user_id']
            if uid not in by_user:
                by_user[uid] = {'hours': 0, 'entries': []}
            by_user[uid]['hours'] += entry.get('duration_hours', 0)
            by_user[uid]['entries'].append(entry)
        
        # הוספת שמות
        clients = load_data()
        users = load_users()
        clients_dict = {c['id']: c for c in clients}
        
        for cid, data in by_client.items():
            client = clients_dict.get(cid, {})
            data['client_name'] = client.get('name', 'לא ידוע')
        
        for uid, data in by_user.items():
            data['user_name'] = users.get(uid, {}).get('name', 'לא ידוע')
        
        # הוספת שמות לכל הרשומות
        for entry in entries:
            client = clients_dict.get(entry['client_id'], {})
            entry['client_name'] = client.get('name', 'לא ידוע')
            
            # מציאת פרויקט ומשימה
            project = None
            task = None
            for p in client.get('projects', []):
                if p.get('id') == entry['project_id']:
                    project = p
                    for t in p.get('tasks', []):
                        if t.get('id') == entry['task_id']:
                            task = t
                            break
                    break
            
            entry['project_title'] = project.get('title', 'לא ידוע') if project else 'לא ידוע'
            entry['task_title'] = task.get('title', task.get('desc', 'לא ידוע')) if task else 'לא ידוע'
            entry['user_name'] = users.get(entry['user_id'], {}).get('name', 'לא ידוע')
        
        return jsonify({
            'success': True,
            'month': month,
            'total_hours': total_hours,
            'total_entries': len(entries),
            'by_client': by_client,
            'by_user': by_user,
            'entries': entries
        })
    except Exception as e:
        print(f"Error in api_time_tracking_report: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/time_tracking/entry/<entry_id>', methods=['PUT'])
@login_required
@csrf.exempt
def api_time_tracking_update(entry_id):
    """עדכון רשומת מדידת זמן קיימת"""
    try:
        data = request.get_json() if request.is_json else request.form.to_dict()
        
        time_data = load_time_tracking()
        entries = time_data.get('entries', [])
        
        # מציאת הרשומה
        entry_index = None
        for i, entry in enumerate(entries):
            if entry.get('id') == entry_id:
                entry_index = i
                break
        
        if entry_index is None:
            return jsonify({'success': False, 'error': 'רשומה לא נמצאה'}), 404
        
        entry = entries[entry_index]
        
        # בדיקת הרשאות - רק מנהלים או הבעלים יכולים לערוך
        if current_user.role not in ['admin', 'manager'] and entry.get('user_id') != current_user.id:
            return jsonify({'success': False, 'error': 'אין הרשאה לערוך רשומה זו'}), 403
        
        # עדכון השדות
        if 'start_time' in data:
            entry['start_time'] = data['start_time']
        if 'end_time' in data:
            entry['end_time'] = data['end_time']
        if 'note' in data:
            entry['note'] = data['note']
        if 'duration_hours' in data:
            entry['duration_hours'] = float(data['duration_hours'])
        
        # חישוב מחדש של duration_hours אם שונו הזמנים
        if 'start_time' in data or 'end_time' in data:
            start_dt = _parse_start_time(entry['start_time'])
            end_dt = _parse_start_time(entry['end_time'])
            if start_dt and end_dt:
                duration_seconds = (end_dt - start_dt).total_seconds()
                entry['duration_hours'] = round(duration_seconds / 3600, 2)
                entry['date'] = start_dt.date().isoformat()
        
        entries[entry_index] = entry
        time_data['entries'] = entries
        save_time_tracking(time_data)
        
        return jsonify({
            'success': True,
            'entry': entry,
            'message': 'הרשומה עודכנה בהצלחה'
        })
    except Exception as e:
        print(f"Error in api_time_tracking_update: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/time_tracking/entry/<entry_id>', methods=['DELETE'])
@login_required
@csrf.exempt
def api_time_tracking_delete(entry_id):
    """מחיקת רשומת מדידת זמן"""
    try:
        time_data = load_time_tracking()
        entries = time_data.get('entries', [])
        
        # מציאת הרשומה
        entry_index = None
        entry_to_delete = None
        for i, entry in enumerate(entries):
            if entry.get('id') == entry_id:
                entry_index = i
                entry_to_delete = entry
                break
        
        if entry_index is None:
            return jsonify({'success': False, 'error': 'רשומה לא נמצאה'}), 404
        
        # בדיקת הרשאות - רק מנהלים או הבעלים יכולים למחוק
        if current_user.role not in ['admin', 'manager'] and entry_to_delete.get('user_id') != current_user.id:
            return jsonify({'success': False, 'error': 'אין הרשאה למחוק רשומה זו'}), 403
        
        # מחיקת הרשומה
        entries.pop(entry_index)
        time_data['entries'] = entries
        save_time_tracking(time_data)
        
        return jsonify({
            'success': True,
            'message': 'הרשומה נמחקה בהצלחה'
        })
    except Exception as e:
        print(f"Error in api_time_tracking_delete: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/time_tracking/manual', methods=['POST'])
@login_required
@csrf.exempt
def api_time_tracking_manual():
    """הוספת רשומת מדידת זמן ידנית"""
    try:
        data = request.get_json() if request.is_json else request.form.to_dict()
        
        # בדיקת שדות חובה
        required_fields = ['client_id', 'project_id', 'task_id', 'date', 'duration_hours']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'error': f'שדה חובה חסר: {field}'}), 400
        
        # אם יש user_id בנתונים ומשתמש הוא מנהל, השתמש בו. אחרת השתמש במשתמש הנוכחי
        user_id = current_user.id
        if data.get('user_id') and current_user.role in ['admin', 'manager']:
            user_id = data['user_id']
        
        duration_hours = float(data['duration_hours'])
        date_str = data['date']  # בפורמט YYYY-MM-DD
        
        # יצירת זמני התחלה וסיום
        if data.get('start_time') and data.get('end_time'):
            start_time = data['start_time']
            end_time = data['end_time']
        else:
            # אם לא סופקו זמנים, יצירת זמנים לפי התאריך ומשך הזמן
            start_dt = datetime.strptime(date_str, '%Y-%m-%d').replace(hour=9, minute=0, second=0)
            end_dt = start_dt + timedelta(hours=duration_hours)
            start_time = start_dt.isoformat() + 'Z'
            end_time = end_dt.isoformat() + 'Z'
        
        # יצירת הרשומה
        entry = {
            'id': str(uuid.uuid4()),
            'user_id': user_id,
            'client_id': data['client_id'],
            'project_id': data['project_id'],
            'task_id': data['task_id'],
            'start_time': start_time,
            'end_time': end_time,
            'duration_hours': duration_hours,
            'note': data.get('note', 'הוספה ידנית'),
            'date': date_str,
            'manual_entry': True  # סימון שזו רשומה ידנית
        }
        
        time_data = load_time_tracking()
        time_data.setdefault('entries', []).append(entry)
        save_time_tracking(time_data)
        
        return jsonify({
            'success': True,
            'entry': entry,
            'message': 'הרשומה נוספה בהצלחה'
        })
    except Exception as e:
        print(f"Error in api_time_tracking_manual: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/time_tracking/adjust/<entry_id>', methods=['POST'])
@login_required
@csrf.exempt
def api_time_tracking_adjust(entry_id):
    """הוספה או הורדה של שעות מרשומה קיימת"""
    try:
        data = request.get_json() if request.is_json else request.form.to_dict()
        
        adjustment_hours = data.get('adjustment_hours')
        if adjustment_hours is None:
            return jsonify({'success': False, 'error': 'חסר פרמטר adjustment_hours'}), 400
        
        adjustment_hours = float(adjustment_hours)
        
        time_data = load_time_tracking()
        entries = time_data.get('entries', [])
        
        # מציאת הרשומה
        entry_index = None
        for i, entry in enumerate(entries):
            if entry.get('id') == entry_id:
                entry_index = i
                break
        
        if entry_index is None:
            return jsonify({'success': False, 'error': 'רשומה לא נמצאה'}), 404
        
        entry = entries[entry_index]
        
        # בדיקת הרשאות
        if current_user.role not in ['admin', 'manager'] and entry.get('user_id') != current_user.id:
            return jsonify({'success': False, 'error': 'אין הרשאה לערוך רשומה זו'}), 403
        
        # עדכון השעות
        old_hours = entry.get('duration_hours', 0)
        new_hours = max(0, old_hours + adjustment_hours)  # לא מאפשר שעות שליליות
        entry['duration_hours'] = round(new_hours, 2)
        
        # עדכון זמן הסיום בהתאם
        start_dt = _parse_start_time(entry['start_time'])
        if start_dt:
            end_dt = start_dt + timedelta(hours=new_hours)
            entry['end_time'] = end_dt.isoformat().replace('+00:00', 'Z') if end_dt.tzinfo else end_dt.isoformat() + 'Z'
        
        # הוספת הערה על ההתאמה
        adjustment_note = f"התאמה ידנית: {'+' if adjustment_hours > 0 else ''}{adjustment_hours} שעות"
        if entry.get('note'):
            entry['note'] = f"{entry['note']} | {adjustment_note}"
        else:
            entry['note'] = adjustment_note
        
        entries[entry_index] = entry
        time_data['entries'] = entries
        save_time_tracking(time_data)
        
        return jsonify({
            'success': True,
            'entry': entry,
            'message': f'השעות עודכנו: {old_hours} → {new_hours}'
        })
    except Exception as e:
        print(f"Error in api_time_tracking_adjust: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/logout', methods=['GET', 'POST'])
@csrf.exempt
def logout():
    logout_user()
    wants_json = request.headers.get('Accept', '').find('application/json') != -1 or \
                request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if wants_json:
        return jsonify({'success': True})
    return redirect(url_for('login'))

# ============================================
# React Frontend Serving (Production)
# ============================================
REACT_BUILD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'dist')

@app.route('/assets/<path:filename>')
def serve_react_assets(filename):
    """Serve React static assets (JS, CSS, etc.)"""
    return send_from_directory(os.path.join(REACT_BUILD_DIR, 'assets'), filename)

@app.route('/app/assets/<path:filename>')
def serve_react_app_assets(filename):
    """Serve React static assets under /app/assets/ path"""
    return send_from_directory(os.path.join(REACT_BUILD_DIR, 'assets'), filename)

# React SPA catch-all routes - serve index.html for client-side routing
REACT_ROUTES = ['/dashboard', '/all_clients', '/finance', '/events', '/suppliers', 
                '/quotes', '/forms', '/admin', '/archive', '/my_tasks', '/time_tracking']

@app.route('/app')
@app.route('/app/')
@app.route('/app/<path:path>')
def serve_react_app(path=''):
    """Serve React SPA for /app routes - React handles its own auth
    This catch-all ensures that page refresh on any /app/* route
    returns the React app, allowing React Router to handle the routing.
    """
    react_index = os.path.join(REACT_BUILD_DIR, 'index.html')
    if os.path.exists(react_index):
        return send_from_directory(REACT_BUILD_DIR, 'index.html')
    return "React build not found. Run 'npm run build' first.", 404

# Redirect old routes to React app
@app.route('/dashboard')
@login_required  
def dashboard_redirect():
    return redirect('/app/')

@app.route('/time_tracking')
@login_required  
def time_tracking_redirect():
    return redirect('/app/time_tracking')

if __name__ == '__main__':
    # Railway deployment configuration
    port = int(os.environ.get('PORT', 5000))
    host = os.environ.get('HOST', '0.0.0.0')
    debug = os.environ.get('FLASK_ENV') != 'production'
    app.run(host=host, port=port, debug=debug)
