"""
Events Blueprint
Contains routes for event management
"""
import os
import uuid
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file
from flask_login import login_required, current_user
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from backend.extensions import csrf
from backend.utils.helpers import (
    load_events, save_events, load_users, load_suppliers,
    load_equipment_bank, save_equipment_bank,
    load_checklist_templates
)
from backend.utils.permissions import get_user_role, is_manager_or_admin

# Create blueprint
events_bp = Blueprint('events', __name__)

# Directory paths
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
STATIC_FOLDER = os.path.join(BASE_DIR, 'static')


@events_bp.route('/api/events')
@login_required
def api_events():
    """Get all events"""
    try:
        events = load_events()
        users = load_users()
        suppliers = load_suppliers()
        equipment_bank = load_equipment_bank()
        templates = load_checklist_templates()
        user_role = get_user_role(current_user.id)
        
        # Filter events based on permissions
        if not is_manager_or_admin(current_user.id, user_role):
            events = [e for e in events 
                     if current_user.id in e.get('assigned_users', [])]
        
        # Sort by date
        events.sort(key=lambda x: x.get('date', ''), reverse=True)
        
        return jsonify({
            'success': True,
            'events': events,
            'users': {uid: {'name': u.get('name', uid)} for uid, u in users.items()},
            'suppliers': suppliers,
            'equipment_bank': equipment_bank,
            'checklist_templates': templates
        })
    except Exception as e:
        print(f"Error in api_events: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@events_bp.route('/add_event', methods=['POST'])
@login_required
@csrf.exempt
def add_event():
    """Add new event"""
    try:
        events = load_events()
        
        name = request.form.get('name', '').strip()
        if not name:
            return jsonify({'success': False, 'error': 'שם אירוע חסר'}), 400
        
        new_event = {
            'id': str(uuid.uuid4()),
            'name': name,
            'date': request.form.get('date', ''),
            'time': request.form.get('time', ''),
            'location': request.form.get('location', ''),
            'client': request.form.get('client', ''),
            'description': request.form.get('description', ''),
            'event_type': request.form.get('event_type', ''),
            'guests_count': int(request.form.get('guests_count', 0) or 0),
            'budget': float(request.form.get('budget', 0) or 0),
            'assigned_users': request.form.getlist('assigned_users'),
            'checklist': [],
            'suppliers': [],
            'equipment': [],
            'management_table': [],
            'shopping_list': [],
            'charges': [],
            'active': True,
            'created_at': datetime.now().isoformat(),
            'created_by': current_user.id
        }
        
        # Add checklist from template
        event_type = request.form.get('event_type', '')
        if event_type:
            templates = load_checklist_templates()
            if event_type in templates:
                new_event['checklist'] = [
                    {'text': item, 'done': False}
                    for item in templates[event_type]
                ]
        
        events.append(new_event)
        save_events(events)
        
        return jsonify({'success': True, 'event': new_event})
    except Exception as e:
        print(f"Error in add_event: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@events_bp.route('/update_event/<event_id>', methods=['POST'])
@login_required
@csrf.exempt
def update_event(event_id):
    """Update event details"""
    try:
        events = load_events()
        event = next((e for e in events if e['id'] == event_id), None)
        
        if not event:
            return jsonify({'success': False, 'error': 'Event not found'}), 404
        
        # Update fields
        for field in ['name', 'date', 'time', 'location', 'client', 'description', 'event_type']:
            if field in request.form:
                event[field] = request.form.get(field)
        
        if 'guests_count' in request.form:
            event['guests_count'] = int(request.form.get('guests_count', 0) or 0)
        
        if 'budget' in request.form:
            event['budget'] = float(request.form.get('budget', 0) or 0)
        
        if 'assigned_users' in request.form:
            event['assigned_users'] = request.form.getlist('assigned_users')
        
        event['updated_at'] = datetime.now().isoformat()
        save_events(events)
        
        return jsonify({'success': True, 'event': event})
    except Exception as e:
        print(f"Error in update_event: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@events_bp.route('/toggle_event_active/<event_id>', methods=['POST'])
@login_required
@csrf.exempt
def toggle_event_active(event_id):
    """Toggle event active status"""
    try:
        events = load_events()
        event = next((e for e in events if e['id'] == event_id), None)
        
        if not event:
            return jsonify({'success': False, 'error': 'Event not found'}), 404
        
        event['active'] = not event.get('active', True)
        save_events(events)
        
        return jsonify({'success': True, 'active': event['active']})
    except Exception as e:
        print(f"Error in toggle_event_active: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@events_bp.route('/update_event_checklist/<event_id>', methods=['POST'])
@login_required
@csrf.exempt
def update_event_checklist(event_id):
    """Update event checklist"""
    try:
        events = load_events()
        event = next((e for e in events if e['id'] == event_id), None)
        
        if not event:
            return jsonify({'success': False, 'error': 'Event not found'}), 404
        
        checklist = request.json.get('checklist', [])
        event['checklist'] = checklist
        save_events(events)
        
        return jsonify({'success': True})
    except Exception as e:
        print(f"Error in update_event_checklist: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@events_bp.route('/add_checklist_item/<event_id>', methods=['POST'])
@login_required
@csrf.exempt
def add_checklist_item(event_id):
    """Add item to event checklist"""
    try:
        events = load_events()
        event = next((e for e in events if e['id'] == event_id), None)
        
        if not event:
            return jsonify({'success': False, 'error': 'Event not found'}), 404
        
        text = request.form.get('text', '').strip()
        if not text:
            return jsonify({'success': False, 'error': 'טקסט חסר'}), 400
        
        if 'checklist' not in event:
            event['checklist'] = []
        
        event['checklist'].append({'text': text, 'done': False})
        save_events(events)
        
        return jsonify({'success': True, 'checklist': event['checklist']})
    except Exception as e:
        print(f"Error in add_checklist_item: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@events_bp.route('/remove_checklist_item/<event_id>/<item_index>', methods=['POST'])
@login_required
@csrf.exempt
def remove_checklist_item(event_id, item_index):
    """Remove item from event checklist"""
    try:
        events = load_events()
        event = next((e for e in events if e['id'] == event_id), None)
        
        if not event:
            return jsonify({'success': False, 'error': 'Event not found'}), 404
        
        idx = int(item_index)
        if 0 <= idx < len(event.get('checklist', [])):
            event['checklist'].pop(idx)
            save_events(events)
        
        return jsonify({'success': True})
    except Exception as e:
        print(f"Error in remove_checklist_item: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@events_bp.route('/add_event_supplier/<event_id>', methods=['POST'])
@login_required
@csrf.exempt
def add_event_supplier(event_id):
    """Add supplier to event"""
    try:
        events = load_events()
        event = next((e for e in events if e['id'] == event_id), None)
        
        if not event:
            return jsonify({'success': False, 'error': 'Event not found'}), 404
        
        if 'suppliers' not in event:
            event['suppliers'] = []
        
        supplier_entry = {
            'supplier_id': request.form.get('supplier_id', ''),
            'supplier_name': request.form.get('supplier_name', ''),
            'service': request.form.get('service', ''),
            'cost': float(request.form.get('cost', 0) or 0),
            'notes': request.form.get('notes', ''),
            'confirmed': request.form.get('confirmed', '').lower() == 'true'
        }
        
        event['suppliers'].append(supplier_entry)
        save_events(events)
        
        return jsonify({'success': True, 'suppliers': event['suppliers']})
    except Exception as e:
        print(f"Error in add_event_supplier: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@events_bp.route('/update_event_equipment/<event_id>', methods=['POST'])
@login_required
@csrf.exempt
def update_event_equipment(event_id):
    """Update event equipment"""
    try:
        events = load_events()
        event = next((e for e in events if e['id'] == event_id), None)
        
        if not event:
            return jsonify({'success': False, 'error': 'Event not found'}), 404
        
        equipment = request.json.get('equipment', [])
        event['equipment'] = equipment
        save_events(events)
        
        return jsonify({'success': True})
    except Exception as e:
        print(f"Error in update_event_equipment: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@events_bp.route('/add_equipment_to_bank', methods=['POST'])
@login_required
@csrf.exempt
def add_equipment_to_bank():
    """Add new equipment to bank"""
    try:
        equipment_bank = load_equipment_bank()
        
        name = request.form.get('name', '').strip()
        if not name:
            return jsonify({'success': False, 'error': 'שם ציוד חסר'}), 400
        
        if name not in equipment_bank:
            equipment_bank.append(name)
            save_equipment_bank(equipment_bank)
        
        return jsonify({'success': True, 'equipment_bank': equipment_bank})
    except Exception as e:
        print(f"Error in add_equipment_to_bank: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@events_bp.route('/add_event_charge/<event_id>', methods=['POST'])
@login_required
@csrf.exempt
def add_event_charge(event_id):
    """Add charge to event"""
    try:
        events = load_events()
        event = next((e for e in events if e['id'] == event_id), None)
        
        if not event:
            return jsonify({'success': False, 'error': 'Event not found'}), 404
        
        if 'charges' not in event:
            event['charges'] = []
        
        new_charge = {
            'id': str(uuid.uuid4()),
            'description': request.form.get('description', ''),
            'amount': float(request.form.get('amount', 0) or 0),
            'our_cost': float(request.form.get('our_cost', 0) or 0),
            'date': request.form.get('date', datetime.now().strftime('%Y-%m-%d')),
            'paid': request.form.get('paid', '').lower() == 'true',
            'created_at': datetime.now().isoformat()
        }
        
        event['charges'].append(new_charge)
        save_events(events)
        
        return jsonify({'success': True, 'charge': new_charge})
    except Exception as e:
        print(f"Error in add_event_charge: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@events_bp.route('/export_event_equipment/<event_id>')
@login_required
def export_event_equipment(event_id):
    """Export event equipment to Excel"""
    try:
        events = load_events()
        event = next((e for e in events if e['id'] == event_id), None)
        
        if not event:
            return jsonify({'success': False, 'error': 'Event not found'}), 404
        
        wb = Workbook()
        ws = wb.active
        ws.title = "ציוד"
        ws.sheet_view.rightToLeft = True
        
        # Title
        ws['A1'] = f"ציוד לאירוע: {event.get('name', '')}"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Headers
        headers = ['פריט', 'כמות', 'הערות']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Data
        row = 4
        for item in event.get('equipment', []):
            if isinstance(item, dict):
                ws.cell(row=row, column=1, value=item.get('name', ''))
                ws.cell(row=row, column=2, value=item.get('quantity', 1))
                ws.cell(row=row, column=3, value=item.get('notes', ''))
            else:
                ws.cell(row=row, column=1, value=str(item))
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 30
        
        filename = f"equipment_{event_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(STATIC_FOLDER, filename)
        wb.save(filepath)
        
        return send_file(filepath, as_attachment=True, download_name=filename)
    except Exception as e:
        print(f"Error in export_event_equipment: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
