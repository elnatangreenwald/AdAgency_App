"""
Time Tracking Blueprint
Contains routes for time tracking functionality
"""
import uuid
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file
from flask_login import login_required, current_user
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import os

from backend.extensions import csrf, limiter
from backend.utils.helpers import load_time_tracking, save_time_tracking, load_data, load_users
from backend.utils.permissions import get_user_role, is_manager_or_admin

# Create blueprint
time_tracking_bp = Blueprint('time_tracking', __name__)

# Directory paths
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
STATIC_FOLDER = os.path.join(BASE_DIR, 'static')


@time_tracking_bp.route('/api/time_tracking/start', methods=['POST'])
@login_required
@csrf.exempt
def api_time_tracking_start():
    """Start time tracking for a task"""
    try:
        user_id = current_user.id
        time_data = load_time_tracking()
        
        # Get task info from request
        client_id = request.form.get('client_id', '')
        client_name = request.form.get('client_name', '')
        project_id = request.form.get('project_id', '')
        project_name = request.form.get('project_name', '')
        task_id = request.form.get('task_id', '')
        task_title = request.form.get('task_title', '')
        
        # Check if user already has active session
        if user_id in time_data.get('active_sessions', {}):
            return jsonify({
                'success': False,
                'error': 'כבר קיימת מדידה פעילה. עצור אותה לפני התחלה חדשה.'
            }), 400
        
        # Create active session
        if 'active_sessions' not in time_data:
            time_data['active_sessions'] = {}
        
        time_data['active_sessions'][user_id] = {
            'client_id': client_id,
            'client_name': client_name,
            'project_id': project_id,
            'project_name': project_name,
            'task_id': task_id,
            'task_title': task_title,
            'start_time': datetime.now().isoformat()
        }
        
        save_time_tracking(time_data)
        
        return jsonify({
            'success': True,
            'session': time_data['active_sessions'][user_id]
        })
    except Exception as e:
        print(f"Error in api_time_tracking_start: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@time_tracking_bp.route('/api/time_tracking/stop', methods=['POST'])
@login_required
@csrf.exempt
def api_time_tracking_stop():
    """Stop time tracking and save entry"""
    try:
        user_id = current_user.id
        time_data = load_time_tracking()
        
        if user_id not in time_data.get('active_sessions', {}):
            return jsonify({'success': False, 'error': 'אין מדידה פעילה לעצירה'}), 400
        
        session = time_data['active_sessions'][user_id]
        start_time = datetime.fromisoformat(session['start_time'])
        end_time = datetime.now()
        duration_seconds = int((end_time - start_time).total_seconds())
        
        # Add notes from request
        notes = request.form.get('notes', '')
        
        # Create entry
        if 'entries' not in time_data:
            time_data['entries'] = []
        
        entry = {
            'id': str(uuid.uuid4()),
            'user_id': user_id,
            'client_id': session.get('client_id', ''),
            'client_name': session.get('client_name', ''),
            'project_id': session.get('project_id', ''),
            'project_name': session.get('project_name', ''),
            'task_id': session.get('task_id', ''),
            'task_title': session.get('task_title', ''),
            'start_time': session['start_time'],
            'end_time': end_time.isoformat(),
            'duration_seconds': duration_seconds,
            'notes': notes,
            'created_at': datetime.now().isoformat()
        }
        
        time_data['entries'].append(entry)
        
        # Remove active session
        del time_data['active_sessions'][user_id]
        
        save_time_tracking(time_data)
        
        return jsonify({'success': True, 'entry': entry})
    except Exception as e:
        print(f"Error in api_time_tracking_stop: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@time_tracking_bp.route('/api/time_tracking/cancel', methods=['POST'])
@login_required
@csrf.exempt
def api_time_tracking_cancel():
    """Cancel active time tracking without saving"""
    try:
        user_id = current_user.id
        time_data = load_time_tracking()
        
        if user_id not in time_data.get('active_sessions', {}):
            return jsonify({'success': False, 'error': 'אין מדידה פעילה לביטול'}), 400
        
        del time_data['active_sessions'][user_id]
        save_time_tracking(time_data)
        
        return jsonify({'success': True, 'message': 'המדידה בוטלה בהצלחה'})
    except Exception as e:
        print(f"Error in api_time_tracking_cancel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@time_tracking_bp.route('/api/time_tracking/active')
@login_required
@limiter.exempt  # פטור מ-rate limiting כי זה נקרא הרבה פעמים (auto-refresh)
def api_time_tracking_active():
    """Get active time tracking session for current user"""
    try:
        user_id = current_user.id
        time_data = load_time_tracking()
        
        session = time_data.get('active_sessions', {}).get(user_id)
        
        return jsonify({
            'success': True,
            'active_session': session
        })
    except Exception as e:
        print(f"Error in api_time_tracking_active: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@time_tracking_bp.route('/api/time_tracking/entries')
@login_required
def api_time_tracking_entries():
    """Get time tracking entries"""
    try:
        user_id = current_user.id
        user_role = get_user_role(user_id)
        time_data = load_time_tracking()
        users = load_users()
        
        entries = time_data.get('entries', [])
        
        # Filter based on permissions
        if not is_manager_or_admin(user_id, user_role):
            entries = [e for e in entries if e.get('user_id') == user_id]
        
        # Sort by date (newest first)
        entries.sort(key=lambda x: x.get('start_time', ''), reverse=True)
        
        # Add user names
        for entry in entries:
            entry['user_name'] = users.get(entry.get('user_id', ''), {}).get('name', '')
        
        return jsonify({
            'success': True,
            'entries': entries
        })
    except Exception as e:
        print(f"Error in api_time_tracking_entries: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@time_tracking_bp.route('/api/time_tracking/report')
@login_required
def api_time_tracking_report():
    """Get time tracking report"""
    try:
        user_id = current_user.id
        user_role = get_user_role(user_id)
        time_data = load_time_tracking()
        users = load_users()
        clients = load_data()
        
        # Get filter params
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        filter_user = request.args.get('user_id', '')
        filter_client = request.args.get('client_id', '')
        
        entries = time_data.get('entries', [])
        
        # Filter based on permissions
        if not is_manager_or_admin(user_id, user_role):
            entries = [e for e in entries if e.get('user_id') == user_id]
        
        # Apply filters
        if filter_user:
            entries = [e for e in entries if e.get('user_id') == filter_user]
        
        if filter_client:
            entries = [e for e in entries if e.get('client_id') == filter_client]
        
        if start_date:
            entries = [e for e in entries if e.get('start_time', '') >= start_date]
        
        if end_date:
            entries = [e for e in entries if e.get('start_time', '')[:10] <= end_date]
        
        # Calculate totals
        total_seconds = sum(e.get('duration_seconds', 0) for e in entries)
        
        # Group by user
        by_user = {}
        for entry in entries:
            uid = entry.get('user_id', '')
            if uid not in by_user:
                by_user[uid] = {
                    'user_id': uid,
                    'user_name': users.get(uid, {}).get('name', uid),
                    'total_seconds': 0,
                    'hourly_rate': users.get(uid, {}).get('hourly_rate', 0),
                    'entries': []
                }
            by_user[uid]['total_seconds'] += entry.get('duration_seconds', 0)
            by_user[uid]['entries'].append(entry)
        
        # Group by client
        by_client = {}
        for entry in entries:
            cid = entry.get('client_id', '')
            cname = entry.get('client_name', '')
            if cid not in by_client:
                by_client[cid] = {
                    'client_id': cid,
                    'client_name': cname,
                    'total_seconds': 0
                }
            by_client[cid]['total_seconds'] += entry.get('duration_seconds', 0)
        
        return jsonify({
            'success': True,
            'entries': entries,
            'total_seconds': total_seconds,
            'by_user': list(by_user.values()),
            'by_client': list(by_client.values()),
            'users': {uid: {'name': u.get('name', uid)} for uid, u in users.items()},
            'clients': [{'id': c['id'], 'name': c.get('name', '')} for c in clients]
        })
    except Exception as e:
        print(f"Error in api_time_tracking_report: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@time_tracking_bp.route('/api/time_tracking/export')
@login_required
def api_time_tracking_export():
    """Export time tracking to Excel"""
    try:
        user_id = current_user.id
        user_role = get_user_role(user_id)
        time_data = load_time_tracking()
        users = load_users()
        
        # Get filter params
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        filter_user = request.args.get('user_id', '')
        
        entries = time_data.get('entries', [])
        
        # Filter based on permissions
        if not is_manager_or_admin(user_id, user_role):
            entries = [e for e in entries if e.get('user_id') == user_id]
        
        # Apply filters
        if filter_user:
            entries = [e for e in entries if e.get('user_id') == filter_user]
        if start_date:
            entries = [e for e in entries if e.get('start_time', '') >= start_date]
        if end_date:
            entries = [e for e in entries if e.get('start_time', '')[:10] <= end_date]
        
        # Create Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "דוח שעות"
        ws.sheet_view.rightToLeft = True
        
        # Headers
        headers = ['עובד', 'לקוח', 'פרויקט', 'משימה', 'תאריך', 'התחלה', 'סיום', 'משך (שעות)', 'הערות']
        header_font = Font(bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
        
        # Data
        row = 2
        for entry in entries:
            user_name = users.get(entry.get('user_id', ''), {}).get('name', '')
            duration_hours = round(entry.get('duration_seconds', 0) / 3600, 2)
            
            start_dt = entry.get('start_time', '')[:16].replace('T', ' ') if entry.get('start_time') else ''
            end_dt = entry.get('end_time', '')[:16].replace('T', ' ') if entry.get('end_time') else ''
            
            ws.cell(row=row, column=1, value=user_name)
            ws.cell(row=row, column=2, value=entry.get('client_name', ''))
            ws.cell(row=row, column=3, value=entry.get('project_name', ''))
            ws.cell(row=row, column=4, value=entry.get('task_title', ''))
            ws.cell(row=row, column=5, value=start_dt[:10] if start_dt else '')
            ws.cell(row=row, column=6, value=start_dt[11:] if start_dt else '')
            ws.cell(row=row, column=7, value=end_dt[11:] if end_dt else '')
            ws.cell(row=row, column=8, value=duration_hours)
            ws.cell(row=row, column=9, value=entry.get('notes', ''))
            row += 1
        
        # Total row
        total_hours = sum(e.get('duration_seconds', 0) for e in entries) / 3600
        ws.cell(row=row+1, column=7, value='סה"כ:')
        ws.cell(row=row+1, column=8, value=round(total_hours, 2))
        ws.cell(row=row+1, column=7).font = header_font
        ws.cell(row=row+1, column=8).font = header_font
        
        # Column widths
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        filename = f"time_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(STATIC_FOLDER, filename)
        wb.save(filepath)
        
        return send_file(filepath, as_attachment=True, download_name=filename)
    except Exception as e:
        print(f"Error in api_time_tracking_export: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
