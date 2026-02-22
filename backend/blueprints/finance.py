"""
Finance Blueprint
Contains routes for financial management, invoices, and charges
"""
import os
import uuid
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file
from flask_login import login_required, current_user
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from backend.extensions import csrf, limiter
from backend.utils.helpers import load_data, save_data, load_users, get_next_charge_number
from backend.utils.permissions import get_user_role, can_user_access_client, is_manager_or_admin, filter_active_clients

# Create blueprint
finance_bp = Blueprint('finance', __name__)

# Directory paths
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
STATIC_FOLDER = os.path.join(BASE_DIR, 'static')


@finance_bp.route('/api/finance')
@login_required
def api_finance():
    """Get finance data for all clients"""
    try:
        data = load_data()
        users = load_users()
        user_role = get_user_role(current_user.id)
        
        # Filter clients based on permissions
        if is_manager_or_admin(current_user.id, user_role):
            clients = filter_active_clients(data)
        else:
            clients = [c for c in filter_active_clients(data) 
                      if can_user_access_client(current_user.id, user_role, c)]
        
        # Calculate finance summary
        total_retainer = sum(float(c.get('retainer', 0) or 0) for c in clients)
        total_charges = 0
        total_paid = 0
        total_unpaid = 0
        
        for client in clients:
            for charge in client.get('extra_charges', []):
                amount = float(charge.get('amount', 0) or 0)
                total_charges += amount
                if charge.get('paid'):
                    total_paid += amount
                else:
                    total_unpaid += amount
        
        return jsonify({
            'success': True,
            'clients': clients,
            'users': {uid: {'name': u.get('name', uid)} for uid, u in users.items()},
            'summary': {
                'total_retainer': total_retainer,
                'total_charges': total_charges,
                'total_paid': total_paid,
                'total_unpaid': total_unpaid
            }
        })
    except Exception as e:
        print(f"Error in api_finance: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@finance_bp.route('/quick_add_charge', methods=['POST'])
@login_required
@csrf.exempt
def quick_add_charge():
    """Add charge to client"""
    try:
        data = load_data()
        client_id = request.form.get('client_id')
        
        client = next((c for c in data if c['id'] == client_id), None)
        if not client:
            return jsonify({'success': False, 'error': 'Client not found'}), 404
        
        if 'extra_charges' not in client:
            client['extra_charges'] = []
        
        new_charge = {
            'id': str(uuid.uuid4()),
            'charge_number': get_next_charge_number(client),
            'description': request.form.get('description', ''),
            'amount': float(request.form.get('amount', 0) or 0),
            'our_cost': float(request.form.get('our_cost', 0) or 0),
            'date': request.form.get('date', datetime.now().strftime('%Y-%m-%d')),
            'paid': request.form.get('paid', '').lower() == 'true',
            'notes': request.form.get('notes', ''),
            'created_at': datetime.now().isoformat()
        }
        
        client['extra_charges'].append(new_charge)
        save_data(data)
        
        return jsonify({'success': True, 'charge': new_charge})
    except Exception as e:
        print(f"Error in quick_add_charge: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@finance_bp.route('/toggle_charge_status/<client_id>/<charge_id>', methods=['POST'])
@login_required
@csrf.exempt
def toggle_charge_status(client_id, charge_id):
    """Toggle charge paid/completed status"""
    try:
        data = load_data()
        client = next((c for c in data if c['id'] == client_id), None)
        
        if not client:
            return jsonify({'success': False, 'error': 'Client not found'}), 404
        
        charge = next((ch for ch in client.get('extra_charges', []) if ch['id'] == charge_id), None)
        
        if not charge:
            return jsonify({'success': False, 'error': 'Charge not found'}), 404
        
        current_status = charge.get('paid', False) or charge.get('completed', False)
        new_status = not current_status
        charge['paid'] = new_status
        charge['completed'] = new_status
        save_data(data)
        
        return jsonify({'success': True, 'paid': new_status, 'completed': new_status})
    except Exception as e:
        print(f"Error in toggle_charge_status: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@finance_bp.route('/update_charge_our_cost/<client_id>/<charge_id>', methods=['POST'])
@login_required
@csrf.exempt
def update_charge_our_cost(client_id, charge_id):
    """Update charge our cost"""
    try:
        data = load_data()
        client = next((c for c in data if c['id'] == client_id), None)
        
        if not client:
            return jsonify({'success': False, 'error': 'Client not found'}), 404
        
        charge = next((ch for ch in client.get('extra_charges', []) if ch['id'] == charge_id), None)
        
        if not charge:
            return jsonify({'success': False, 'error': 'Charge not found'}), 404
        
        charge['our_cost'] = float(request.form.get('our_cost', 0) or 0)
        save_data(data)
        
        return jsonify({'success': True, 'our_cost': charge['our_cost']})
    except Exception as e:
        print(f"Error in update_charge_our_cost: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@finance_bp.route('/delete_charge/<client_id>/<charge_id>', methods=['POST'])
@login_required
@csrf.exempt
@limiter.exempt  # פטור מ-rate limiting כי זו פעולה רגילה של משתמש
def delete_charge(client_id, charge_id):
    """Delete a charge"""
    try:
        data = load_data()
        client = next((c for c in data if c['id'] == client_id), None)
        
        if not client:
            return jsonify({'success': False, 'error': 'Client not found'}), 404
        
        client['extra_charges'] = [ch for ch in client.get('extra_charges', []) if ch['id'] != charge_id]
        save_data(data)
        
        return jsonify({'success': True})
    except Exception as e:
        print(f"Error in delete_charge: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@finance_bp.route('/update_finance/<client_id>', methods=['POST'])
@login_required
@csrf.exempt
def update_finance(client_id):
    """Update client financial info"""
    try:
        data = load_data()
        client = next((c for c in data if c['id'] == client_id), None)
        
        if not client:
            return jsonify({'success': False, 'error': 'Client not found'}), 404
        
        if 'retainer' in request.form:
            client['retainer'] = float(request.form.get('retainer', 0) or 0)
        
        save_data(data)
        
        return jsonify({'success': True})
    except Exception as e:
        print(f"Error in update_finance: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@finance_bp.route('/generate_invoice/<client_id>')
@login_required
def generate_invoice(client_id):
    """Generate invoice Excel for client"""
    try:
        data = load_data()
        client = next((c for c in data if c['id'] == client_id), None)
        
        if not client:
            return jsonify({'success': False, 'error': 'Client not found'}), 404
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "חשבונית"
        ws.sheet_view.rightToLeft = True
        
        # Styles
        header_font = Font(name='Arial', bold=True, size=14)
        title_font = Font(name='Arial', bold=True, size=18)
        
        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = f"חשבונית - {client.get('name', '')}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Date
        ws['A3'] = f"תאריך: {datetime.now().strftime('%d/%m/%Y')}"
        
        # Headers
        headers = ['תיאור', 'תאריך', 'סכום', 'עלות לנו', 'שולם']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
        
        # Charges
        row = 6
        total = 0
        for charge in client.get('extra_charges', []):
            ws.cell(row=row, column=1, value=charge.get('description', ''))
            ws.cell(row=row, column=2, value=charge.get('date', ''))
            ws.cell(row=row, column=3, value=charge.get('amount', 0))
            ws.cell(row=row, column=4, value=charge.get('our_cost', 0))
            ws.cell(row=row, column=5, value='כן' if charge.get('paid') else 'לא')
            total += float(charge.get('amount', 0) or 0)
            row += 1
        
        # Total
        ws.cell(row=row+1, column=2, value='סה"כ:')
        ws.cell(row=row+1, column=3, value=total)
        ws.cell(row=row+1, column=2).font = header_font
        ws.cell(row=row+1, column=3).font = header_font
        
        # Adjust column widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Save file
        filename = f"invoice_{client_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(STATIC_FOLDER, filename)
        wb.save(filepath)
        
        return send_file(filepath, as_attachment=True, download_name=filename)
    except Exception as e:
        print(f"Error in generate_invoice: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@finance_bp.route('/export_open_charges')
@login_required
def export_open_charges():
    """Export all open charges to Excel"""
    try:
        data = load_data()
        user_role = get_user_role(current_user.id)
        
        # Filter clients
        if is_manager_or_admin(current_user.id, user_role):
            clients = filter_active_clients(data)
        else:
            clients = [c for c in filter_active_clients(data) 
                      if can_user_access_client(current_user.id, user_role, c)]
        
        # Create Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "חיובים פתוחים"
        ws.sheet_view.rightToLeft = True
        
        # Headers
        headers = ['לקוח', 'תיאור', 'תאריך', 'סכום', 'עלות לנו']
        header_font = Font(bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
        
        # Data
        row = 2
        total = 0
        for client in clients:
            for charge in client.get('extra_charges', []):
                if not charge.get('paid'):
                    ws.cell(row=row, column=1, value=client.get('name', ''))
                    ws.cell(row=row, column=2, value=charge.get('description', ''))
                    ws.cell(row=row, column=3, value=charge.get('date', ''))
                    ws.cell(row=row, column=4, value=charge.get('amount', 0))
                    ws.cell(row=row, column=5, value=charge.get('our_cost', 0))
                    total += float(charge.get('amount', 0) or 0)
                    row += 1
        
        # Total row
        ws.cell(row=row+1, column=3, value='סה"כ:')
        ws.cell(row=row+1, column=4, value=total)
        ws.cell(row=row+1, column=3).font = header_font
        ws.cell(row=row+1, column=4).font = header_font
        
        # Column widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Save and send
        filename = f"open_charges_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(STATIC_FOLDER, filename)
        wb.save(filepath)
        
        return send_file(filepath, as_attachment=True, download_name=filename)
    except Exception as e:
        print(f"Error in export_open_charges: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
